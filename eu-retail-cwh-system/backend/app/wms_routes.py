import os
import sqlite3
from datetime import datetime
from flask import Blueprint, request, jsonify, current_app, send_file, g
from werkzeug.utils import secure_filename
import openpyxl
import io
import re
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

inventory_bp = Blueprint('inventory', __name__)
delivery_bp = Blueprint('delivery', __name__)

from . import db as db_module

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
CALIBRATION_HEADERS = [
    'Bom Code',
    'Product Number',
    'Product Description',
    'Category',
    'Category 2',
    'Instruction',
    'Inventory (Pcs)',
    'Unit Price (USD)',
    'TTL Amount (USD)',
    '最近入库时间 Last Inbound date',
    '最近出库时间 Last Outbound date',
    '库龄 Dos (Days of Supply)',
    '库存静置时长 Days without Stock',
    '备注 comment',
    'SKU条目创建时间',
    '图片 Photo',
]
CALIBRATION_SHEETS = {
    'Lodz': 'Lodz warehouse',
    'Bydgoszcz': 'Bydgoszcz warehouse',
}
SUMMARY_ROW_MARKERS = {'TOTAL', 'GRAND TOTAL', 'SUMA KONCOWA', 'SUMA KOŃCOWA'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def uploaded_file_size(file):
    pos = file.stream.tell()
    file.stream.seek(0, os.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(pos)
    return size

def save_uploaded_source_file(file, prefix):
    upload_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = secure_filename(file.filename or 'upload.xlsx') or 'upload.xlsx'
    filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    path = os.path.join(upload_dir, filename)
    file.save(path)
    try:
        file.stream.seek(0)
    except Exception:
        pass
    return filename, path

def remove_uploaded_source_file(path):
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except OSError:
        pass

def get_db():
    return db_module.get_wms_db()

def query_db(query, args=(), one=False):
    d = db_module.get_wms_db()
    cur = d.execute(query, args)
    rows = cur.fetchall()
    cur.close()
    return (rows[0] if rows else None) if one else rows

def resolve_product_identity(d, product_number='', bom_code=''):
    product_number = (product_number or '').strip()
    bom_code = (bom_code or '').strip()
    if product_number and bom_code:
        return product_number, bom_code
    if product_number:
        row = d.execute("""
            SELECT bom_code FROM warehouse_inventory
            WHERE product_number = ?
            ORDER BY updated_at DESC LIMIT 1
        """, (product_number,)).fetchone()
        if row:
            return product_number, row['bom_code']
        row = d.execute("""
            SELECT bom_code FROM product_master
            WHERE product_number = ?
            ORDER BY updated_at DESC LIMIT 1
        """, (product_number,)).fetchone()
        if row:
            return product_number, row['bom_code']
        row = d.execute("""
            SELECT bom_code FROM sku_master
            WHERE product_number = ?
            ORDER BY updated_at DESC LIMIT 1
        """, (product_number,)).fetchone()
        return product_number, row['bom_code'] if row else product_number
    if bom_code:
        row = d.execute("SELECT product_number FROM product_master WHERE bom_code = ? ORDER BY updated_at DESC LIMIT 1", (bom_code,)).fetchone()
        if not row:
            row = d.execute("SELECT product_number FROM sku_master WHERE bom_code = ?", (bom_code,)).fetchone()
        return (row['product_number'] if row and row['product_number'] else bom_code), bom_code
    return '', ''

def clean_text(value):
    return str(value or '').strip()

def clean_upper(value):
    return clean_text(value).upper()

def split_filter_values(value):
    return [normalize_dimension(v, 'category2') for v in re.split(r'[,;，；]+', str(value or '')) if v.strip()]

def split_dimension_filter(value, kind='generic'):
    return [normalize_dimension(v, kind) for v in re.split(r'[,;，；]+', str(value or '')) if v.strip()]

def safe_export_part(value, fallback='all'):
    text = re.sub(r'[^A-Za-z0-9\u4e00-\u9fff_-]+', '_', str(value or '').strip()).strip('_')
    return (text[:40] or fallback)

def maintenance_password_ok(password, allow_admin=True):
    if allow_admin and bool(getattr(g, 'is_admin', False)):
        return True
    return (password or '') == 'Q84405995'

def normalize_dimension(value, kind='generic'):
    text = re.sub(r'\s+', ' ', clean_upper(value)).strip()
    compact = re.sub(r'[\s/_-]+', '', text)
    if kind == 'category2':
        fixes = {
            'ADUIO': 'AUDIO',
            'AUDIO': 'AUDIO',
            'IOT': 'IOT',
            'OTHER': 'OTHER',
            'PHONE': 'PHONE',
            'TABLET': 'TABLET',
            'WEARABLE': 'WEARABLE',
        }
        return fixes.get(compact, text)
    if kind == 'category':
        fixes = {
            'ACRYLICPROP': 'ACRYLIC PROP',
            'DUMMY': 'DUMMY',
            'FURNITURE': 'FURNITURE',
            'GIFT': 'GIFT',
            'PROP': 'PROP',
            'SECURITYSYSTEM': 'SECURITY SYSTEM',
            'TOOLS/ACCESSORIES': 'TOOLS/ACCESSORIES',
            'TOOLSACCESSORIES': 'TOOLS/ACCESSORIES',
            'TOOLACCESSORIES': 'TOOLS/ACCESSORIES',
            'TRAININGMATERIALS': 'TRAINING MATERIALS',
            'UNIFORM': 'UNIFORM',
            'WOODENOVERLAY': 'WOODEN OVERLAY',
        }
        return fixes.get(compact, text)
    if kind == 'instruction':
        return compact if compact else text
    return text

def is_local_bom_code(value):
    text = clean_upper(value)
    compact = re.sub(r'[\s/_-]+', '', text)
    return text in {'/', '／'} or compact in {'本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', 'NOBOM', 'NOBOMCODE'}

LOCAL_BOM_SQL_EXPR = "REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, ''))), ' ', ''), '_', ''), '-', '')"
LOCAL_BOM_SQL_VALUES = "('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')"
LOCAL_WI_BOM_SQL = f"({LOCAL_BOM_SQL_EXPR} IN {LOCAL_BOM_SQL_VALUES} OR COALESCE(NULLIF(wi.bom_code, ''), '') = '')"

def clean_code(value):
    text = clean_upper(value)
    if re.fullmatch(r'.+\.0+', text):
        text = re.sub(r'\.0+$', '', text)
    return text

def code_match_expr(column):
    return f"REPLACE(REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE({column}, ''))), ' ', ''), '_', ''), '-', ''), '.0', '')"

def code_match_value(value):
    return re.sub(r'[\s_-]+', '', clean_code(value)).replace('.0', '')

def parse_operation_id(value):
    raw = clean_text(value)
    if not raw:
        return '', '', ''
    parts = [p for p in raw.split('_') if p]
    prefix = parts[0].upper() if parts else ''
    order_date = ''
    target_parts = parts[1:]
    if parts and re.fullmatch(r'\d{8}', parts[-1]):
        ddmmyyyy = parts[-1]
        order_date = f"{ddmmyyyy[4:8]}-{ddmmyyyy[2:4]}-{ddmmyyyy[0:2]}"
        target_parts = parts[1:-1]
    target = ' '.join(target_parts).strip()
    return prefix, target, order_date

def header_key(value):
    return clean_text(value).lower().replace(' ', '').replace('_', '')

def find_header_row(ws, required_headers, max_scan_rows=30, aliases=None):
    aliases = aliases or {}
    required = {header_key(h) for h in required_headers}
    alias_to_canonical = {}
    for canonical, names in aliases.items():
        canonical_key = header_key(canonical)
        for name in [canonical] + list(names):
            alias_to_canonical[header_key(name)] = canonical_key
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        col_map = {}
        for idx, value in enumerate(values):
            if not value:
                continue
            key = header_key(value)
            col_map[key] = idx + 1
            if key in alias_to_canonical:
                col_map[alias_to_canonical[key]] = idx + 1
        if required.issubset(set(col_map)):
            return row_idx, col_map
    return None, {}

def get_cell(ws, row_idx, col_map, header_name):
    col = col_map.get(header_key(header_name))
    return ws.cell(row_idx, col).value if col else None

def to_int(value, default=0):
    try:
        if value is None or str(value).strip() == '':
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default

def to_int_strict(value):
    if value is None or str(value).strip() == '':
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None

def to_float(value, default=0):
    try:
        if value is None or str(value).strip() == '':
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def date_text(value):
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return clean_text(value)[:10]

def normalized_header(value):
    return re.sub(r'[^A-Z0-9]+', '', clean_text(value).upper())

def is_summary_row_value(value):
    text = clean_upper(value)
    normalized = normalized_header(text)
    return text in SUMMARY_ROW_MARKERS or normalized in {normalized_header(v) for v in SUMMARY_ROW_MARKERS}

def get_row_value_by_headers(row, header_map, header_name):
    idx = header_map.get(normalized_header(header_name))
    if idx is None or idx >= len(row):
        return None
    return row[idx]

def parse_date_text(value):
    if value is None or clean_text(value) == '':
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    text = clean_text(value)
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return text[:10]

def find_exact_header_map(ws, headers, max_scan_rows=10):
    required = [normalized_header(h) for h in headers]
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        header_map = {}
        for idx, value in enumerate(values):
            key = normalized_header(value)
            if key:
                header_map[key] = idx
        if all(key in header_map for key in required[:8]):
            return row_idx, header_map
    return None, {}

def upsert_inventory_calibration_row(d, warehouse, data, report):
    bom_code = clean_code(data.get('Bom Code'))
    product_number = clean_code(data.get('Product Number'))
    if not product_number:
        report['skipped'] += 1
        return
    if is_summary_row_value(product_number) or is_summary_row_value(bom_code):
        report['skipped'] += 1
        return
    product_description = clean_text(data.get('Product Description'))
    category = normalize_dimension(data.get('Category'), 'category')
    category_2 = normalize_dimension(data.get('Category 2'), 'category2')
    instruction = normalize_dimension(data.get('Instruction'), 'instruction')
    inventory = to_int(data.get('Inventory (Pcs)'), 0)
    unit_price = to_float(data.get('Unit Price (USD)'), 0)
    ttl_amount = to_float(data.get('TTL Amount (USD)'), None)
    if ttl_amount is None:
        ttl_amount = inventory * unit_price
    last_inbound = parse_date_text(data.get('最近入库时间 Last Inbound date'))
    last_outbound = parse_date_text(data.get('最近出库时间 Last Outbound date'))
    comment = clean_text(data.get('备注 comment'))
    sku_created_at = parse_date_text(data.get('SKU条目创建时间'))
    if not last_inbound and sku_created_at:
        last_inbound = sku_created_at
    if not bom_code:
        bom_code = product_number

    existing_product = d.execute("SELECT 1 FROM product_master WHERE product_number = ?", (product_number,)).fetchone()
    existing_inv = d.execute("SELECT inventory FROM warehouse_inventory WHERE warehouse = ? AND product_number = ?",
                             (warehouse, product_number)).fetchone()
    old_qty = existing_inv['inventory'] if existing_inv else None

    d.execute("""INSERT INTO product_master
                    (product_number, bom_code, product_description, category, category_2, unit_price,
                     last_inbound_date, last_outbound_date, remark, created_at, updated_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE(NULLIF(?, ''), CURRENT_TIMESTAMP), CURRENT_TIMESTAMP)
                 ON CONFLICT(product_number) DO UPDATE SET
                    bom_code=excluded.bom_code,
                    product_description=excluded.product_description,
                    category=excluded.category,
                    category_2=excluded.category_2,
                    unit_price=CASE
                        WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                        ELSE product_master.unit_price
                    END,
                    last_inbound_date=COALESCE(excluded.last_inbound_date, product_master.last_inbound_date),
                    last_outbound_date=COALESCE(excluded.last_outbound_date, product_master.last_outbound_date),
                    remark=excluded.remark,
                    updated_at=CURRENT_TIMESTAMP""",
              (product_number, bom_code, product_description, category, category_2, unit_price,
               last_inbound or None, last_outbound or None, comment, sku_created_at or None))
    if not is_local_bom_code(bom_code):
        d.execute("""INSERT INTO sku_master
                        (bom_code, product_number, product_description, category, category_2, unit_price, updated_at)
                     VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                     ON CONFLICT(bom_code) DO UPDATE SET
                        product_number=excluded.product_number,
                        product_description=excluded.product_description,
                        category=excluded.category,
                        category_2=excluded.category_2,
                        unit_price=CASE
                            WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                            ELSE sku_master.unit_price
                        END,
                        updated_at=CURRENT_TIMESTAMP""",
                  (bom_code, product_number, product_description, category, category_2, unit_price))
    d.execute("""INSERT INTO warehouse_inventory
                    (warehouse, bom_code, product_number, instruction, inventory, last_inbound_date, last_outbound_date, comment, updated_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                 ON CONFLICT(warehouse, product_number) DO UPDATE SET
                    bom_code=excluded.bom_code,
                    instruction=excluded.instruction,
                    inventory=excluded.inventory,
                    last_inbound_date=COALESCE(excluded.last_inbound_date, warehouse_inventory.last_inbound_date),
                    last_outbound_date=COALESCE(excluded.last_outbound_date, warehouse_inventory.last_outbound_date),
                    comment=excluded.comment,
                    updated_at=CURRENT_TIMESTAMP""",
              (warehouse, bom_code, product_number, instruction, inventory, last_inbound or None, last_outbound or None, comment))
    if old_qty != inventory:
        report['inventory_changed'] += 1
        if old_qty is not None:
            d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, comment)
                         VALUES (?, ?, ?, 'calibration', ?, ?)""",
                      (warehouse, bom_code, product_number, inventory - old_qty, f"库存校准 Excel: {old_qty} → {inventory}"))
    if existing_product:
        report['updated'] += 1
    else:
        report['created'] += 1
    report['imported'] += 1

def code_key(value):
    return re.sub(r'[^A-Z0-9]+', '', clean_code(value))

def image_code_candidates(filename_or_code):
    stem = os.path.splitext(os.path.basename(clean_text(filename_or_code)))[0].strip()
    stem = re.sub(r'\s*\(\d+\)$', '', stem).strip()
    candidates = [stem]
    if '_' in stem:
        candidates.append(stem.split('_', 1)[1].strip())
    if '-' in stem:
        candidates.append(stem.split('-', 1)[1].strip())
    seen = set()
    result = []
    for c in candidates:
        key = code_key(c)
        if c and key and key not in seen:
            result.append(clean_code(c))
            seen.add(key)
    return result

def is_summary_product_number(product_number):
    text = clean_upper(product_number)
    summary_tokens = ('SUMA', 'TOTAL', '汇总', 'KONCOWA', 'KOŃCOWA', 'GRANDTOTAL')
    compact = header_key(text)
    return any(token in text for token in summary_tokens) or 'grandtotal' in compact

def find_existing_image_filename(code):
    target = code_key(code)
    if not target:
        return None
    folder = current_app.config['UPLOAD_FOLDER']
    if not os.path.exists(folder):
        return None
    for name in os.listdir(folder):
        if not allowed_file(name):
            continue
        stem_key = code_key(os.path.splitext(name)[0])
        if stem_key == target or stem_key.endswith(target):
            return name
    return None

def clear_image_folder(folder):
    deleted = 0
    failed = []
    if not folder or not os.path.exists(folder):
        return deleted, failed
    base = os.path.abspath(folder)
    for root, dirs, files in os.walk(base, topdown=False):
        for name in files:
            path = os.path.abspath(os.path.join(root, name))
            if not path.startswith(base + os.sep):
                continue
            try:
                os.remove(path)
                deleted += 1
            except OSError:
                failed.append(name)
        for name in dirs:
            path = os.path.abspath(os.path.join(root, name))
            if not path.startswith(base + os.sep):
                continue
            try:
                os.rmdir(path)
            except OSError:
                pass
    return deleted, failed

def image_key_for_sku(bom_code, product_number):
    bom_code = clean_code(bom_code)
    product_number = clean_code(product_number)
    if product_number and (not bom_code or is_local_bom_code(bom_code)):
        return product_number
    return bom_code or product_number

def repair_image_mappings(d):
    rows = d.execute("""
        SELECT bom_code, product_number
        FROM product_master
        UNION
        SELECT bom_code, product_number
        FROM warehouse_inventory
    """).fetchall()
    repaired = 0
    for row in rows:
        bom_code = clean_code(row['bom_code'])
        product_number = clean_code(row['product_number'])
        match_code = image_key_for_sku(bom_code, product_number)
        if not match_code:
            continue
        filename = find_existing_image_filename(match_code)
        if not filename:
            continue
        image_bom_key = image_key_for_sku(bom_code, product_number)
        d.execute("""INSERT INTO sku_master (bom_code, product_number, image_path, updated_at)
                     VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                     ON CONFLICT(bom_code) DO UPDATE SET
                         product_number=COALESCE(NULLIF(excluded.product_number, ''), product_number),
                         image_path=excluded.image_path,
                         updated_at=CURRENT_TIMESTAMP""",
                  (image_bom_key, product_number, filename))
        repaired += 1
    return repaired

def resolve_image_bom(d, filename_or_code):
    candidates = image_code_candidates(filename_or_code)
    for candidate in candidates:
        row = d.execute("""
            SELECT product_number, bom_code FROM product_master WHERE UPPER(TRIM(product_number)) = UPPER(TRIM(?))
            UNION
            SELECT product_number, bom_code FROM warehouse_inventory WHERE UPPER(TRIM(product_number)) = UPPER(TRIM(?))
            LIMIT 1
        """, (candidate, candidate)).fetchone()
        if row:
            return image_key_for_sku(row['bom_code'], row['product_number'])
        row = d.execute("""
            SELECT bom_code FROM sku_master WHERE UPPER(TRIM(bom_code)) = UPPER(TRIM(?))
            UNION
            SELECT bom_code FROM product_master WHERE UPPER(TRIM(bom_code)) = UPPER(TRIM(?))
            UNION
            SELECT bom_code FROM product_master WHERE UPPER(TRIM(product_number)) = UPPER(TRIM(?))
            UNION
            SELECT bom_code FROM warehouse_inventory WHERE UPPER(TRIM(product_number)) = UPPER(TRIM(?))
            LIMIT 1
        """, (candidate, candidate, candidate, candidate)).fetchone()
        if row:
            return row['bom_code']
        key = code_key(candidate)
        row = d.execute("""
            SELECT product_number, bom_code FROM product_master WHERE {expr_product} = ?
            UNION
            SELECT product_number, bom_code FROM warehouse_inventory WHERE {expr_inv} = ?
            LIMIT 1
        """.format(expr_product=code_match_expr('product_number'), expr_inv=code_match_expr('product_number')),
            (key, key)).fetchone()
        if row:
            return image_key_for_sku(row['bom_code'], row['product_number'])
        row = d.execute("""
            SELECT bom_code FROM sku_master WHERE REPLACE(REPLACE(REPLACE(UPPER(TRIM(bom_code)), ' ', ''), '_', ''), '-', '') = ?
            UNION
            SELECT bom_code FROM product_master WHERE REPLACE(REPLACE(REPLACE(UPPER(TRIM(bom_code)), ' ', ''), '_', ''), '-', '') = ?
            LIMIT 1
        """, (key, key)).fetchone()
        if row:
            return row['bom_code']
        rows = d.execute("""
            SELECT bom_code FROM sku_master
            UNION
            SELECT bom_code FROM product_master
            UNION
            SELECT bom_code FROM warehouse_inventory
        """).fetchall()
        for row in rows:
            if code_key(row['bom_code']) == key:
                return row['bom_code']
    return None

# ==================== SKU 管理 ====================

@inventory_bp.route('/api/skus', methods=['GET'])
def get_skus():
    d = get_db()
    keyword = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    category2 = request.args.get('category2', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    where_clauses = []
    params = []
    if keyword:
        where_clauses.append("(pm.bom_code LIKE ? OR pm.product_number LIKE ? OR pm.product_description LIKE ?)")
        params.extend([f'%{keyword}%'] * 3)
    if category:
        where_clauses.append("UPPER(TRIM(COALESCE(pm.category, ''))) = UPPER(TRIM(?))")
        params.append(category)
    category2_values = split_filter_values(category2)
    if category2_values:
        placeholders = ','.join(['?'] * len(category2_values))
        where_clauses.append(f"UPPER(TRIM(COALESCE(pm.category_2, ''))) IN ({placeholders})")
        params.extend(category2_values)

    where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    offset = (page - 1) * per_page

    total = d.execute(f"SELECT COUNT(*) FROM product_master pm {where}", params).fetchone()[0]
    rows = d.execute(f"""
        SELECT pm.*, simg.image_path
        FROM product_master pm
        LEFT JOIN sku_master simg ON pm.bom_code = simg.bom_code
        {where}
        ORDER BY pm.updated_at DESC LIMIT ? OFFSET ?
    """,
                     params + [per_page, offset]).fetchall()
    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


@inventory_bp.route('/api/skus', methods=['POST'])
def create_sku():
    d = get_db()
    data = request.get_json()
    try:
        d.execute("""INSERT INTO sku_master (bom_code, product_number, product_description, category, category_2, unit_price)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (data['bom_code'], data.get('product_number', ''), data.get('product_description', ''),
                   normalize_dimension(data.get('category', ''), 'category'),
                   normalize_dimension(data.get('category_2', ''), 'category2'), data.get('unit_price', 0)))
        d.commit()
        return jsonify({'success': True, 'bom_code': data['bom_code']})
    except sqlite3.IntegrityError:
        return jsonify({'error': f"BOM Code '{data['bom_code']}' already exists"}), 409


def update_sku_master_by_bom(d, bom_code, data):
    fields = []
    params = []
    for key in ['product_number', 'product_description', 'category', 'category_2', 'unit_price', 'image_path']:
        if key in data:
            fields.append(f"{key} = ?")
            if key == 'category':
                params.append(normalize_dimension(data[key], 'category'))
            elif key == 'category_2':
                params.append(normalize_dimension(data[key], 'category2'))
            else:
                params.append(data[key])
    if fields:
        fields.append("updated_at = CURRENT_TIMESTAMP")
        params.append(bom_code)
        d.execute(f"UPDATE sku_master SET {', '.join(fields)} WHERE bom_code = ?", params)
    return bool(fields)


@inventory_bp.route('/api/skus/<bom_code>', methods=['PUT'])
def update_sku(bom_code):
    d = get_db()
    data = request.get_json()
    if not update_sku_master_by_bom(d, bom_code, data):
        return jsonify({'error': 'No fields to update'}), 400
    d.commit()
    return jsonify({'success': True})


@inventory_bp.route('/api/inventory/product/<path:product_number>/maintenance', methods=['PUT'])
def update_product_maintenance(product_number):
    d = get_db()
    data = request.get_json() or {}
    product_number = clean_code(product_number)
    if not product_number:
        return jsonify({'error': 'Product Number 不能为空'}), 400
    product = d.execute("SELECT * FROM product_master WHERE product_number = ?", (product_number,)).fetchone()
    if not product:
        return jsonify({'error': 'Product Number 不存在'}), 404

    fields = []
    params = []
    for key in ['bom_code', 'product_description', 'category', 'category_2', 'unit_price',
                'last_inbound_date', 'last_outbound_date', 'remark', 'dos_threshold', 'idle_threshold']:
        if key in data:
            fields.append(f"{key} = ?")
            if key == 'category':
                params.append(normalize_dimension(data[key], 'category'))
            elif key == 'category_2':
                params.append(normalize_dimension(data[key], 'category2'))
            elif key in ('dos_threshold', 'idle_threshold'):
                params.append(to_int(data[key], 180))
            else:
                params.append(data[key])
    if not fields:
        return jsonify({'error': 'No fields to update'}), 400
    fields.append("updated_at = CURRENT_TIMESTAMP")
    params.append(product_number)
    d.execute(f"UPDATE product_master SET {', '.join(fields)} WHERE product_number = ?", params)

    bom_code = clean_code(data.get('bom_code') or product['bom_code'] or '')
    if bom_code and not is_local_bom_code(bom_code):
        d.execute("""INSERT INTO sku_master (bom_code, product_number, product_description, category, category_2, unit_price, updated_at)
                     VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                     ON CONFLICT(bom_code) DO UPDATE SET
                        product_number=excluded.product_number,
                        product_description=excluded.product_description,
                        category=excluded.category,
                        category_2=excluded.category_2,
                        unit_price=excluded.unit_price,
                        updated_at=CURRENT_TIMESTAMP""",
                  (bom_code, product_number, data.get('product_description', product['product_description'] or ''),
                   normalize_dimension(data.get('category', product['category'] or ''), 'category'),
                   normalize_dimension(data.get('category_2', product['category_2'] or ''), 'category2'),
                   to_float(data.get('unit_price', product['unit_price'] or 0))))
        d.execute("""UPDATE warehouse_inventory
                     SET bom_code = ?,
                         last_inbound_date = COALESCE(NULLIF(?, ''), last_inbound_date),
                         last_outbound_date = COALESCE(NULLIF(?, ''), last_outbound_date),
                         comment = COALESCE(NULLIF(?, ''), comment),
                         updated_at = CURRENT_TIMESTAMP
                     WHERE product_number = ?""",
                  (bom_code, data.get('last_inbound_date', ''), data.get('last_outbound_date', ''),
                   data.get('remark', ''), product_number))
    d.commit()
    return jsonify({'success': True})


@inventory_bp.route('/api/skus/<bom_code>', methods=['DELETE'])
def delete_sku(bom_code):
    d = get_db()
    inv = d.execute("SELECT COUNT(*) FROM warehouse_inventory WHERE bom_code = ?", (bom_code,)).fetchone()[0]
    if inv > 0:
        return jsonify({'error': 'Cannot delete: inventory records exist for this SKU'}), 409
    items = d.execute("SELECT COUNT(*) FROM delivery_order_items WHERE bom_code = ?", (bom_code,)).fetchone()[0]
    if items > 0:
        return jsonify({'error': 'Cannot delete: delivery order items reference this SKU'}), 409
    d.execute("DELETE FROM sku_master WHERE bom_code = ?", (bom_code,))
    d.commit()
    return jsonify({'success': True})


# ==================== 库存管理 ====================

@inventory_bp.route('/api/inventory', methods=['GET'])
def get_inventory():
    d = get_db()
    warehouse = request.args.get('warehouse')
    if warehouse is None:
        warehouse = 'Lodz warehouse'
    warehouse = warehouse.strip()
    keyword = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    category2 = request.args.get('category2', '')
    instruction = request.args.get('instruction', '')
    quality = request.args.get('quality', '').strip().lower()
    sort_by = request.args.get('sort_by', 'updated_at')
    sort_dir = 'ASC' if request.args.get('sort_dir', 'desc').lower() == 'asc' else 'DESC'
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 100))

    where_clauses = []
    params = []
    if warehouse:
        where_clauses.append("wi.warehouse = ?")
        params.append(warehouse)
    if keyword:
        where_clauses.append("(wi.bom_code LIKE ? OR wi.product_number LIKE ? OR pm.product_description LIKE ?)")
        params.extend([f'%{keyword}%'] * 3)
    category_values = split_filter_values(category)
    if category_values:
        placeholders = ','.join(['?'] * len(category_values))
        where_clauses.append(f"UPPER(TRIM(COALESCE(pm.category, ''))) IN ({placeholders})")
        params.extend(category_values)
    category2_values = split_filter_values(category2)
    if category2_values:
        placeholders = ','.join(['?'] * len(category2_values))
        where_clauses.append(f"UPPER(TRIM(COALESCE(pm.category_2, ''))) IN ({placeholders})")
        params.extend(category2_values)
    instruction_values = split_filter_values(instruction)
    if instruction_values:
        placeholders = ','.join(['?'] * len(instruction_values))
        where_clauses.append(f"UPPER(TRIM(COALESCE(wi.instruction, ''))) IN ({placeholders})")
        params.extend(instruction_values)
    if quality == 'missing_price':
        where_clauses.append("wi.inventory > 0")
        where_clauses.append("COALESCE(pm.unit_price, 0) <= 0")
    elif quality == 'missing_image':
        where_clauses.append("wi.inventory > 0")
        where_clauses.append("COALESCE(pimg.image_path, simg.image_path, '') = ''")
    elif quality == 'missing_inbound':
        where_clauses.append("wi.inventory > 0")
        where_clauses.append("COALESCE(wi.last_inbound_date, '') = ''")

    where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    offset = (page - 1) * per_page
    inventory_from = f"""
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        LEFT JOIN sku_master simg ON wi.bom_code = simg.bom_code AND NOT {LOCAL_WI_BOM_SQL}
        LEFT JOIN sku_master pimg ON {LOCAL_WI_BOM_SQL}
                                 AND wi.product_number = pimg.bom_code
    """
    sort_columns = {
        'bom_code': 'wi.bom_code',
        'product_number': 'wi.product_number',
        'product_description': 'pm.product_description',
        'category': 'UPPER(TRIM(COALESCE(pm.category, "")))',
        'category_2': 'UPPER(TRIM(COALESCE(pm.category_2, "")))',
        'instruction': 'UPPER(TRIM(COALESCE(wi.instruction, "")))',
        'inventory': 'wi.inventory',
        'unit_price': 'pm.unit_price',
        'ttl_amount': 'ttl_amount',
        'last_inbound_date': 'wi.last_inbound_date',
        'last_outbound_date': 'wi.last_outbound_date',
        'dos': 'dos',
        'days_without_stock': 'days_without_stock',
        'updated_at': 'wi.updated_at'
    }
    order_expr = sort_columns.get(sort_by, 'wi.updated_at')

    total = d.execute(f"""
        SELECT COUNT(*) FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        LEFT JOIN sku_master simg ON wi.bom_code = simg.bom_code AND NOT {LOCAL_WI_BOM_SQL}
        LEFT JOIN sku_master pimg ON {LOCAL_WI_BOM_SQL}
                                 AND wi.product_number = pimg.bom_code
        {where}
    """, params).fetchone()[0]
    last_updated_row = d.execute(f"""
        SELECT MAX(wi.updated_at) AS last_updated_at
        {inventory_from}
        {where}
    """, params).fetchone()

    rows = d.execute(f"""
        SELECT wi.id, wi.warehouse, wi.bom_code,
               wi.product_number AS product_number,
               wi.inventory, wi.last_inbound_date, wi.last_outbound_date, wi.comment, wi.updated_at,
               pm.product_description,
               UPPER(TRIM(COALESCE(pm.category, ''))) AS category,
               UPPER(TRIM(COALESCE(pm.category_2, ''))) AS category_2,
               UPPER(TRIM(COALESCE(wi.instruction, ''))) AS instruction,
               pm.unit_price, COALESCE(pimg.image_path, simg.image_path) AS image_path,
               pm.unit_price * wi.inventory AS ttl_amount,
               CAST(JULIANDAY('now') - JULIANDAY(wi.last_inbound_date) AS INTEGER) AS dos,
               CAST(JULIANDAY('now') - JULIANDAY(wi.last_outbound_date) AS INTEGER) AS days_without_stock
        {inventory_from}
        {where}
        ORDER BY {order_expr} {sort_dir}, wi.updated_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    return jsonify({
        'data': [dict(r) for r in rows], 'total': total,
        'page': page, 'per_page': per_page, 'warehouse': warehouse,
        'last_updated_at': last_updated_row['last_updated_at'] if last_updated_row else ''
    })


@inventory_bp.route('/api/inventory/<int:inv_id>', methods=['PUT'])
def update_inventory(inv_id):
    d = get_db()
    data = request.get_json()
    fields = []
    params = []
    for key in ['instruction', 'inventory', 'last_inbound_date', 'last_outbound_date', 'comment']:
        if key in data:
            fields.append(f"{key} = ?")
            params.append(data[key])
    if not fields:
        return jsonify({'error': 'No fields to update'}), 400
    fields.append("updated_at = CURRENT_TIMESTAMP")
    params.append(inv_id)
    d.execute(f"UPDATE warehouse_inventory SET {', '.join(fields)} WHERE id = ?", params)
    d.commit()
    return jsonify({'success': True})


@inventory_bp.route('/api/inventory/bulk', methods=['POST'])
def bulk_inventory():
    d = get_db()
    data = request.get_json()
    warehouse = data.get('warehouse', '')
    items = data.get('items', [])
    count = 0
    for item in items:
        product_number, bom_code = resolve_product_identity(d, item.get('product_number', ''), item.get('bom_code', ''))
        if not product_number:
            continue
        d.execute("""INSERT INTO warehouse_inventory (warehouse, bom_code, product_number, instruction, inventory, last_inbound_date, comment)
                     VALUES (?, ?, ?, ?, ?, ?, ?)
                     ON CONFLICT(warehouse, product_number) DO UPDATE SET
                     bom_code = excluded.bom_code,
                     instruction = COALESCE(excluded.instruction, instruction),
                     inventory = COALESCE(excluded.inventory, inventory),
                     last_inbound_date = COALESCE(excluded.last_inbound_date, last_inbound_date),
                     comment = COALESCE(excluded.comment, comment),
                     updated_at = CURRENT_TIMESTAMP""",
                  (warehouse, bom_code, product_number, item.get('instruction', ''),
                   item.get('inventory', 0), item.get('last_inbound_date', ''),
                   item.get('comment', '')))
        count += 1
    d.commit()
    return jsonify({'success': True, 'count': count})


@inventory_bp.route('/api/inventory/stats', methods=['GET'])
def inventory_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')

    if warehouse:
        base_where = "WHERE warehouse = ?"
        params = [warehouse]
    else:
        base_where = ""
        params = []

    total_skus = d.execute(f"SELECT COUNT(*) FROM warehouse_inventory {base_where}", params).fetchone()[0]
    total_pcs = d.execute(f"SELECT COALESCE(SUM(inventory), 0) FROM warehouse_inventory {base_where}", params).fetchone()[0]
    total_value = d.execute(f"""
        SELECT COALESCE(SUM(wi.inventory * pm.unit_price), 0)
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {base_where.replace('warehouse_inventory', 'wi')}
    """, params).fetchone()[0]
    missing_image_skus = d.execute(f"""
        SELECT COUNT(DISTINCT wi.product_number)
        FROM warehouse_inventory wi
        LEFT JOIN sku_master simg ON wi.bom_code = simg.bom_code AND NOT {LOCAL_WI_BOM_SQL}
        LEFT JOIN sku_master pimg ON {LOCAL_WI_BOM_SQL}
                                 AND wi.product_number = pimg.bom_code
        {base_where.replace('warehouse_inventory', 'wi')}
        {'AND' if warehouse else 'WHERE'} COALESCE(pimg.image_path, simg.image_path, '') = ''
    """, params).fetchone()[0]

    stock_where = ["wi.inventory <= 5", "wi.inventory > 0"]
    stock_params = []
    if warehouse:
        stock_where.append("wi.warehouse = ?")
        stock_params.append(warehouse)
    low_stock = d.execute(f"""
        SELECT wi.warehouse, wi.bom_code, wi.product_number, wi.inventory, pm.product_description
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        WHERE {' AND '.join(stock_where)}
        ORDER BY wi.inventory ASC
        LIMIT 20
    """, stock_params).fetchall()

    idle_where = ["wi.last_outbound_date IS NOT NULL", "JULIANDAY('now') - JULIANDAY(wi.last_outbound_date) > 90", "wi.inventory > 0"]
    idle_params = []
    if warehouse:
        idle_where.append("wi.warehouse = ?")
        idle_params.append(warehouse)
    long_idle = d.execute(f"""
        SELECT wi.warehouse, wi.bom_code, wi.product_number, wi.inventory, wi.last_outbound_date, pm.product_description,
               CAST(JULIANDAY('now') - JULIANDAY(wi.last_outbound_date) AS INTEGER) AS idle_days
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        WHERE {' AND '.join(idle_where)}
        ORDER BY idle_days DESC
        LIMIT 20
    """, idle_params).fetchall()

    cat_dist = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(pm.category, ''))) AS category, COUNT(*) AS cnt, SUM(wi.inventory) AS total_pcs
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {'WHERE wi.warehouse = ?' if warehouse else ''}
        GROUP BY UPPER(TRIM(COALESCE(pm.category, '')))
        ORDER BY total_pcs DESC
    """, params).fetchall()

    country_dist = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(wi.instruction, ''))) AS country, COUNT(*) AS cnt, SUM(wi.inventory) AS total_pcs
        FROM warehouse_inventory wi
        {'WHERE wi.warehouse = ?' if warehouse else ''}
        GROUP BY UPPER(TRIM(COALESCE(wi.instruction, '')))
        ORDER BY total_pcs DESC
    """, params).fetchall()

    return jsonify({
        'total_skus': total_skus,
        'total_pcs': total_pcs,
        'total_value': round(total_value, 2) if total_value else 0,
        'missing_image_skus': missing_image_skus or 0,
        'low_stock': [dict(r) for r in low_stock],
        'long_idle': [dict(r) for r in long_idle],
        'category_dist': [dict(r) for r in cat_dist],
        'country_dist': [dict(r) for r in country_dist]
    })


# ==================== 数据分析看板 ====================

@inventory_bp.route('/api/inventory/stats/category-ttl', methods=['GET'])
def category_ttl_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    where = "WHERE wi.warehouse = ?" if warehouse else ""
    params = [warehouse] if warehouse else []
    rows = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(pm.category, ''))) AS category,
               COUNT(DISTINCT wi.product_number) AS sku_count,
               SUM(wi.inventory) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {where}
        GROUP BY UPPER(TRIM(COALESCE(pm.category, '')))
        ORDER BY ttl_amount DESC
    """, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/stats/instruction-ttl', methods=['GET'])
def instruction_ttl_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    where = "WHERE wi.warehouse = ?" if warehouse else ""
    params = [warehouse] if warehouse else []
    rows = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(wi.instruction, ''))) AS instruction,
               COUNT(DISTINCT wi.product_number) AS sku_count,
               SUM(wi.inventory) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {where}
        GROUP BY UPPER(TRIM(COALESCE(wi.instruction, '')))
        ORDER BY ttl_amount DESC
    """, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/stats/category2-ttl', methods=['GET'])
def category2_ttl_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    where = "WHERE wi.warehouse = ?" if warehouse else ""
    params = [warehouse] if warehouse else []
    rows = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(pm.category_2, ''))) AS category_2,
               COUNT(DISTINCT wi.product_number) AS sku_count,
               SUM(wi.inventory) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {where}
        GROUP BY UPPER(TRIM(COALESCE(pm.category_2, '')))
        ORDER BY ttl_amount DESC
    """, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/stats/value-pcs-matrix', methods=['GET'])
def value_pcs_matrix_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    where = "WHERE wi.warehouse = ?" if warehouse else ""
    params = [warehouse] if warehouse else []
    rows = d.execute(f"""
        SELECT UPPER(TRIM(COALESCE(pm.category, 'UNCATEGORIZED'))) AS category,
               UPPER(TRIM(COALESCE(pm.category_2, 'UNSPECIFIED'))) AS category_2,
               UPPER(TRIM(COALESCE(wi.instruction, 'UNSPECIFIED'))) AS instruction,
               COUNT(DISTINCT wi.product_number) AS sku_count,
               COALESCE(SUM(wi.inventory), 0) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount,
               CASE
                   WHEN COALESCE(SUM(wi.inventory), 0) > 0
                   THEN COALESCE(SUM(wi.inventory * pm.unit_price), 0) / SUM(wi.inventory)
                   ELSE 0
               END AS avg_unit_value
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {where}
        GROUP BY UPPER(TRIM(COALESCE(pm.category, 'UNCATEGORIZED'))),
                 UPPER(TRIM(COALESCE(pm.category_2, 'UNSPECIFIED'))),
                 UPPER(TRIM(COALESCE(wi.instruction, 'UNSPECIFIED')))
        ORDER BY ttl_amount DESC, total_pcs DESC
    """, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


def inventory_aggregate(where_clauses, params):
    d = get_db()
    where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    row = d.execute(f"""
        SELECT COUNT(DISTINCT wi.product_number) AS sku_count,
               COALESCE(SUM(wi.inventory), 0) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        {where}
    """, params).fetchone()
    return {
        'sku_count': row['sku_count'] or 0,
        'total_pcs': row['total_pcs'] or 0,
        'ttl_amount': round(row['ttl_amount'] or 0, 2)
    }


@inventory_bp.route('/api/inventory/stats/health', methods=['GET'])
def inventory_health_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    base = ["wi.inventory > 0"]
    params = []
    if warehouse:
        base.append("wi.warehouse = ?")
        params.append(warehouse)

    total = inventory_aggregate(base, params)
    dos_rows = [dict(r) for r in overdue_inventory_rows('dos', warehouse)]
    idle_rows = [dict(r) for r in overdue_inventory_rows('idle', warehouse)]
    dos_total = {
        'sku_count': len(dos_rows),
        'total_pcs': sum(r.get('inventory') or 0 for r in dos_rows),
        'ttl_amount': round(sum(r.get('ttl_amount') or 0 for r in dos_rows), 2),
        'avg_overdue_days': round(sum(r.get('overdue_days') or 0 for r in dos_rows) / len(dos_rows), 1) if dos_rows else 0
    }
    idle_total = {
        'sku_count': len(idle_rows),
        'total_pcs': sum(r.get('inventory') or 0 for r in idle_rows),
        'ttl_amount': round(sum(r.get('ttl_amount') or 0 for r in idle_rows), 2),
        'avg_overdue_days': round(sum(r.get('overdue_days') or 0 for r in idle_rows) / len(idle_rows), 1) if idle_rows else 0
    }

    never_outbound = inventory_aggregate(base + ["wi.last_outbound_date IS NULL"], params)
    missing_price = inventory_aggregate(base + ["COALESCE(pm.unit_price, 0) <= 0"], params)
    missing_inbound = inventory_aggregate(base + ["wi.last_inbound_date IS NULL"], params)
    missing_image_where = "WHERE " + " AND ".join(base + ["COALESCE(pimg.image_path, simg.image_path, '') = ''"])
    missing_image = dict(d.execute(f"""
        SELECT COUNT(DISTINCT wi.product_number) AS sku_count,
               COALESCE(SUM(wi.inventory), 0) AS total_pcs,
               COALESCE(SUM(wi.inventory * pm.unit_price), 0) AS ttl_amount
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        LEFT JOIN sku_master simg ON wi.bom_code = simg.bom_code AND NOT {LOCAL_WI_BOM_SQL}
        LEFT JOIN sku_master pimg ON {LOCAL_WI_BOM_SQL}
                                 AND wi.product_number = pimg.bom_code
        {missing_image_where}
    """, params).fetchone())
    missing_image['ttl_amount'] = round(missing_image.get('ttl_amount') or 0, 2)

    def summarize_by_category(rows):
        grouped = {}
        for r in rows:
            key = r.get('category') or 'UNSPECIFIED'
            cur = grouped.setdefault(key, {'category': key, 'sku_count': 0, 'total_pcs': 0, 'ttl_amount': 0, 'max_overdue_days': 0})
            cur['sku_count'] += 1
            cur['total_pcs'] += r.get('inventory') or 0
            cur['ttl_amount'] += r.get('ttl_amount') or 0
            cur['max_overdue_days'] = max(cur['max_overdue_days'], r.get('overdue_days') or 0)
        return sorted(
            [{**v, 'ttl_amount': round(v['ttl_amount'], 2)} for v in grouped.values()],
            key=lambda x: x['ttl_amount'],
            reverse=True
        )[:8]

    return jsonify({
        'total': total,
        'dos': dos_total,
        'idle': idle_total,
        'never_outbound': never_outbound,
        'data_quality': {
            'missing_price': missing_price,
            'missing_image': missing_image,
            'missing_inbound': missing_inbound
        },
        'top_dos': sorted(dos_rows, key=lambda r: (r.get('ttl_amount') or 0, r.get('overdue_days') or 0), reverse=True)[:8],
        'top_idle': sorted(idle_rows, key=lambda r: (r.get('ttl_amount') or 0, r.get('overdue_days') or 0), reverse=True)[:8],
        'dos_by_category': summarize_by_category(dos_rows),
        'idle_by_category': summarize_by_category(idle_rows)
    })


def overdue_inventory_rows(kind, warehouse=''):
    d = get_db()
    is_idle = kind == 'idle'
    date_col = 'wi.last_outbound_date' if is_idle else 'wi.last_inbound_date'
    days_alias = 'idle_days' if is_idle else 'dos'
    threshold_expr = 'COALESCE(pm.idle_threshold, 180)' if is_idle else 'COALESCE(pm.dos_threshold, 180)'
    where_clauses = ["wi.inventory > 0", f"{date_col} IS NOT NULL", f"JULIANDAY('now') - JULIANDAY({date_col}) >= {threshold_expr}"]
    params = []
    if warehouse:
        where_clauses.append("wi.warehouse = ?")
        params.append(warehouse)
    where = "WHERE " + " AND ".join(where_clauses)
    rows = d.execute(f"""
        SELECT wi.warehouse, wi.bom_code, wi.product_number, wi.inventory,
               wi.last_inbound_date, wi.last_outbound_date,
               pm.product_description,
               COALESCE(pimg.image_path, simg.image_path) AS image_path,
               UPPER(TRIM(COALESCE(pm.category, ''))) AS category,
               UPPER(TRIM(COALESCE(pm.category_2, ''))) AS category_2,
               UPPER(TRIM(COALESCE(wi.instruction, ''))) AS instruction,
               COALESCE(pm.unit_price, 0) AS unit_price,
               COALESCE(pm.unit_price, 0) * wi.inventory AS ttl_amount,
               COALESCE(pm.dos_threshold, 180) AS dos_threshold,
               COALESCE(pm.idle_threshold, 180) AS idle_threshold,
               CAST(JULIANDAY('now') - JULIANDAY({date_col}) AS INTEGER) AS {days_alias},
               CAST(JULIANDAY('now') - JULIANDAY({date_col}) - {threshold_expr} AS INTEGER) AS overdue_days
        FROM warehouse_inventory wi
        LEFT JOIN product_master pm ON wi.product_number = pm.product_number
        LEFT JOIN sku_master simg ON wi.bom_code = simg.bom_code AND NOT {LOCAL_WI_BOM_SQL}
        LEFT JOIN sku_master pimg ON {LOCAL_WI_BOM_SQL}
                                 AND wi.product_number = pimg.bom_code
        {where}
        ORDER BY overdue_days DESC, ttl_amount DESC
    """, params).fetchall()
    return rows


@inventory_bp.route('/api/inventory/stats/dos180', methods=['GET'])
def dos180_stats():
    warehouse = request.args.get('warehouse', '')
    rows = overdue_inventory_rows('dos', warehouse)
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/stats/idle180', methods=['GET'])
def idle180_stats():
    warehouse = request.args.get('warehouse', '')
    rows = overdue_inventory_rows('idle', warehouse)
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/export/overdue', methods=['GET'])
def export_overdue_inventory():
    warehouse = request.args.get('warehouse', '')
    kind = request.args.get('kind', 'dos')
    if kind not in ('dos', 'idle'):
        kind = 'dos'
    rows = [dict(r) for r in overdue_inventory_rows(kind, warehouse)]
    title = '长期未使用库存报告' if kind == 'idle' else 'DOS 超期库存报告'
    metric_label = '未使用天数' if kind == 'idle' else '库龄天数'
    threshold_label = '静置提醒线' if kind == 'idle' else 'DOS 超期线'
    metric_key = 'idle_days' if kind == 'idle' else 'dos'
    threshold_key = 'idle_threshold' if kind == 'idle' else 'dos_threshold'

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = '统计总览'
    detail_ws = wb.create_sheet('明细清单')

    dark = '1F2937'
    accent = '2563EB' if kind == 'dos' else 'D97706'
    danger = 'DC2626'
    muted = '64748B'
    header_fill = PatternFill('solid', fgColor=dark)
    sub_fill = PatternFill('solid', fgColor='EAF1FF' if kind == 'dos' else 'FFF7ED')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_value = round(sum(float(r.get('ttl_amount') or 0) for r in rows), 2)
    total_pcs = sum(int(r.get('inventory') or 0) for r in rows)
    avg_overdue = round(sum(int(r.get('overdue_days') or 0) for r in rows) / len(rows), 1) if rows else 0
    severe_count = sum(1 for r in rows if int(r.get('overdue_days') or 0) >= 90)

    summary_ws.merge_cells('A1:H1')
    summary_ws['A1'] = title
    summary_ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
    summary_ws['A1'].fill = PatternFill('solid', fgColor=dark)
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    summary_ws.merge_cells('A2:H2')
    summary_ws['A2'] = f"仓库：{warehouse or '全部仓库'}    导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_ws['A2'].font = Font(color=muted)
    summary_ws['A2'].alignment = Alignment(horizontal='center')

    kpis = [
        ('超期 SKU', len(rows)),
        ('超期数量 PCS', total_pcs),
        ('超期货值 USD', total_value),
        ('平均超期天数', avg_overdue),
        ('90+ 超期 SKU', severe_count),
    ]
    for idx, (label, value) in enumerate(kpis, start=1):
        col = 1 + (idx - 1) * 2
        summary_ws.cell(4, col, label)
        summary_ws.cell(5, col, value)
        summary_ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        summary_ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        summary_ws.cell(4, col).fill = sub_fill
        summary_ws.cell(4, col).font = Font(bold=True, color=dark)
        summary_ws.cell(5, col).font = Font(bold=True, size=15, color=accent if idx != 5 else danger)
        summary_ws.cell(4, col).alignment = Alignment(horizontal='center')
        summary_ws.cell(5, col).alignment = Alignment(horizontal='center')
        summary_ws.cell(5, col).number_format = '$#,##0.00' if 'USD' in label else '#,##0.0' if '平均' in label else '#,##0'

    category_map = {}
    for r in rows:
        cat = r.get('category') or 'UNSPECIFIED'
        cur = category_map.setdefault(cat, {'sku': 0, 'pcs': 0, 'value': 0})
        cur['sku'] += 1
        cur['pcs'] += int(r.get('inventory') or 0)
        cur['value'] += float(r.get('ttl_amount') or 0)
    summary_headers = ['物料品类', 'SKU 数', '数量 PCS', '货值 USD', '数量占比', '货值占比']
    summary_ws.append([])
    summary_ws.append(summary_headers)
    start_row = summary_ws.max_row
    for cell in summary_ws[start_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for cat, values in sorted(category_map.items(), key=lambda kv: kv[1]['value'], reverse=True):
        row_idx = summary_ws.max_row + 1
        summary_ws.append([
            cat,
            values['sku'],
            values['pcs'],
            round(values['value'], 2),
            values['pcs'] / total_pcs if total_pcs else 0,
            values['value'] / total_value if total_value else 0,
        ])
        for cell in summary_ws[row_idx]:
            cell.border = border
        summary_ws.cell(row_idx, 4).number_format = '$#,##0.00'
        summary_ws.cell(row_idx, 5).number_format = '0.0%'
        summary_ws.cell(row_idx, 6).number_format = '0.0%'

    headers = [
        '图片', '仓库', 'BOM 编号', 'Product Number', '产品描述', '物料品类', '适用产品', '归属国家',
        '库存 PCS', '单价 USD', '总货值 USD', '最近入库', '最近出库', metric_label, threshold_label, '超期天数'
    ]
    detail_ws.append(headers)
    for cell in detail_ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    image_folder = current_app.config['UPLOAD_FOLDER']
    for r in rows:
        row_idx = detail_ws.max_row + 1
        detail_ws.append([
            '', r.get('warehouse', ''), r.get('bom_code', ''), r.get('product_number', ''),
            r.get('product_description', ''), r.get('category', ''), r.get('category_2', ''), r.get('instruction', ''),
            r.get('inventory', 0), r.get('unit_price', 0), r.get('ttl_amount', 0),
            r.get('last_inbound_date', ''), r.get('last_outbound_date', ''),
            r.get(metric_key, 0), r.get(threshold_key, 180), r.get('overdue_days', 0)
        ])
        detail_ws.row_dimensions[row_idx].height = 58
        for cell in detail_ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=cell.column in (5,))
        detail_ws.cell(row_idx, 9).number_format = '#,##0'
        detail_ws.cell(row_idx, 10).number_format = '$#,##0.00'
        detail_ws.cell(row_idx, 11).number_format = '$#,##0.00'
        detail_ws.cell(row_idx, 16).font = Font(bold=True, color=danger)
        image_path = r.get('image_path') or ''
        if image_path:
            full_path = os.path.abspath(os.path.join(image_folder, os.path.basename(image_path)))
            if os.path.exists(full_path):
                try:
                    img = ExcelImage(full_path)
                    img.width = 54
                    img.height = 54
                    detail_ws.add_image(img, f'A{row_idx}')
                except Exception:
                    detail_ws.cell(row_idx, 1, '图片读取失败')

    if rows:
        table_ref = f"A1:P{detail_ws.max_row}"
        tab = Table(displayName=f"{'Idle' if kind == 'idle' else 'Dos'}OverdueTable", ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        detail_ws.add_table(tab)
    detail_ws.freeze_panes = 'A2'
    detail_ws.auto_filter.ref = f"A1:P{max(detail_ws.max_row, 1)}"

    summary_widths = [24, 12, 14, 16, 12, 12, 12, 12, 12, 12]
    for idx, width in enumerate(summary_widths, start=1):
        summary_ws.column_dimensions[get_column_letter(idx)].width = width
    detail_widths = [10, 20, 16, 22, 46, 18, 16, 14, 12, 12, 14, 14, 14, 12, 12, 12]
    for idx, width in enumerate(detail_widths, start=1):
        detail_ws.column_dimensions[get_column_letter(idx)].width = width
    summary_ws.freeze_panes = 'A8'
    summary_ws.sheet_view.showGridLines = False
    detail_ws.sheet_view.showGridLines = False

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename_kind = 'idle' if kind == 'idle' else 'dos'
    filename_wh = (warehouse or 'all_warehouses').replace(' ', '_')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'VN22_{filename_kind}_overdue_{filename_wh}.xlsx'
    )


@inventory_bp.route('/api/inventory/movements', methods=['GET'])
def list_inventory_movements():
    movement_type = clean_text(request.args.get('movement_type', '')).lower()
    search = clean_text(request.args.get('q', ''))
    warehouse = clean_text(request.args.get('warehouse', ''))
    start_date = clean_text(request.args.get('start_date', ''))
    end_date = clean_text(request.args.get('end_date', ''))
    category_values = split_dimension_filter(request.args.get('category', ''), 'category')
    category2_values = split_dimension_filter(request.args.get('category2', ''), 'category2')
    instruction_values = split_dimension_filter(request.args.get('instruction', ''), 'instruction')
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(max(10, request.args.get('per_page', 20, type=int)), 100)
    sort_by = clean_text(request.args.get('sort_by', 'movement_date'))
    sort_dir = 'ASC' if clean_text(request.args.get('sort_dir', 'desc')).lower() == 'asc' else 'DESC'

    if movement_type not in ('', 'inbound', 'outbound'):
        movement_type = ''
    def valid_date(value):
        if not value:
            return ''
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%Y-%m-%d')
        except ValueError:
            return ''
    start_date = valid_date(start_date)
    end_date = valid_date(end_date)

    where = ["1=1"]
    params = []
    if movement_type:
        where.append("LOWER(im.movement_type) = ?")
        params.append(movement_type)
    if start_date:
        where.append("date(im.movement_date) >= date(?)")
        params.append(start_date)
    if end_date:
        where.append("date(im.movement_date) <= date(?)")
        params.append(end_date)
    if warehouse:
        where.append("wi.warehouse = ?")
        params.append(warehouse)
    if category_values:
        where.append("UPPER(TRIM(COALESCE(pm.category, ''))) IN ({})".format(','.join('?' for _ in category_values)))
        params.extend(category_values)
    if category2_values:
        where.append("UPPER(TRIM(COALESCE(pm.category_2, ''))) IN ({})".format(','.join('?' for _ in category2_values)))
        params.extend(category2_values)
    if instruction_values:
        where.append("UPPER(TRIM(COALESCE(wi.instruction, ''))) IN ({})".format(','.join('?' for _ in instruction_values)))
        params.extend(instruction_values)
    if search:
        like = f"%{search.upper()}%"
        where.append("""(
            UPPER(COALESCE(im.product_number, '')) LIKE ?
            OR UPPER(COALESCE(im.operation_id, '')) LIKE ?
            OR UPPER(COALESCE(im.operation_target, '')) LIKE ?
            OR UPPER(COALESCE(pm.product_description, '')) LIKE ?
            OR UPPER(COALESCE(pm.bom_code, wi.bom_code, '')) LIKE ?
        )""")
        params.extend([like] * 5)

    d = get_db()
    base_sql = f"""
        FROM inventory_movements im
        LEFT JOIN product_master pm ON im.product_number = pm.product_number
        LEFT JOIN warehouse_inventory wi ON im.product_number = wi.product_number
        WHERE {' AND '.join(where)}
        GROUP BY im.id
    """
    total = d.execute(f"SELECT COUNT(*) AS cnt FROM (SELECT im.id {base_sql})", params).fetchone()['cnt']
    sort_columns = {
        'movement_date': 'date(im.movement_date)',
        'movement_type': 'LOWER(im.movement_type)',
        'product_number': 'UPPER(im.product_number)',
        'product_description': 'UPPER(pm.product_description)',
        'category': 'UPPER(pm.category)',
        'category_2': 'UPPER(pm.category_2)',
        'instruction': "UPPER(GROUP_CONCAT(DISTINCT UPPER(TRIM(COALESCE(wi.instruction, '')))))",
        'warehouse': 'UPPER(GROUP_CONCAT(DISTINCT wi.warehouse))',
        'operation_id': 'UPPER(im.operation_id)',
        'operation_target': 'UPPER(im.operation_target)',
        'quantity': 'CAST(im.quantity AS REAL)',
        'current_inventory': 'SUM(COALESCE(wi.inventory, 0))',
        'value_usd': '(COALESCE(pm.unit_price, 0) * ABS(CAST(COALESCE(im.quantity, 0) AS REAL)))',
    }
    order_expr = sort_columns.get(sort_by, 'date(im.movement_date)')
    rows = d.execute(f"""
        SELECT im.id,
               im.movement_type,
               im.sheet_name,
               im.product_number,
               im.movement_date,
               im.operation_id,
               im.operation_prefix,
               im.operation_target,
               im.quantity,
               im.source_file,
               COALESCE(pm.bom_code, wi.bom_code, '') AS bom_code,
               COALESCE(pm.product_description, '') AS product_description,
               COALESCE(pm.category, '') AS category,
               COALESCE(pm.category_2, '') AS category_2,
               GROUP_CONCAT(DISTINCT wi.warehouse) AS warehouses,
               GROUP_CONCAT(DISTINCT UPPER(TRIM(COALESCE(wi.instruction, '')))) AS instructions,
               SUM(COALESCE(wi.inventory, 0)) AS current_inventory,
               COALESCE(pm.unit_price, 0) AS unit_price
        {base_sql}
        ORDER BY {order_expr} {sort_dir}, im.id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()

    def movement_label(value):
        value = clean_text(value).lower()
        if value == 'inbound':
            return '入库'
        if value == 'outbound':
            return '出库'
        return value or '/'

    items = []
    for r in rows:
        item = dict(r)
        item['movement_label'] = movement_label(item.get('movement_type'))
        item['value_usd'] = round(float(item.get('unit_price') or 0) * abs(float(item.get('quantity') or 0)), 2)
        items.append(item)

    return jsonify({
        'items': items,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page if per_page else 1
    })


@inventory_bp.route('/api/inventory/export/movements', methods=['GET'])
def export_inventory_movements():
    warehouse = clean_text(request.args.get('warehouse', ''))
    movement_type = clean_text(request.args.get('movement_type', '')).lower()
    start_date = clean_text(request.args.get('start_date', ''))
    end_date = clean_text(request.args.get('end_date', ''))
    category_values = split_dimension_filter(request.args.get('category', ''), 'category')
    category2_values = split_dimension_filter(request.args.get('category2', ''), 'category2')
    instruction_values = split_dimension_filter(request.args.get('instruction', ''), 'instruction')

    if movement_type not in ('', 'inbound', 'outbound'):
        movement_type = ''

    def valid_date(value):
        if not value:
            return ''
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%Y-%m-%d')
        except ValueError:
            return ''

    start_date = valid_date(start_date)
    end_date = valid_date(end_date)

    where = ["1=1"]
    params = []
    if movement_type:
        where.append("LOWER(im.movement_type) = ?")
        params.append(movement_type)
    if start_date:
        where.append("date(im.movement_date) >= date(?)")
        params.append(start_date)
    if end_date:
        where.append("date(im.movement_date) <= date(?)")
        params.append(end_date)
    if warehouse:
        where.append("wi.warehouse = ?")
        params.append(warehouse)
    if category_values:
        where.append("UPPER(TRIM(COALESCE(pm.category, ''))) IN ({})".format(','.join('?' for _ in category_values)))
        params.extend(category_values)
    if category2_values:
        where.append("UPPER(TRIM(COALESCE(pm.category_2, ''))) IN ({})".format(','.join('?' for _ in category2_values)))
        params.extend(category2_values)
    if instruction_values:
        where.append("UPPER(TRIM(COALESCE(wi.instruction, ''))) IN ({})".format(','.join('?' for _ in instruction_values)))
        params.extend(instruction_values)

    d = get_db()
    rows = d.execute(f"""
        SELECT im.id,
               im.movement_type,
               im.sheet_name,
               im.product_number,
               im.movement_date,
               im.operation_id,
               im.operation_prefix,
               im.operation_target,
               im.order_date,
               im.quantity,
               im.uom,
               im.source_file,
               im.created_at,
               COALESCE(pm.bom_code, wi.bom_code, '') AS bom_code,
               COALESCE(pm.product_description, '') AS product_description,
               COALESCE(pm.category, '') AS category,
               COALESCE(pm.category_2, '') AS category_2,
               GROUP_CONCAT(DISTINCT wi.warehouse) AS warehouses,
               GROUP_CONCAT(DISTINCT UPPER(TRIM(COALESCE(wi.instruction, '')))) AS instructions,
               SUM(COALESCE(wi.inventory, 0)) AS current_inventory,
               COALESCE(pm.unit_price, 0) AS unit_price
        FROM inventory_movements im
        LEFT JOIN product_master pm ON im.product_number = pm.product_number
        LEFT JOIN warehouse_inventory wi ON im.product_number = wi.product_number
        WHERE {' AND '.join(where)}
        GROUP BY im.id
        ORDER BY date(im.movement_date) DESC, im.id DESC
    """, params).fetchall()
    rows = [dict(r) for r in rows]

    def order_type_label(prefix, operation_id):
        prefix_text = clean_upper(prefix)
        operation_text = clean_text(operation_id)
        if prefix_text == 'PL':
            label = 'PL订单'
        elif prefix_text == 'EU':
            label = '地区部订单'
        else:
            label = clean_text(prefix) or '/'
        return '/' if operation_text and operation_text == label else label

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = '统计总览'
    detail_ws = wb.create_sheet('出入库流水')

    dark = '1F2937'
    blue = '2563EB'
    green = '059669'
    orange = 'D97706'
    muted = '64748B'
    header_fill = PatternFill('solid', fgColor=dark)
    sub_fill = PatternFill('solid', fgColor='EAF1FF')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    inbound_qty = sum(float(r.get('quantity') or 0) for r in rows if clean_text(r.get('movement_type')).lower() == 'inbound')
    outbound_qty = sum(float(r.get('quantity') or 0) for r in rows if clean_text(r.get('movement_type')).lower() == 'outbound')
    sku_count = len({clean_upper(r.get('product_number')) for r in rows if r.get('product_number')})
    value_total = sum(float(r.get('unit_price') or 0) * abs(float(r.get('quantity') or 0)) for r in rows)
    wh_label = warehouse or '全部仓库'
    date_label = f"{start_date or '不限'} 至 {end_date or '不限'}"

    summary_ws.merge_cells('A1:H1')
    summary_ws['A1'] = '出入库流水导出'
    summary_ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
    summary_ws['A1'].fill = header_fill
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    summary_ws.merge_cells('A2:H2')
    summary_ws['A2'] = f"仓库：{wh_label}    时间段：{date_label}    导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_ws['A2'].font = Font(color=muted)
    summary_ws['A2'].alignment = Alignment(horizontal='center')

    kpis = [
        ('流水条数', len(rows)),
        ('SKU 数', sku_count),
        ('入库数量 PCS', inbound_qty),
        ('出库数量 PCS', outbound_qty),
        ('涉及货值 USD', value_total),
    ]
    for idx, (label, value) in enumerate(kpis, start=1):
        col = 1 + (idx - 1) * 2
        summary_ws.cell(4, col, label)
        summary_ws.cell(5, col, value)
        summary_ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        summary_ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        summary_ws.cell(4, col).fill = sub_fill
        summary_ws.cell(4, col).font = Font(bold=True, color=dark)
        summary_ws.cell(5, col).font = Font(bold=True, size=15, color=green if '入库' in label else orange if '出库' in label else blue)
        summary_ws.cell(4, col).alignment = Alignment(horizontal='center')
        summary_ws.cell(5, col).alignment = Alignment(horizontal='center')
        summary_ws.cell(5, col).number_format = '$#,##0.00' if 'USD' in label else '#,##0'

    type_summary = {}
    for r in rows:
        label = '入库' if clean_text(r.get('movement_type')).lower() == 'inbound' else '出库' if clean_text(r.get('movement_type')).lower() == 'outbound' else r.get('movement_type') or '/'
        cur = type_summary.setdefault(label, {'rows': 0, 'qty': 0, 'value': 0})
        qty = float(r.get('quantity') or 0)
        cur['rows'] += 1
        cur['qty'] += qty
        cur['value'] += float(r.get('unit_price') or 0) * abs(qty)

    summary_ws.append([])
    summary_ws.append(['流水类型', '条数', '数量 PCS', '涉及货值 USD', '条数占比', '货值占比'])
    summary_header_row = summary_ws.max_row
    for cell in summary_ws[summary_header_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for label, values in sorted(type_summary.items()):
        row_idx = summary_ws.max_row + 1
        summary_ws.append([
            label,
            values['rows'],
            values['qty'],
            round(values['value'], 2),
            values['rows'] / len(rows) if rows else 0,
            values['value'] / value_total if value_total else 0,
        ])
        for cell in summary_ws[row_idx]:
            cell.border = border
        summary_ws.cell(row_idx, 4).number_format = '$#,##0.00'
        summary_ws.cell(row_idx, 5).number_format = '0.0%'
        summary_ws.cell(row_idx, 6).number_format = '0.0%'

    headers = [
        'DATA RUCHU', '流水类型', 'Sheet', 'Product Number', 'Bom Code', 'Product Description',
        '物料品类', '适用产品', '归属国家', '仓库', '当前库存 PCS', 'Unit Price USD', '涉及货值 USD',
        'ID_OPERACJI', '订单类型', '对象/国家城市', '下单日期', '数量', 'UoM', 'Source File', '导入时间'
    ]
    detail_ws.append(headers)
    for cell in detail_ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for r in rows:
        qty = float(r.get('quantity') or 0)
        row_idx = detail_ws.max_row + 1
        movement_label = '入库' if clean_text(r.get('movement_type')).lower() == 'inbound' else '出库' if clean_text(r.get('movement_type')).lower() == 'outbound' else r.get('movement_type')
        detail_ws.append([
            r.get('movement_date') or '',
            movement_label,
            r.get('sheet_name') or '',
            r.get('product_number') or '',
            r.get('bom_code') or '',
            r.get('product_description') or '',
            normalize_dimension(r.get('category'), 'category') or '',
            normalize_dimension(r.get('category_2'), 'category2') or '',
            r.get('instructions') or '',
            r.get('warehouses') or '',
            r.get('current_inventory') or 0,
            r.get('unit_price') or 0,
            round(float(r.get('unit_price') or 0) * abs(qty), 2),
            r.get('operation_id') or '',
            order_type_label(r.get('operation_prefix'), r.get('operation_id')),
            r.get('operation_target') or '',
            r.get('order_date') or '',
            qty,
            r.get('uom') or '',
            r.get('source_file') or '',
            r.get('created_at') or '',
        ])
        for cell in detail_ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=cell.column in (6, 14, 20))
        detail_ws.cell(row_idx, 11).number_format = '#,##0'
        detail_ws.cell(row_idx, 12).number_format = '$#,##0.00'
        detail_ws.cell(row_idx, 13).number_format = '$#,##0.00'
        detail_ws.cell(row_idx, 18).number_format = '#,##0'

    if rows:
        table_ref = f"A1:U{detail_ws.max_row}"
        tab = Table(displayName='InventoryMovementTable', ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        detail_ws.add_table(tab)
    detail_ws.freeze_panes = 'A2'
    detail_ws.auto_filter.ref = f"A1:U{max(detail_ws.max_row, 1)}"
    detail_ws.sheet_view.showGridLines = False
    summary_ws.sheet_view.showGridLines = False

    for idx, width in enumerate([16, 12, 18, 22, 16, 42, 16, 14, 18, 24, 14, 14, 14, 28, 14, 20, 14, 12, 10, 26, 20], start=1):
        detail_ws.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate([18, 14, 14, 16, 12, 12, 12, 12, 12, 12], start=1):
        summary_ws.column_dimensions[get_column_letter(idx)].width = width
    summary_ws.freeze_panes = 'A8'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"VN57_出入库流水_{safe_export_part(wh_label)}_{safe_export_part(start_date or 'all')}-{safe_export_part(end_date or 'all')}_{ts}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@inventory_bp.route('/api/inventory/settings/thresholds', methods=['GET'])
def get_inventory_threshold_settings():
    d = get_db()
    rows = d.execute("""
        SELECT UPPER(TRIM(COALESCE(pm.category, 'UNCATEGORIZED'))) AS category,
               COUNT(DISTINCT pm.product_number) AS sku_count,
               COALESCE(SUM(COALESCE(wi.inventory, 0)), 0) AS inventory_pcs,
               ROUND(COALESCE(SUM(COALESCE(wi.inventory, 0) * COALESCE(pm.unit_price, 0)), 0), 2) AS ttl_amount,
               ROUND(
                   CASE
                       WHEN COALESCE(SUM(COALESCE(wi.inventory, 0)), 0) > 0
                       THEN COALESCE(SUM(COALESCE(wi.inventory, 0) * COALESCE(pm.unit_price, 0)), 0) / SUM(COALESCE(wi.inventory, 0))
                       ELSE COALESCE(AVG(NULLIF(pm.unit_price, 0)), 0)
                   END,
                   2
               ) AS avg_unit_price,
               MIN(COALESCE(pm.dos_threshold, 180)) AS dos_threshold,
               MIN(COALESCE(pm.idle_threshold, 180)) AS idle_threshold
        FROM product_master pm
        LEFT JOIN warehouse_inventory wi ON wi.product_number = pm.product_number
        GROUP BY UPPER(TRIM(COALESCE(pm.category, 'UNCATEGORIZED')))
        ORDER BY category
    """).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@inventory_bp.route('/api/inventory/settings/thresholds', methods=['POST'])
def save_inventory_threshold_settings():
    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows') or []
    if not isinstance(rows, list):
        return jsonify({'error': 'rows 必须是数组'}), 400

    d = get_db()
    updated = 0
    for row in rows:
        category = clean_upper(row.get('category'))
        if not category:
            continue
        try:
            dos_threshold = max(1, int(row.get('dos_threshold') or 180))
            idle_threshold = max(1, int(row.get('idle_threshold') or 180))
        except (TypeError, ValueError):
            return jsonify({'error': f'{category} 的超期线必须是数字'}), 400
        cur = d.execute("""
            UPDATE product_master
            SET dos_threshold = ?, idle_threshold = ?, updated_at = CURRENT_TIMESTAMP
            WHERE UPPER(TRIM(COALESCE(category, 'UNCATEGORIZED'))) = ?
        """, (dos_threshold, idle_threshold, category))
        updated += cur.rowcount

    d.commit()
    return jsonify({'success': True, 'updated': updated})


@inventory_bp.route('/api/inventory/stats/scrap', methods=['GET'])
def scrap_stats():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    where_clauses = ["t.change_type = 'scrap'"]
    params = []
    if warehouse:
        where_clauses.append("t.warehouse = ?")
        params.append(warehouse)
    where = "WHERE " + " AND ".join(where_clauses)

    total_records = d.execute(f"SELECT COUNT(*) FROM inventory_transactions t {where}", params).fetchone()[0]
    total_qty = d.execute(f"SELECT COALESCE(SUM(ABS(t.quantity_change)), 0) FROM inventory_transactions t {where}", params).fetchone()[0]
    total_value = d.execute(f"""
        SELECT COALESCE(SUM(ABS(t.quantity_change) * sm.unit_price), 0)
        FROM inventory_transactions t
        LEFT JOIN sku_master sm ON t.bom_code = sm.bom_code
        {where}
    """, params).fetchone()[0]

    by_cat = d.execute(f"""
        SELECT sm.category, COUNT(*) AS cnt, SUM(ABS(t.quantity_change)) AS total_qty,
               COALESCE(SUM(ABS(t.quantity_change) * sm.unit_price), 0) AS total_value
        FROM inventory_transactions t
        LEFT JOIN sku_master sm ON t.bom_code = sm.bom_code
        {where}
        GROUP BY sm.category ORDER BY total_value DESC
    """, params).fetchall()

    by_month = d.execute(f"""
        SELECT strftime('%Y-%m', t.created_at) AS month, COUNT(*) AS cnt,
               SUM(ABS(t.quantity_change)) AS total_qty,
               COALESCE(SUM(ABS(t.quantity_change) * sm.unit_price), 0) AS total_value
        FROM inventory_transactions t
        LEFT JOIN sku_master sm ON t.bom_code = sm.bom_code
        {where}
        GROUP BY month ORDER BY month DESC LIMIT 12
    """, params).fetchall()

    return jsonify({
        'total_records': total_records,
        'total_qty': total_qty,
        'total_value': round(total_value, 2) if total_value else 0,
        'by_category': [dict(r) for r in by_cat],
        'by_month': [dict(r) for r in by_month]
    })


@inventory_bp.route('/api/inventory/upload-list', methods=['POST'])
def upload_inventory_list():
    return jsonify({
        'error': '旧 Stock List 导入已停用。请使用“供应商库存 + 维护表”导入，以 Product Number 作为库存主键。'
    }), 410

    d = get_db()
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传 .xlsx 文件'}), 400
    if uploaded_file_size(f) > 20 * 1024 * 1024:
        return jsonify({'error': 'Excel 文件不能超过 20MB'}), 400

    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'inventory_upload')
    wb = openpyxl.load_workbook(f, data_only=True)
    report = {'sheets': {}, 'total_updated': 0, 'total_created': 0, 'errors': []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sn = sheet_name.strip().lower()
        if 'lodz' in sn:
            warehouse = 'Lodz warehouse'
        elif 'bydgoszcz' in sn:
            warehouse = 'Bydgoszcz warehouse'
        else:
            report['sheets'][sheet_name] = {'skipped': True, 'reason': '无法识别仓库名'}
            continue

        sheet_report = {'warehouse': warehouse, 'rows': 0, 'updated': 0, 'created_sku': 0, 'updated_inv': 0, 'errors': []}

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not any(v not in (None, '') for v in row):
                continue

            product_number = clean_code(row[1]) if row[1] else ''
            if not product_number:
                sheet_report['errors'].append(f"第 {row_idx} 行缺少 Product Number，已跳过")
                continue
            bom_code = clean_code(row[0]) if row[0] else '/'
            if not bom_code or is_local_bom_code(bom_code):
                bom_code = '/'

            sheet_report['rows'] += 1
            product_description = str(row[2]).strip() if row[2] else ''
            category = normalize_dimension(row[3], 'category')
            category_2 = normalize_dimension(row[4], 'category2')
            instruction = normalize_dimension(row[5], 'instruction')
            inventory = int(row[6]) if row[6] is not None and str(row[6]).strip() != '' else 0
            unit_price = float(row[7]) if row[7] is not None and str(row[7]).strip() != '' else 0
            last_inbound = str(row[9]).strip() if row[9] else ''
            last_outbound = str(row[10]).strip() if row[10] else ''
            comment = str(row[13]).strip() if len(row) > 13 and row[13] else ''

            if not is_local_bom_code(bom_code):
                existing = d.execute("SELECT bom_code FROM sku_master WHERE bom_code = ?", (bom_code,)).fetchone()
                if existing:
                    d.execute("""UPDATE sku_master SET product_number=?, product_description=?, category=?, category_2=?,
                                 unit_price=CASE WHEN COALESCE(?, 0) > 0 THEN ? ELSE unit_price END,
                                 updated_at=CURRENT_TIMESTAMP WHERE bom_code=?""",
                              (product_number, product_description, category, category_2, unit_price, unit_price, bom_code))
                else:
                    d.execute("""INSERT INTO sku_master (bom_code, product_number, product_description, category, category_2, unit_price)
                                 VALUES (?, ?, ?, ?, ?, ?)""",
                              (bom_code, product_number, product_description, category, category_2, unit_price))
                    sheet_report['created_sku'] += 1
            d.execute("""INSERT INTO product_master (product_number, bom_code, product_description, category, category_2, unit_price, updated_at)
                         VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                         ON CONFLICT(product_number) DO UPDATE SET
                             bom_code=excluded.bom_code,
                             product_description=excluded.product_description,
                             category=excluded.category,
                             category_2=excluded.category_2,
                             unit_price=CASE
                                WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                                ELSE product_master.unit_price
                             END,
                             updated_at=CURRENT_TIMESTAMP""",
                      (product_number, bom_code, product_description, category, category_2, unit_price))

            inv_existing = d.execute("SELECT inventory FROM warehouse_inventory WHERE warehouse=? AND product_number=?",
                                     (warehouse, product_number)).fetchone()
            old_qty = inv_existing['inventory'] if inv_existing else None
            d.execute("""INSERT INTO warehouse_inventory (warehouse, bom_code, product_number, instruction, inventory, last_inbound_date, last_outbound_date, comment)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                         ON CONFLICT(warehouse, product_number) DO UPDATE SET
                         bom_code=excluded.bom_code, instruction=excluded.instruction, inventory=excluded.inventory,
                         last_inbound_date=COALESCE(NULLIF(excluded.last_inbound_date,''), last_inbound_date),
                         last_outbound_date=COALESCE(NULLIF(excluded.last_outbound_date,''), last_outbound_date),
                         comment=excluded.comment, updated_at=CURRENT_TIMESTAMP""",
                      (warehouse, bom_code, product_number, instruction, inventory, last_inbound or None, last_outbound or None, comment))

            if old_qty != inventory:
                sheet_report['updated_inv'] += 1
                if old_qty is not None:
                    change = inventory - old_qty
                    d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, comment)
                                 VALUES (?, ?, ?, 'calibration', ?, ?)""",
                              (warehouse, bom_code, product_number, change, f"库存校准: {old_qty} → {inventory}"))

        d.commit()
        report['sheets'][sheet_name] = sheet_report
        report['total_updated'] += sheet_report['updated_inv']
        report['total_created'] += sheet_report['created_sku']

    return jsonify({'success': True, 'report': report})


@inventory_bp.route('/api/inventory/import-supplier', methods=['POST'])
def import_supplier_inventory():
    d = get_db()
    supplier_file = request.files.get('supplier_file')
    maintenance_file = request.files.get('maintenance_file')
    password = request.form.get('password', '')
    warehouse = 'Lodz warehouse'
    has_supplier = bool(supplier_file and supplier_file.filename)
    has_maintenance = bool(maintenance_file and maintenance_file.filename)
    if not has_supplier and not has_maintenance:
        return jsonify({'error': '请至少上传供应商库存表或库存维护表'}), 400
    if has_supplier and not supplier_file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '供应商库存表必须是 Excel 文件'}), 400
    if has_maintenance and not maintenance_file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '库存维护表必须是 Excel 文件'}), 400
    if (has_supplier and uploaded_file_size(supplier_file) > 20 * 1024 * 1024) or (has_maintenance and uploaded_file_size(maintenance_file) > 20 * 1024 * 1024):
        return jsonify({'error': 'Excel 文件不能超过 20MB'}), 400

    supplier_required = ['Product Number', 'Instruction', 'Category', 'Category 2', 'Product Description', 'Sum of Quantity']
    maintenance_required = ['Product Number', 'Bom Code', 'Unit Price', 'Last Inbound Date', 'Last Outbound Date', 'Remark', 'DOS Threshold', 'Idle Threshold']
    supplier_aliases = {
        'Instruction': ['Instructions', 'Instruction Code', 'InstructionCode'],
        'Category 2': ['Category2', 'Category II', 'Product Category', 'Applicable Product'],
        'Product Description': ['Production Description', 'Description', 'Product Desc', 'Item Description'],
        'Sum of Quantity': ['Sum Quantity', 'Quantity', 'Qty', 'SUMA Z QUANTITY', 'Suma z Quantity', 'Suma of Quantity', 'Sum z Quantity', 'Sum of Qty']
    }
    maintenance_aliases = {
        'Bom Code': ['BOM Code', 'BomCode', 'BOM', 'BOM No'],
        'Unit Price': ['UnitPrice', 'Price', 'Unit Cost'],
        'Last Inbound Date': ['Last Inbound', 'Inbound Date', 'Last Receipt Date'],
        'Last Outbound Date': ['Last Outbound', 'Outbound Date', 'Last Issue Date'],
        'Remark': ['Remarks', 'Comment', 'Comments', 'Notes'],
        'DOS Threshold': ['DOS', 'DOS Days', 'DOS Line'],
        'Idle Threshold': ['Idle', 'Idle Days', 'Idle Line']
    }

    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'supplier_inventory_import')
    report = {
        'warehouse': warehouse,
        'supplier_header_row': None,
        'maintenance_header_row': None,
        'supplier_rows': 0,
        'maintenance_rows': 0,
        'imported': 0,
        'created_products': 0,
        'updated_products': 0,
        'updated_inventory': 0,
        'removed_inventory': 0,
        'skipped_supplier_rows': 0,
        'used_existing_maintenance': 0,
        'inferred_bom': [],
        'missing_bom': [],
        'missing_images': [],
        'errors': []
    }

    def find_image(bom_code, product_number=''):
        return find_existing_image_filename(image_key_for_sku(bom_code, product_number))

    def infer_bom_code(product_number):
        if '_' in product_number:
            return clean_code(product_number.split('_', 1)[1].strip())
        return clean_code(product_number)

    if has_maintenance:
        maintenance_wb = openpyxl.load_workbook(maintenance_file, data_only=True)
        maintenance_ws = maintenance_wb.active
        maintenance_header_row, maintenance_cols = find_header_row(maintenance_ws, maintenance_required, aliases=maintenance_aliases)
        report['maintenance_header_row'] = maintenance_header_row
        if not maintenance_header_row:
            return jsonify({'error': '维护表未找到表头行，请确认包含 Product Number / Bom Code / Unit Price 等字段'}), 400

        for row_idx in range(maintenance_header_row + 1, maintenance_ws.max_row + 1):
            product_number = clean_code(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Product Number'))
            if not product_number:
                continue
            report['maintenance_rows'] += 1
            bom_code = clean_code(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Bom Code'))
            if not bom_code or is_local_bom_code(bom_code):
                bom_code = '/'
            unit_price = to_float(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Unit Price'))
            last_inbound = date_text(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Last Inbound Date'))
            last_outbound = date_text(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Last Outbound Date'))
            remark = clean_text(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Remark'))
            dos_threshold = to_int(get_cell(maintenance_ws, row_idx, maintenance_cols, 'DOS Threshold'), 180)
            idle_threshold = to_int(get_cell(maintenance_ws, row_idx, maintenance_cols, 'Idle Threshold'), 180)
            product_exists = d.execute("SELECT 1 FROM product_master WHERE product_number = ?", (product_number,)).fetchone()
            image_filename = find_image(bom_code, product_number)

            d.execute("""INSERT INTO product_master
                            (product_number, bom_code, unit_price, last_inbound_date, last_outbound_date, remark, dos_threshold, idle_threshold, updated_at)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                         ON CONFLICT(product_number) DO UPDATE SET
                            bom_code=excluded.bom_code,
                            unit_price=CASE
                                WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                                ELSE product_master.unit_price
                            END,
                            last_inbound_date=excluded.last_inbound_date,
                            last_outbound_date=excluded.last_outbound_date,
                            remark=excluded.remark,
                            dos_threshold=excluded.dos_threshold,
                            idle_threshold=excluded.idle_threshold,
                            updated_at=CURRENT_TIMESTAMP""",
                      (product_number, bom_code, unit_price, last_inbound or None, last_outbound or None, remark, dos_threshold, idle_threshold))
            if is_local_bom_code(bom_code):
                image_key = image_key_for_sku(bom_code, product_number)
                if image_filename:
                    d.execute("""INSERT INTO sku_master (bom_code, product_number, image_path, updated_at)
                                 VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                                 ON CONFLICT(bom_code) DO UPDATE SET
                                    product_number=excluded.product_number,
                                    image_path=COALESCE(excluded.image_path, image_path),
                                    updated_at=CURRENT_TIMESTAMP""",
                              (image_key, product_number, image_filename))
            else:
                d.execute("""INSERT INTO sku_master (bom_code, product_number, unit_price, image_path, updated_at)
                             VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                             ON CONFLICT(bom_code) DO UPDATE SET
                                product_number=excluded.product_number,
                                unit_price=CASE
                                    WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                                    ELSE sku_master.unit_price
                                END,
                                image_path=COALESCE(excluded.image_path, image_path),
                                updated_at=CURRENT_TIMESTAMP""",
                          (bom_code, product_number, unit_price, image_filename))
            d.execute("""UPDATE warehouse_inventory
                         SET bom_code = ?,
                             last_inbound_date = ?,
                             last_outbound_date = ?,
                             comment = ?,
                             updated_at = CURRENT_TIMESTAMP
                         WHERE product_number = ?""",
                      (bom_code, last_inbound or None, last_outbound or None, remark, product_number))
            if product_exists:
                report['updated_products'] += 1
            else:
                report['created_products'] += 1
            if not image_filename:
                report['missing_images'].append(bom_code)

    if has_supplier:
        supplier_wb = openpyxl.load_workbook(supplier_file, data_only=True)
        supplier_ws = supplier_wb.active
        supplier_header_row, supplier_cols = find_header_row(supplier_ws, supplier_required, aliases=supplier_aliases)
        report['supplier_header_row'] = supplier_header_row
        if not supplier_header_row:
            return jsonify({'error': '供应商库存表未找到表头行，请确认包含 Product Number / Instruction / Product Description / Sum of Quantity'}), 400

        uploaded_products = set()
        for row_idx in range(supplier_header_row + 1, supplier_ws.max_row + 1):
            product_number = clean_code(get_cell(supplier_ws, row_idx, supplier_cols, 'Product Number'))
            if not product_number:
                continue
            report['supplier_rows'] += 1
            product_description = clean_text(get_cell(supplier_ws, row_idx, supplier_cols, 'Product Description'))
            category = normalize_dimension(get_cell(supplier_ws, row_idx, supplier_cols, 'Category'), 'category')
            category_2 = normalize_dimension(get_cell(supplier_ws, row_idx, supplier_cols, 'Category 2'), 'category2')
            instruction = normalize_dimension(get_cell(supplier_ws, row_idx, supplier_cols, 'Instruction'), 'instruction')
            raw_inventory = get_cell(supplier_ws, row_idx, supplier_cols, 'Sum of Quantity')
            inventory = to_int_strict(raw_inventory)
            invalid_reasons = []
            if is_summary_product_number(product_number):
                invalid_reasons.append('汇总行')
            if not category:
                invalid_reasons.append('缺 Category')
            if not category_2:
                invalid_reasons.append('缺 Category 2')
            if not instruction:
                invalid_reasons.append('缺 Instruction')
            if inventory is None:
                invalid_reasons.append('数量无效')
            if invalid_reasons:
                report['skipped_supplier_rows'] += 1
                if len(report['errors']) < 20:
                    report['errors'].append(f"供应商库存第 {row_idx} 行已跳过：{product_number}（{', '.join(invalid_reasons)}）")
                continue
            uploaded_products.add(product_number)
            existing = d.execute("SELECT * FROM product_master WHERE product_number = ?", (product_number,)).fetchone()
            inv_existing = d.execute("SELECT * FROM warehouse_inventory WHERE warehouse = ? AND product_number = ?",
                                     (warehouse, product_number)).fetchone()
            product_exists = bool(existing)
            if existing:
                bom_code = clean_code(existing['bom_code']) or infer_bom_code(product_number)
                unit_price = existing['unit_price'] or 0
                last_inbound = existing['last_inbound_date'] or ''
                last_outbound = existing['last_outbound_date'] or ''
                remark = existing['remark'] or ''
                dos_threshold = existing['dos_threshold'] or 180
                idle_threshold = existing['idle_threshold'] or 180
                report['used_existing_maintenance'] += 1
            elif inv_existing:
                bom_code = clean_code(inv_existing['bom_code']) or infer_bom_code(product_number)
                unit_price = 0
                last_inbound = inv_existing['last_inbound_date'] or ''
                last_outbound = inv_existing['last_outbound_date'] or ''
                remark = inv_existing['comment'] or ''
                dos_threshold = 180
                idle_threshold = 180
            else:
                bom_code = infer_bom_code(product_number)
                unit_price = 0
                last_inbound = ''
                last_outbound = ''
                remark = ''
                dos_threshold = 180
                idle_threshold = 180
                report['inferred_bom'].append(product_number)
                report['missing_bom'].append(product_number)

            image_filename = find_image(bom_code, product_number)
            image_key = image_key_for_sku(bom_code, product_number)
            if not image_filename:
                report['missing_images'].append(image_key or product_number)

            if not product_exists:
                d.execute("""INSERT INTO product_master
                                (product_number, bom_code, product_description, category, category_2, unit_price,
                                 last_inbound_date, last_outbound_date, remark, dos_threshold, idle_threshold, updated_at)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                             ON CONFLICT(product_number) DO NOTHING""",
                          (product_number, bom_code, product_description, category, category_2, unit_price,
                           last_inbound or None, last_outbound or None, remark, dos_threshold, idle_threshold))
                report['created_products'] += 1

            if product_exists:
                if image_filename:
                    d.execute("""INSERT INTO sku_master (bom_code, product_number, image_path, updated_at)
                                 VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                                 ON CONFLICT(bom_code) DO UPDATE SET
                                    product_number=COALESCE(NULLIF(excluded.product_number, ''), product_number),
                                    image_path=COALESCE(excluded.image_path, image_path),
                                    updated_at=CURRENT_TIMESTAMP""",
                              (image_key, product_number, image_filename))
            elif is_local_bom_code(bom_code):
                if image_filename:
                    d.execute("""INSERT INTO sku_master (bom_code, product_number, image_path, updated_at)
                                 VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                                 ON CONFLICT(bom_code) DO UPDATE SET
                                    product_number=excluded.product_number,
                                    image_path=COALESCE(excluded.image_path, image_path),
                                    updated_at=CURRENT_TIMESTAMP""",
                              (image_key, product_number, image_filename))
            else:
                d.execute("""INSERT INTO sku_master (bom_code, product_number, product_description, category, category_2, unit_price, image_path, updated_at)
                             VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                             ON CONFLICT(bom_code) DO UPDATE SET
                                product_number=excluded.product_number,
                                product_description=excluded.product_description,
                                category=excluded.category,
                                category_2=excluded.category_2,
                                unit_price=CASE
                                    WHEN COALESCE(excluded.unit_price, 0) > 0 THEN excluded.unit_price
                                    ELSE sku_master.unit_price
                                END,
                                image_path=COALESCE(excluded.image_path, image_path),
                                updated_at=CURRENT_TIMESTAMP""",
                          (bom_code, product_number, product_description, category, category_2, unit_price, image_filename))

            old_qty = inv_existing['inventory'] if inv_existing else None
            if inv_existing:
                d.execute("""UPDATE warehouse_inventory
                             SET inventory = ?, updated_at = CURRENT_TIMESTAMP
                             WHERE warehouse = ? AND product_number = ?""",
                          (inventory, warehouse, product_number))
            else:
                d.execute("""INSERT INTO warehouse_inventory
                                (warehouse, bom_code, product_number, instruction, inventory, last_inbound_date, last_outbound_date, comment, updated_at)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)""",
                          (warehouse, bom_code, product_number, instruction, inventory, last_inbound or None, last_outbound or None, remark))
            if old_qty != inventory:
                report['updated_inventory'] += 1
                if old_qty is not None:
                    d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, comment)
                                 VALUES (?, ?, ?, 'calibration', ?, ?)""",
                              (warehouse, clean_code(inv_existing['bom_code']) if inv_existing else bom_code, product_number, inventory - old_qty, f"供应商库存导入校准: {old_qty} → {inventory}"))
            report['imported'] += 1

        old_rows = d.execute("SELECT product_number, bom_code, inventory FROM warehouse_inventory WHERE warehouse = ?", (warehouse,)).fetchall()
        for row in old_rows:
            if row['product_number'] not in uploaded_products:
                d.execute("DELETE FROM warehouse_inventory WHERE warehouse = ? AND product_number = ?", (warehouse, row['product_number']))
                d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, comment)
                             VALUES (?, ?, ?, 'calibration', ?, ?)""",
                          (warehouse, row['bom_code'], row['product_number'], -int(row['inventory'] or 0), "供应商库存刷新移除：新表未包含该 Product Number"))
                report['removed_inventory'] += 1

    report['inferred_bom'] = sorted(set(report['inferred_bom']))
    report['missing_bom'] = sorted(set(report['missing_bom']))
    report['missing_images'] = sorted(set(report['missing_images']))
    d.commit()
    return jsonify({'success': True, 'report': report})


@inventory_bp.route('/api/inventory/import-calibration', methods=['POST'])
def import_inventory_calibration():
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传库存校准 Excel'}), 400
    if uploaded_file_size(f) > 25 * 1024 * 1024:
        return jsonify({'error': 'Excel 文件不能超过 25MB'}), 400

    try:
        db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'inventory_calibration_import')
        saved_filename, saved_path = save_uploaded_source_file(f, 'inventory_calibration')
        wb = openpyxl.load_workbook(saved_path, data_only=True)
        d = get_db()
        report = {'source_file': saved_filename, 'sheets': {}, 'imported': 0, 'created': 0, 'updated': 0, 'inventory_changed': 0, 'skipped': 0, 'errors': []}

        for sheet_title, warehouse in CALIBRATION_SHEETS.items():
            if sheet_title not in wb.sheetnames:
                report['sheets'][sheet_title] = {'warehouse': warehouse, 'missing': True, 'imported': 0, 'skipped': 0}
                report['errors'].append(f"缺少 Sheet: {sheet_title}")
                continue
            ws = wb[sheet_title]
            header_row, header_map = find_exact_header_map(ws, CALIBRATION_HEADERS)
            sheet_report = {'warehouse': warehouse, 'header_row': header_row, 'imported': 0, 'created': 0, 'updated': 0, 'inventory_changed': 0, 'skipped': 0}
            if not header_row:
                sheet_report['error'] = '未找到校准表头'
                report['errors'].append(f"{sheet_title} 未找到校准表头")
                report['sheets'][sheet_title] = sheet_report
                continue
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                before = dict(sheet_report)
                data = {h: get_row_value_by_headers(row, header_map, h) for h in CALIBRATION_HEADERS}
                if not any(clean_text(v) for v in data.values()):
                    continue
                upsert_inventory_calibration_row(d, warehouse, data, sheet_report)
                for key in ('imported', 'created', 'updated', 'inventory_changed', 'skipped'):
                    report[key] += sheet_report[key] - before[key]
            report['sheets'][sheet_title] = sheet_report

        d.commit()
        remove_uploaded_source_file(saved_path)
        return jsonify({'success': True, 'report': report})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        return jsonify({'error': f'库存校准导入失败：{str(e)}'}), 500


@inventory_bp.route('/api/inventory/import-specyfikacja', methods=['POST'])
def import_specyfikacja_movements():
    files = [f for f in request.files.getlist('files') if f and f.filename]
    single = request.files.get('file')
    if single and single.filename:
        files.append(single)
    if not files:
        return jsonify({'error': '请上传 Specyfikacja Excel'}), 400
    bad = [f.filename for f in files if not f.filename.endswith(('.xlsx', '.xls'))]
    if bad:
        return jsonify({'error': '只支持 .xlsx / .xls 文件：' + ', '.join(bad[:5])}), 400
    if any(uploaded_file_size(f) > 30 * 1024 * 1024 for f in files):
        return jsonify({'error': '单个 Excel 文件不能超过 30MB'}), 400

    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'specyfikacja_import')
    d = get_db()
    inbound_latest = {}
    outbound_latest = {}
    report = {'source_file': '', 'source_files': [], 'file_count': 0, 'sheets': {}, 'files': [], 'inbound_rows': 0, 'outbound_rows': 0, 'updated_products': 0, 'updated_inventory_rows': 0, 'skipped': 0, 'errors': []}
    required = ['DATA RUCHU', 'SKU', 'ID_OPERACJI', 'ILOŚĆ', 'UoM']
    aliases = {
        'DATA RUCHU': ['DATA', 'Data ruchu', 'Movement Date'],
        'SKU': ['Product Number', 'PRODUCT NUMBER', 'Indeks', 'Towar'],
        'ID_OPERACJI': ['ID OPERACJI', 'ID', 'Operation ID'],
        'ILOŚĆ': ['ILOSC', 'ILOSĆ', 'Quantity', 'Qty'],
        'UoM': ['UOM', 'Unit', 'Unit of Measure']
    }

    def latest_set(target, sku, date_value):
        if not sku or not date_value:
            return
        old = target.get(sku)
        if not old or date_value > old:
            target[sku] = date_value

    saved_paths = []
    for f in files:
        try:
            saved_filename, saved_path = save_uploaded_source_file(f, 'specyfikacja')
            saved_paths.append(saved_path)
            report['source_files'].append(saved_filename)
            wb = openpyxl.load_workbook(saved_path, data_only=True)
            file_report = {'source_file': saved_filename, 'sheets': {}, 'inbound_rows': 0, 'outbound_rows': 0, 'skipped': 0, 'errors': []}
            for sheet_name in wb.sheetnames:
                normalized_sheet = sheet_name.strip().lower()
                sheet_key = f"{saved_filename} / {sheet_name}"
                if normalized_sheet == 'arkusz1':
                    sheet_report = {'ignored': True}
                    report['sheets'][sheet_key] = sheet_report
                    file_report['sheets'][sheet_name] = sheet_report
                    continue
                is_inbound = normalized_sheet == 'inbound'
                is_outbound = normalized_sheet in ('outbound polska', 'outbound europa')
                if not is_inbound and not is_outbound:
                    sheet_report = {'ignored': True, 'reason': '非出入库流水 Sheet'}
                    report['sheets'][sheet_key] = sheet_report
                    file_report['sheets'][sheet_name] = sheet_report
                    continue
                ws = wb[sheet_name]
                header_row, col_map = find_header_row(ws, required, aliases=aliases)
                sheet_report = {'header_row': header_row, 'type': 'inbound' if is_inbound else 'outbound', 'rows': 0, 'skipped': 0}
                if not header_row:
                    sheet_report['error'] = '未找到 DATA RUCHU / SKU / ID_OPERACJI / ILOŚĆ / UoM 表头'
                    msg = f"{saved_filename} / {sheet_name} 未找到必要表头"
                    report['errors'].append(msg)
                    file_report['errors'].append(msg)
                    report['sheets'][sheet_key] = sheet_report
                    file_report['sheets'][sheet_name] = sheet_report
                    continue
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    sku = clean_code(get_cell(ws, row_idx, col_map, 'SKU'))
                    move_date = parse_date_text(get_cell(ws, row_idx, col_map, 'DATA RUCHU'))
                    operation_id = clean_text(get_cell(ws, row_idx, col_map, 'ID_OPERACJI'))
                    qty = to_float(get_cell(ws, row_idx, col_map, 'ILOŚĆ'), 0)
                    uom = clean_text(get_cell(ws, row_idx, col_map, 'UoM'))
                    if not sku or is_summary_row_value(sku) or not move_date:
                        sheet_report['skipped'] += 1
                        file_report['skipped'] += 1
                        report['skipped'] += 1
                        continue
                    sheet_report['rows'] += 1
                    op_prefix, op_target, order_date = parse_operation_id(operation_id)
                    d.execute("""INSERT INTO inventory_movements
                                    (movement_type, sheet_name, product_number, movement_date, operation_id,
                                     operation_prefix, operation_target, order_date, quantity, uom, source_file)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                              ('inbound' if is_inbound else 'outbound', sheet_name, sku, move_date, operation_id,
                               op_prefix, op_target, order_date or None, qty, uom, saved_filename))
                    if is_inbound:
                        latest_set(inbound_latest, sku, move_date)
                        file_report['inbound_rows'] += 1
                        report['inbound_rows'] += 1
                    else:
                        latest_set(outbound_latest, sku, move_date)
                        file_report['outbound_rows'] += 1
                        report['outbound_rows'] += 1
                report['sheets'][sheet_key] = sheet_report
                file_report['sheets'][sheet_name] = sheet_report
            report['files'].append(file_report)
        except Exception as e:
            msg = f"{f.filename}: {str(e)}"
            report['errors'].append(msg)
            report['files'].append({'source_file': f.filename, 'error': str(e)})
            if len(files) == 1:
                d.rollback()
                return jsonify({'error': f'Specyfikacja 导入失败：{str(e)}'}), 500
            continue

    all_skus = sorted(set(inbound_latest) | set(outbound_latest))
    for sku in all_skus:
        existing = d.execute("SELECT created_at, last_inbound_date, last_outbound_date FROM product_master WHERE product_number = ?", (sku,)).fetchone()
        fallback_inbound = ''
        if existing and not inbound_latest.get(sku) and not existing['last_inbound_date']:
            fallback_inbound = parse_date_text(existing['created_at'])
        inbound = inbound_latest.get(sku) or fallback_inbound
        outbound = outbound_latest.get(sku) or inbound
        product_cur = d.execute("""UPDATE product_master
                                   SET last_inbound_date = COALESCE(NULLIF(?, ''), last_inbound_date),
                                       last_outbound_date = COALESCE(NULLIF(?, ''), last_outbound_date),
                                       updated_at = CURRENT_TIMESTAMP
                                   WHERE product_number = ?""", (inbound, outbound, sku))
        report['updated_products'] += product_cur.rowcount if product_cur.rowcount and product_cur.rowcount > 0 else 0
        cur = d.execute("""UPDATE warehouse_inventory
                           SET last_inbound_date = COALESCE(NULLIF(?, ''), last_inbound_date),
                               last_outbound_date = COALESCE(NULLIF(?, ''), last_outbound_date),
                               updated_at = CURRENT_TIMESTAMP
                           WHERE product_number = ?""", (inbound, outbound, sku))
        report['updated_inventory_rows'] += cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0

    # If a SKU has inbound history but no outbound history, treat the first
    # inbound date as the idle baseline instead of leaving Last Outbound blank.
    d.execute("""
        UPDATE product_master
        SET last_outbound_date = last_inbound_date,
            updated_at = CURRENT_TIMESTAMP
        WHERE (last_outbound_date IS NULL OR TRIM(last_outbound_date) = '')
          AND last_inbound_date IS NOT NULL
          AND TRIM(last_inbound_date) != ''
    """)
    d.execute("""
        UPDATE warehouse_inventory
        SET last_outbound_date = last_inbound_date,
            updated_at = CURRENT_TIMESTAMP
        WHERE (last_outbound_date IS NULL OR TRIM(last_outbound_date) = '')
          AND last_inbound_date IS NOT NULL
          AND TRIM(last_inbound_date) != ''
    """)

    d.commit()
    report['file_count'] = len(report['source_files'])
    report['source_file'] = '；'.join(report['source_files'])
    for saved_path in saved_paths:
        remove_uploaded_source_file(saved_path)
    return jsonify({'success': True, 'report': report})


@inventory_bp.route('/api/inventory/product/<path:product_number>')
def get_inventory_product_detail(product_number):
    d = get_db()
    requested_product_number = clean_code(product_number)
    match_key = code_match_value(requested_product_number)
    candidates = set([requested_product_number]) if requested_product_number else set()
    direct_identity_rows = d.execute(f"""
        SELECT bom_code FROM product_master WHERE {code_match_expr('product_number')} = ?
        UNION
        SELECT bom_code FROM warehouse_inventory WHERE {code_match_expr('product_number')} = ?
        UNION
        SELECT bom_code FROM sku_master WHERE {code_match_expr('product_number')} = ?
    """, (match_key, match_key, match_key)).fetchall()
    local_product_identity = bool(direct_identity_rows) and not any(
        row['bom_code'] and not is_local_bom_code(row['bom_code']) for row in direct_identity_rows
    )

    # Product Number may differ by case, spacing, underscore, hyphen, or Excel ".0".
    # Build a candidate set first, then use it for master data, stock, and movement lookup.
    if local_product_identity:
        lookup_rows = d.execute(f"""
            SELECT product_number, bom_code FROM product_master
            WHERE {code_match_expr('product_number')} = ?
            UNION
            SELECT product_number, bom_code FROM warehouse_inventory
            WHERE {code_match_expr('product_number')} = ?
            UNION
            SELECT product_number, bom_code FROM sku_master
            WHERE {code_match_expr('product_number')} = ?
        """, (match_key, match_key, match_key)).fetchall()
    else:
        lookup_rows = d.execute(f"""
            SELECT product_number, bom_code FROM product_master
            WHERE {code_match_expr('product_number')} = ?
               OR {code_match_expr('bom_code')} = ?
            UNION
            SELECT product_number, bom_code FROM warehouse_inventory
            WHERE {code_match_expr('product_number')} = ?
               OR {code_match_expr('bom_code')} = ?
            UNION
            SELECT product_number, bom_code FROM sku_master
            WHERE {code_match_expr('product_number')} = ?
               OR {code_match_expr('bom_code')} = ?
        """, (match_key, match_key, match_key, match_key, match_key, match_key)).fetchall()
    bom_candidates = set()
    for row in lookup_rows:
        if row['product_number']:
            candidates.add(clean_code(row['product_number']))
        if not local_product_identity and row['bom_code'] and not is_local_bom_code(row['bom_code']):
            bom_candidates.add(clean_code(row['bom_code']))
    if bom_candidates:
        placeholders = ','.join('?' for _ in bom_candidates)
        for row in d.execute(f"""
            SELECT product_number FROM product_master WHERE bom_code IN ({placeholders})
            UNION
            SELECT product_number FROM warehouse_inventory WHERE bom_code IN ({placeholders})
            UNION
            SELECT product_number FROM sku_master WHERE bom_code IN ({placeholders})
        """, list(bom_candidates) * 3).fetchall():
            if row['product_number']:
                candidates.add(clean_code(row['product_number']))

    candidate_list = sorted(c for c in candidates if c)
    candidate_keys = sorted(set(code_match_value(c) for c in candidate_list if c))
    exact_placeholders = ','.join('?' for _ in candidate_list) or "''"
    key_placeholders = ','.join('?' for _ in candidate_keys) or "''"

    product = d.execute("""
        SELECT pm.*, COALESCE(pimg.image_path, sm.image_path) AS image_path
        FROM product_master pm
        LEFT JOIN sku_master sm ON pm.bom_code = sm.bom_code
                                 AND NOT (REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE') OR COALESCE(NULLIF(pm.bom_code, ''), '') = '')
        LEFT JOIN sku_master pimg ON (REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE') OR COALESCE(NULLIF(pm.bom_code, ''), '') = '')
                                 AND pm.product_number = pimg.bom_code
        WHERE pm.product_number IN ({})
           OR {} IN ({})
        ORDER BY CASE WHEN pm.product_number = ? THEN 0 ELSE 1 END, pm.updated_at DESC
        LIMIT 1
    """.format(exact_placeholders, code_match_expr('pm.product_number'), key_placeholders),
        candidate_list + candidate_keys + [requested_product_number]).fetchone()
    inventory = d.execute("""
        SELECT id, warehouse, bom_code, product_number, instruction, inventory,
               last_inbound_date, last_outbound_date, comment, updated_at
        FROM warehouse_inventory
        WHERE product_number IN ({})
           OR {} IN ({})
        ORDER BY warehouse
    """.format(exact_placeholders, code_match_expr('product_number'), key_placeholders),
        candidate_list + candidate_keys).fetchall()
    movements = d.execute("""
        SELECT movement_type, sheet_name, product_number, movement_date, operation_id, operation_prefix,
               operation_target, order_date, quantity, uom, source_file
        FROM inventory_movements
        WHERE product_number IN ({})
           OR {} IN ({})
        ORDER BY movement_date DESC, id DESC
        LIMIT 30
    """.format(exact_placeholders, code_match_expr('product_number'), key_placeholders),
        candidate_list + candidate_keys).fetchall()
    if not product and not inventory:
        return jsonify({'error': 'Product Number 不存在'}), 404
    return jsonify({
        'product': dict(product) if product else {},
        'inventory': [dict(r) for r in inventory],
        'movements': [dict(r) for r in movements],
        'matched_product_numbers': candidate_list,
    })


@inventory_bp.route('/api/inventory/adjust', methods=['POST'])
def adjust_inventory():
    d = get_db()
    data = request.get_json()
    warehouse = data.get('warehouse', '')
    product_number, bom_code = resolve_product_identity(d, data.get('product_number', ''), data.get('bom_code', ''))
    change_type = data.get('change_type', '')
    quantity = data.get('quantity', 0)
    comment = data.get('comment', '')

    if not all([warehouse, product_number, change_type, quantity]):
        return jsonify({'error': 'Missing required fields'}), 400

    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], f"inventory_adjust_{product_number}")
    d.execute("""INSERT INTO warehouse_inventory (warehouse, bom_code, product_number, inventory)
                 VALUES (?, ?, ?, 0)
                 ON CONFLICT(warehouse, product_number) DO NOTHING""", (warehouse, bom_code, product_number))

    row = d.execute("SELECT inventory FROM warehouse_inventory WHERE warehouse = ? AND product_number = ?",
                    (warehouse, product_number)).fetchone()
    current = row['inventory'] if row else 0

    if change_type == 'outbound':
        new_qty = max(0, current - quantity)
        actual_change = -(current - new_qty)
    elif change_type == 'scrap':
        new_qty = max(0, current - quantity)
        actual_change = -(current - new_qty)
    elif change_type == 'inbound':
        new_qty = current + quantity
        actual_change = quantity
    else:
        new_qty = quantity
        actual_change = quantity - current

    d.execute("UPDATE warehouse_inventory SET inventory = ?, updated_at = CURRENT_TIMESTAMP WHERE warehouse = ? AND product_number = ?",
              (new_qty, warehouse, product_number))

    if change_type == 'inbound':
        d.execute("UPDATE warehouse_inventory SET last_inbound_date = DATE('now') WHERE warehouse = ? AND product_number = ?",
                  (warehouse, product_number))
    elif change_type in ('outbound', 'scrap'):
        d.execute("UPDATE warehouse_inventory SET last_outbound_date = DATE('now') WHERE warehouse = ? AND product_number = ?",
                  (warehouse, product_number))

    d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, comment)
                 VALUES (?, ?, ?, ?, ?, ?)""",
              (warehouse, bom_code, product_number, change_type, actual_change, comment))

    d.commit()
    return jsonify({'success': True, 'new_quantity': new_qty})


@inventory_bp.route('/api/inventory/transactions', methods=['GET'])
def get_transactions():
    d = get_db()
    warehouse = request.args.get('warehouse', '')
    bom_code = request.args.get('bom_code', '')
    product_number = request.args.get('product_number', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    where_clauses = []
    params = []
    if warehouse:
        where_clauses.append("warehouse = ?")
        params.append(warehouse)
    if bom_code:
        where_clauses.append("t.bom_code = ?")
        params.append(bom_code)
    if product_number:
        where_clauses.append("t.product_number = ?")
        params.append(product_number)

    where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    offset = (page - 1) * per_page

    total = d.execute(f"SELECT COUNT(*) FROM inventory_transactions {where}", params).fetchone()[0]
    rows = d.execute(f"""
        SELECT t.*, sm.product_description
        FROM inventory_transactions t
        LEFT JOIN sku_master sm ON t.bom_code = sm.bom_code
        {where}
        ORDER BY t.created_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


# ==================== 图片上传 ====================

@inventory_bp.route('/api/inventory/images/upload', methods=['POST'])
def upload_images():
    uploaded = []
    not_found = []
    d = get_db()

    files = request.files.getlist('files')
    for f in files:
        if not f or not allowed_file(f.filename):
            continue
        if uploaded_file_size(f) > 5 * 1024 * 1024:
            not_found.append(f"{f.filename} (超过5MB)")
            continue
        original_code = image_code_candidates(f.filename)[0] if image_code_candidates(f.filename) else ''
        bom_code = resolve_image_bom(d, f.filename)
        if not bom_code:
            not_found.append(original_code or f.filename)
            continue
        ext = f.filename.rsplit('.', 1)[1].lower()
        filename = secure_filename(f"{bom_code}.{ext}")
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        f.save(filepath)
        d.execute("""INSERT INTO sku_master (bom_code, image_path, updated_at)
                     VALUES (?, ?, CURRENT_TIMESTAMP)
                     ON CONFLICT(bom_code) DO UPDATE SET image_path = excluded.image_path, updated_at = CURRENT_TIMESTAMP""",
                  (bom_code, filename))
        uploaded.append(bom_code)

    repaired = repair_image_mappings(d)
    d.commit()
    return jsonify({'success': True, 'uploaded': uploaded, 'not_found': not_found, 'repaired': repaired})


@inventory_bp.route('/api/inventory/images', methods=['DELETE'])
def clear_inventory_images():
    data = request.get_json(silent=True) or {}
    if not maintenance_password_ok(data.get('password'), allow_admin=True):
        return jsonify({'error': '清理密码错误'}), 403

    folder = current_app.config['UPLOAD_FOLDER']
    deleted, failed = clear_image_folder(folder)

    d = get_db()
    d.execute("UPDATE sku_master SET image_path = NULL, updated_at = CURRENT_TIMESTAMP WHERE COALESCE(image_path, '') <> ''")
    d.commit()
    return jsonify({'success': True, 'deleted': deleted, 'failed': failed})


# ==================== 发货指令管理 ====================

def _generate_order_number():
    today = datetime.now().strftime('%Y%m%d')
    d = get_db()
    prefix = f"DO-{today}-"
    last = d.execute("SELECT order_number FROM delivery_orders WHERE order_number LIKE ? ORDER BY id DESC LIMIT 1",
                     (f"{prefix}%",)).fetchone()
    if last:
        seq = int(last['order_number'].split('-')[-1]) + 1
    else:
        seq = 1
    return f"{prefix}{seq:03d}"


@delivery_bp.route('/api/orders', methods=['GET'])
def get_orders():
    d = get_db()
    status = request.args.get('status', '')
    warehouse = request.args.get('warehouse', '')
    keyword = request.args.get('q', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    where_clauses = []
    params = []
    if status:
        where_clauses.append("status = ?")
        params.append(status)
    if warehouse:
        where_clauses.append("warehouse = ?")
        params.append(warehouse)
    if keyword:
        where_clauses.append("(order_number LIKE ? OR stt_number LIKE ? OR destination_country LIKE ?)")
        params.extend([f'%{keyword}%'] * 3)

    where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    offset = (page - 1) * per_page

    total = d.execute(f"SELECT COUNT(*) FROM delivery_orders {where}", params).fetchone()[0]
    rows = d.execute(f"""
        SELECT o.*,
               (SELECT COUNT(*) FROM delivery_order_items WHERE order_id = o.id) AS item_count,
               (SELECT COALESCE(SUM(requested_qty), 0) FROM delivery_order_items WHERE order_id = o.id) AS total_requested,
               (SELECT COALESCE(SUM(actual_qty), 0) FROM delivery_order_items WHERE order_id = o.id) AS total_actual
        FROM delivery_orders o
        {where}
        ORDER BY o.created_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


@delivery_bp.route('/api/orders', methods=['POST'])
def create_order():
    d = get_db()
    data = request.get_json()
    warehouse = data.get('warehouse', '')
    dest_country = data.get('destination_country', '')
    items = data.get('items', [])

    if not warehouse or not items:
        return jsonify({'error': 'Warehouse and items are required'}), 400

    order_number = _generate_order_number()
    cur = d.execute("""INSERT INTO delivery_orders (order_number, warehouse, destination_country)
                       VALUES (?, ?, ?)""", (order_number, warehouse, dest_country))
    order_id = cur.lastrowid

    for item in items:
        product_number, bom_code = resolve_product_identity(d, item.get('product_number', ''), item.get('bom_code', ''))
        requested_qty = item.get('requested_qty', 0)
        if not product_number or requested_qty <= 0:
            continue
        d.execute("""INSERT INTO delivery_order_items (order_id, bom_code, product_number, requested_qty)
                     VALUES (?, ?, ?, ?)""", (order_id, bom_code, product_number, requested_qty))

    d.commit()
    return jsonify({'success': True, 'order_id': order_id, 'order_number': order_number})


@delivery_bp.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order_detail(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404

    items = d.execute("""
        SELECT oi.*, sm.product_description, COALESCE(NULLIF(oi.product_number, ''), sm.product_number) AS product_number,
               sm.category, sm.unit_price
        FROM delivery_order_items oi
        LEFT JOIN sku_master sm ON oi.bom_code = sm.bom_code
        WHERE oi.order_id = ?
    """, (order_id,)).fetchall()

    return jsonify({'order': dict(order), 'items': [dict(r) for r in items]})


@delivery_bp.route('/api/orders/<int:order_id>', methods=['PUT'])
def update_order(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'pending':
        return jsonify({'error': 'Only pending orders can be updated'}), 400

    data = request.get_json()
    if 'destination_country' in data:
        d.execute("UPDATE delivery_orders SET destination_country = ? WHERE id = ?",
                  (data['destination_country'], order_id))

    if 'items' in data:
        d.execute("DELETE FROM delivery_order_items WHERE order_id = ?", (order_id,))
        for item in data['items']:
            product_number, bom_code = resolve_product_identity(d, item.get('product_number', ''), item.get('bom_code', ''))
            d.execute("""INSERT INTO delivery_order_items (order_id, bom_code, product_number, requested_qty)
                         VALUES (?, ?, ?, ?)""", (order_id, bom_code, product_number, item['requested_qty']))

    d.commit()
    return jsonify({'success': True})


@delivery_bp.route('/api/orders/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'pending':
        return jsonify({'error': 'Only pending orders can be deleted'}), 400

    d.execute("DELETE FROM delivery_order_items WHERE order_id = ?", (order_id,))
    d.execute("DELETE FROM delivery_orders WHERE id = ?", (order_id,))
    d.commit()
    return jsonify({'success': True})


@delivery_bp.route('/api/orders/<int:order_id>/confirm', methods=['PUT'])
def confirm_order(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'pending':
        return jsonify({'error': 'Only pending orders can be confirmed'}), 400

    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], f"order_confirm_{order['order_number']}")
    data = request.get_json()
    actual_items = data.get('items', [])

    for item in actual_items:
        product_number, bom_code = resolve_product_identity(d, item.get('product_number', ''), item.get('bom_code', ''))
        actual_qty = item.get('actual_qty', 0)
        if not product_number:
            continue
        d.execute("UPDATE delivery_order_items SET actual_qty = ? WHERE order_id = ? AND product_number = ?",
                  (actual_qty, order_id, product_number))

    items = d.execute("SELECT * FROM delivery_order_items WHERE order_id = ?", (order_id,)).fetchall()
    warehouse = order['warehouse']
    for item in items:
        actual_qty = item['actual_qty']
        if actual_qty is None or actual_qty <= 0:
            continue
        product_number = item['product_number'] or item['bom_code']
        bom_code = item['bom_code']
        d.execute("""INSERT INTO warehouse_inventory (warehouse, bom_code, product_number, inventory)
                     VALUES (?, ?, ?, 0) ON CONFLICT(warehouse, product_number) DO NOTHING""",
                  (warehouse, bom_code, product_number))
        d.execute("UPDATE warehouse_inventory SET inventory = MAX(0, inventory - ?), last_outbound_date = DATE('now'), updated_at = CURRENT_TIMESTAMP WHERE warehouse = ? AND product_number = ?",
                  (actual_qty, warehouse, product_number))
        d.execute("""INSERT INTO inventory_transactions (warehouse, bom_code, product_number, change_type, quantity_change, order_id, comment)
                     VALUES (?, ?, ?, 'order_ship', ?, ?, ?)""",
                  (warehouse, bom_code, product_number, -actual_qty, order_id, f"Order {order['order_number']}"))

    d.execute("UPDATE delivery_orders SET status = 'confirmed', confirmed_at = CURRENT_TIMESTAMP WHERE id = ?", (order_id,))
    d.commit()
    return jsonify({'success': True, 'status': 'confirmed'})


@delivery_bp.route('/api/orders/<int:order_id>/ship', methods=['PUT'])
def ship_order(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'confirmed':
        return jsonify({'error': 'Only confirmed orders can be shipped'}), 400

    data = request.get_json()
    stt_number = data.get('stt_number', '').strip()
    if not stt_number:
        return jsonify({'error': 'STT number is required'}), 400

    d.execute("UPDATE delivery_orders SET status = 'shipped', stt_number = ?, shipped_at = CURRENT_TIMESTAMP WHERE id = ?",
              (stt_number, order_id))
    d.commit()
    return jsonify({'success': True, 'status': 'shipped', 'stt_number': stt_number})


@delivery_bp.route('/api/orders/<int:order_id>/deliver', methods=['PUT'])
def deliver_order(order_id):
    d = get_db()
    order = d.execute("SELECT * FROM delivery_orders WHERE id = ?", (order_id,)).fetchone()
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'shipped':
        return jsonify({'error': 'Only shipped orders can be marked as delivered'}), 400

    d.execute("UPDATE delivery_orders SET status = 'delivered', delivered_at = CURRENT_TIMESTAMP WHERE id = ?", (order_id,))
    d.commit()
    return jsonify({'success': True, 'status': 'delivered'})
