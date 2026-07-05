import sqlite3
import re
import os
import uuid
import csv
import io
import json
import hashlib
import subprocess
import tempfile
import zipfile
import urllib.request
import urllib.parse
from datetime import datetime
from flask import Blueprint, jsonify, g, current_app, request, send_file
from werkzeug.utils import secure_filename
from . import db as db_module

main_bp = Blueprint('main', __name__)

APP_VERSION = 'VN62'
APP_BUILD_TIME = '2026-07-05 21:55 CEST'
UPDATE_MANIFEST_URL = os.environ.get(
    'EUCWH_UPDATE_MANIFEST_URL',
    'https://raw.githubusercontent.com/piplupssss/EU-Retail-CWH-/main/eucwh-updates/latest.json'
)
UPDATE_FALLBACK_MANIFEST_URL = 'https://qitengliu.com/eucwh-updates/latest.json'
UPDATE_ALLOWED_HOSTS = {'qitengliu.com', 'www.qitengliu.com', 'raw.githubusercontent.com'}
UPDATE_MAX_BYTES = 300 * 1024 * 1024


def version_number(value):
    match = re.search(r'VN\s*(\d+)', str(value or ''), re.I)
    return int(match.group(1)) if match else 0


def validate_update_url(url, expected_path_prefix=None):
    parsed = urllib.parse.urlparse(str(url or '').strip())
    if parsed.scheme != 'https':
        raise ValueError('更新地址必须使用 HTTPS')
    host = (parsed.hostname or '').lower()
    if host not in UPDATE_ALLOWED_HOSTS:
        raise ValueError(f'更新来源不在白名单：{host or "-"}')
    if expected_path_prefix:
        path = parsed.path or ''
        if host in {'qitengliu.com', 'www.qitengliu.com'}:
            if not path.startswith(expected_path_prefix):
                raise ValueError('更新路径不在允许目录内')
        elif host == 'raw.githubusercontent.com':
            if '/EU-Retail-CWH-/' not in path or '/eucwh-updates/' not in path:
                raise ValueError('GitHub 更新路径不在允许目录内')
    return parsed.geturl()


def fetch_update_manifest():
    errors = []
    raw = None
    for manifest_url in [UPDATE_MANIFEST_URL, UPDATE_FALLBACK_MANIFEST_URL]:
        try:
            url = validate_update_url(manifest_url, '/eucwh-updates/')
            req = urllib.request.Request(url, headers={'User-Agent': f'EUCWH-Updater/{APP_VERSION}'})
            with urllib.request.urlopen(req, timeout=12) as resp:
                if getattr(resp, 'status', 200) >= 400:
                    raise RuntimeError(f'HTTP {resp.status}')
                raw = resp.read(512 * 1024 + 1)
            break
        except Exception as e:
            errors.append(str(e))
    if raw is None:
        raise RuntimeError('更新清单读取失败：' + ' | '.join(errors))
    if len(raw) > 512 * 1024:
        raise RuntimeError('更新清单过大')
    data = json.loads(raw.decode('utf-8-sig'))
    if not isinstance(data, dict):
        raise RuntimeError('更新清单格式错误')
    package_url = validate_update_url(data.get('package_url'), '/eucwh-updates/packages/')
    data['package_url'] = package_url
    data['latest_version'] = str(data.get('latest_version') or data.get('version') or '').strip()
    data['sha256'] = str(data.get('sha256') or '').strip().lower()
    if not data['latest_version'] or not re.fullmatch(r'VN\d+', data['latest_version'], re.I):
        raise RuntimeError('更新清单缺少 latest_version')
    if not re.fullmatch(r'[0-9a-f]{64}', data['sha256']):
        raise RuntimeError('更新清单缺少合法 SHA256')
    notes = data.get('notes') or []
    data['notes'] = notes if isinstance(notes, list) else [str(notes)]
    return data


def software_root_dir():
    base = app_base_dir()
    return os.path.dirname(base) if os.path.basename(base).lower() == '_system' else base


def download_update_zip(package_url, expected_sha256, version):
    update_dir = os.path.join(software_root_dir(), '_downloaded_updates')
    os.makedirs(update_dir, exist_ok=True)
    filename = safe_filename_part(os.path.basename(urllib.parse.urlparse(package_url).path), f'{version}_update.zip')
    if not filename.lower().endswith('.zip'):
        filename += '.zip'
    zip_path = os.path.join(update_dir, filename)
    req = urllib.request.Request(package_url, headers={'User-Agent': f'EUCWH-Updater/{APP_VERSION}'})
    digest = hashlib.sha256()
    total = 0
    with urllib.request.urlopen(req, timeout=60) as resp, open(zip_path, 'wb') as out:
        if getattr(resp, 'status', 200) >= 400:
            raise RuntimeError(f'HTTP {resp.status}')
        while True:
            chunk = resp.read(1024 * 512)
            if not chunk:
                break
            total += len(chunk)
            if total > UPDATE_MAX_BYTES:
                raise RuntimeError('更新包超过大小限制')
            digest.update(chunk)
            out.write(chunk)
    actual = digest.hexdigest()
    if actual.lower() != expected_sha256.lower():
        remove_file_safely(zip_path)
        raise RuntimeError('更新包 SHA256 校验失败')
    return zip_path, total, actual


def safe_extract_update_zip(zip_path):
    root = software_root_dir()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        if not names:
            raise RuntimeError('更新包为空')
        top_dirs = {name.split('/')[0] for name in names if name and not name.startswith('../') and '/' in name}
        if len(top_dirs) != 1:
            raise RuntimeError('更新包结构错误：必须包含单一顶层目录')
        top_dir = next(iter(top_dirs))
        for member in names:
            target = os.path.abspath(os.path.join(root, member))
            if not target.startswith(os.path.abspath(root) + os.sep):
                raise RuntimeError('更新包包含非法路径')
        zf.extractall(root)
    extracted = os.path.join(root, top_dir)
    if not os.path.isfile(os.path.join(extracted, f'upgrade_to_{top_dir.split("-")[0].lower()}.exe')):
        candidates = [n for n in os.listdir(extracted) if n.lower().startswith('upgrade_to_') and n.lower().endswith('.exe')]
        if not candidates:
            raise RuntimeError('更新包缺少升级器 exe')
    return extracted



def safe_upload_filename(filename):
    """Return a unique, filesystem-safe upload filename."""
    base = secure_filename(filename or 'upload')
    if not base:
        base = 'upload'
    return f"{uuid.uuid4().hex}_{base}"


def uploaded_file_size(file):
    pos = file.stream.tell()
    file.stream.seek(0, os.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(pos)
    return size


def remove_file_safely(filepath):
    try:
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
    except OSError:
        pass


def remove_tree_contents(folder):
    removed = 0
    failed = []
    if not folder or not os.path.exists(folder):
        return removed, failed
    for root, dirs, files in os.walk(folder, topdown=False):
        for filename in files:
            path = os.path.join(root, filename)
            try:
                os.remove(path)
                removed += 1
            except OSError:
                failed.append(path)
        for dirname in dirs:
            path = os.path.join(root, dirname)
            try:
                os.rmdir(path)
            except OSError:
                pass
    return removed, failed


def upload_dir_path():
    path = os.path.join(os.path.dirname(current_app.config['DATABASE']), 'uploads')
    os.makedirs(path, exist_ok=True)
    return path


def app_base_dir():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def dir_size_bytes(path):
    total = 0
    if not path or not os.path.exists(path):
        return 0
    if os.path.isfile(path):
        try:
            return os.path.getsize(path)
        except OSError:
            return 0
    for root, _, files in os.walk(path):
        for filename in files:
            try:
                total += os.path.getsize(os.path.join(root, filename))
            except OSError:
                pass
    return total


def report_assets_dir():
    static_dir = os.path.join(os.path.dirname(os.path.dirname(current_app.config['DATABASE'])), 'static')
    path = os.path.join(static_dir, 'report_assets')
    os.makedirs(path, exist_ok=True)
    return path


def source_file_to_disk_name(source_file):
    """Map a stored source_file value back to the physical upload filename."""
    source_file = source_file or ''
    for prefix in ('warehouse_', 'pl_domestic_', 'intl_direct_', 'upload_', 'logistics_', 'warehouse_cost_'):
        if source_file.startswith(prefix):
            return source_file[len(prefix):]
    return source_file


def is_admin_request():
    return bool(getattr(g, 'is_admin', False))


def maintenance_password_ok(password, allow_admin=True):
    if allow_admin and is_admin_request():
        return True
    return (password or '') == 'Q84405995'


def safe_filename_part(value, fallback='all'):
    text = str(value or '').strip()
    if not text:
        text = fallback
    text = text.replace('全部', 'all').replace(' ', '_')
    text = re.sub(r'[\\/:*?"<>|\r\n]+', '_', text)
    text = re.sub(r'_+', '_', text).strip('_.')
    return text[:60] or fallback


def split_request_values(value):
    return [v.strip().upper() for v in re.split(r'[,;，；]+', str(value or '')) if v.strip()]


def normalize_acceptance_status(value):
    text = str(value or '').strip().lower()
    mapping = {
        'verified': 'verified', 'accepted': 'verified', 'accept': 'verified',
        'yes': 'verified', 'y': 'verified', 'true': 'verified', '1': 'verified',
        '已验收': 'verified', '验收': 'verified', '通过': 'verified', '确认': 'verified', '确认验收': 'verified',
        'pending': 'pending', '待验收': 'pending', '待确认': 'pending', '待核实': 'pending', '': '',
        'rejected': 'rejected', 'reject': 'rejected', 'no': 'rejected', 'n': 'rejected', 'false': 'rejected', '0': 'rejected',
        '已拒绝': 'rejected', '拒绝': 'rejected', '异常': 'rejected', '不通过': 'rejected',
    }
    return mapping.get(text, text if text in ('verified', 'pending', 'rejected') else '')


def recalc_invoice_header_status(db, invoice_id):
    rows = db.execute("""
        SELECT COALESCE(NULLIF(acceptance_status, ''), 'pending') AS status
        FROM invoice_items
        WHERE invoice_id = ?
    """, (invoice_id,)).fetchall()
    if not rows:
        return
    statuses = [r['status'] for r in rows]
    if all(s == 'verified' for s in statuses):
        status = 'verified'
    elif all(s == 'rejected' for s in statuses):
        status = 'rejected'
    else:
        status = 'pending'
    db.execute("UPDATE invoice_headers SET status = ? WHERE id = ?", (status, invoice_id))


def _csv_bytes(rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return ('\ufeff' + buf.getvalue()).encode('utf-8')


def _xlsx_response(workbook, filename):
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _workbook_bytes(workbook):
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _delete_upload_source(filename_or_path):
    """Imported Excel/PDF sources are transient; images are managed separately."""
    if not filename_or_path:
        return
    if os.path.isabs(filename_or_path):
        remove_file_safely(filename_or_path)
        return
    remove_file_safely(os.path.join(upload_dir_path(), filename_or_path))


def _norm_header(value):
    return re.sub(r'[\s_/\-（）()]+', '', str(value or '').strip().upper())


def _find_header(ws, required, aliases=None, scan_rows=12):
    aliases = aliases or {}
    wanted = {}
    for key in list(dict.fromkeys(list(required) + list(aliases.keys()))):
        wanted[_norm_header(key)] = key
        for alias in aliases.get(key, []):
            wanted[_norm_header(alias)] = key
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        found = {}
        for col_idx in range(1, ws.max_column + 1):
            label = wanted.get(_norm_header(ws.cell(row_idx, col_idx).value))
            if label and label not in found:
                found[label] = col_idx
        if all(key in found for key in required):
            return row_idx, found
    return None, {}


def _cell(ws, row_idx, col_map, key):
    col = col_map.get(key)
    return ws.cell(row_idx, col).value if col else None


def _clean_cell(value):
    return str(value or '').strip()


def _date_cell(value):
    if value is None or str(value).strip() == '':
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return text[:10] if re.match(r'^\d{4}-\d{2}-\d{2}', text) else text


def _float_cell(value, default=None):
    if value is None or str(value).strip() == '':
        return default
    try:
        return float(str(value).replace(',', '').replace(' ', ''))
    except ValueError:
        return default


def _style_simple_sheet(ws, title=None):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill('solid', fgColor='E2E8F0')
    header_font = Font(color='111827', bold=True)
    thin = Side(style='thin', color='CBD5E1')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = Border(bottom=thin)
    ws.auto_filter.ref = ws.dimensions
    for idx in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, idx).value or '') for row in range(1, min(ws.max_row, 200) + 1)]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max([len(v) for v in values] or [10]) + 2, 10), 34)


def _style_report_sheet(ws, title, subtitle, headers, rows, summary_rows=None):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    summary_rows = summary_rows or []
    col_count = max(len(headers), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(size=16, bold=True, color='1f2937')
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(size=10, color='64748b')
    ws.cell(2, 1).alignment = Alignment(horizontal='center')

    row_idx = 4
    if summary_rows:
        ws.cell(row_idx, 1, '统计总览')
        ws.cell(row_idx, 1).font = Font(bold=True, color='334155')
        row_idx += 1
        for label, value in summary_rows:
            ws.cell(row_idx, 1, label)
            ws.cell(row_idx, 2, value)
            ws.cell(row_idx, 1).font = Font(color='64748b')
            ws.cell(row_idx, 2).font = Font(bold=True, color='0f172a')
            row_idx += 1
        row_idx += 1

    header_row = row_idx
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color='ffffff')
        cell.fill = PatternFill('solid', fgColor='1f2937')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='cbd5e1')
    for r_offset, row in enumerate(rows, 1):
        for col, value in enumerate(row, 1):
            cell = ws.cell(header_row + r_offset, col, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{header_row + len(rows)}"
    for idx, header in enumerate(headers, 1):
        values = [str(header)] + [str(r[idx - 1] if idx - 1 < len(r) and r[idx - 1] is not None else '') for r in rows[:300]]
        width = min(max(max(len(v) for v in values) + 2, 10), 36)
        ws.column_dimensions[get_column_letter(idx)].width = width
    return header_row


@main_bp.route('/api/reports/header-icon', methods=['GET'])
def get_report_header_icon():
    folder = report_assets_dir()
    for ext in ('png', 'jpg', 'jpeg', 'webp', 'gif'):
        filename = f'header_icon.{ext}'
        if os.path.exists(os.path.join(folder, filename)):
            return jsonify({'success': True, 'url': f'/static/report_assets/{filename}'})
    return jsonify({'success': True, 'url': ''})


@main_bp.route('/api/reports/header-icon', methods=['POST'])
def upload_report_header_icon():
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': '请选择 icon 图片'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in {'png', 'jpg', 'jpeg', 'webp', 'gif'}:
        return jsonify({'success': False, 'error': '仅支持 png / jpg / jpeg / webp / gif'}), 400
    folder = report_assets_dir()
    for old_ext in ('png', 'jpg', 'jpeg', 'webp', 'gif'):
        remove_file_safely(os.path.join(folder, f'header_icon.{old_ext}'))
    filename = f'header_icon.{ext}'
    file.save(os.path.join(folder, filename))
    return jsonify({'success': True, 'url': f'/static/report_assets/{filename}'})

# Re-export get_db for convenience
def get_db():
    return db_module.get_db()

@main_bp.teardown_request
def close_db(e=None):
    db_module.close_db(e)

# ========== 运单业务状态计算 ==========

def add_workdays(date_str, days):
    """从 YYYY-MM-DD HH:MM 格式的日期字符串加上 N 个工作日，返回比较用的日期"""
    from datetime import datetime, timedelta
    dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
    added = 0
    while added < days:
        dt += timedelta(days=1)
        if dt.weekday() < 5:  # 0-4 = Mon-Fri
            added += 1
    return dt.strftime('%Y-%m-%d')

def calc_display_status(row):
    """
    根据原始数据 + 手动/自动核实计算业务状态：
    - 已交付：有 actual_delivery_date（手动标记或自动抓取确认）
    - 异常单：manual_status = 'abnormal'
    - 待核实：delivery_date <= 今天，但未确认交付
    - 运输中：delivery_date > 今天，或无 delivery_date 但未超 10 工作日
    - Canceled 保持原样（前端已排除）
    """
    from datetime import datetime
    status = row['shipment_status']
    # 优先：已确认交付（手动或自动）
    actual_dd = row['actual_delivery_date']
    delivery_locked = row['delivery_locked'] if 'delivery_locked' in row.keys() else 0
    if delivery_locked or (actual_dd and actual_dd.strip()):
        return '已交付'
    if status == 'Canceled':
        return 'Canceled'
    # 手动标记的异常单优先
    manual = row['manual_status'] or ''
    if manual == 'abnormal':
        return '异常单'
    if manual in ('auto_in_transport', 'auto_booked'):
        return '运输中'
    delivery_date = row['delivery_date']
    pickup_date = row['pickup_date']
    today_str = datetime.now().strftime('%Y-%m-%d')
    if delivery_date and delivery_date.strip():
        try:
            delivery_dt = delivery_date[:10]
            if delivery_dt <= today_str:
                return '待核实'
            else:
                return '运输中'
        except:
            pass
    if pickup_date and pickup_date.strip():
        try:
            deadline = add_workdays(pickup_date, 10)
            return '异常单' if today_str > deadline else '运输中'
        except:
            return '运输中'
    return '运输中'

def effective_demand_country(row):
    """业务需求国家：优先用手工/导入维护值，缺省回退到收货国家。"""
    try:
        demand = row['demand_country'] if 'demand_country' in row.keys() else ''
    except Exception:
        demand = ''
    try:
        consignee = row['consignee_country_code'] if 'consignee_country_code' in row.keys() else ''
    except Exception:
        consignee = ''
    return (demand or consignee or '').strip().upper()

def dashboard_month_filter(alias=''):
    """Return SQL snippet/params for dashboard-level pickup month filtering."""
    month = (request.args.get('month') or '').strip()
    if not month:
        return '', []
    prefix = f"{alias}." if alias else ''
    return f" AND SUBSTR({prefix}pickup_date, 1, 7) = ?", [month]

# ========== 页面路由 ==========

@main_bp.route('/')
def index():
    vue_index = os.path.join(current_app.static_folder, 'vue', 'index.html')
    if os.path.exists(vue_index):
        return current_app.send_static_file('vue/index.html')
    return current_app.send_static_file('index.html')

# ========== 版本更新 API ==========

@main_bp.route('/api/update/check')
def api_update_check():
    try:
        manifest = fetch_update_manifest()
        latest = manifest['latest_version'].upper()
        current = APP_VERSION.upper()
        return jsonify({
            'success': True,
            'current_version': current,
            'current_time': APP_BUILD_TIME,
            'latest_version': latest,
            'release_time': manifest.get('release_time') or manifest.get('time') or '',
            'update_available': version_number(latest) > version_number(current),
            'sha256': manifest.get('sha256'),
            'package_url': manifest.get('package_url'),
            'notes': manifest.get('notes') or [],
            'download_only': True,
            'security': 'Only HTTPS GET requests to Qiteng GitHub update source or qitengliu.com fallback mirror are used. No logistics, inventory, invoice, image, database, account, or backup data is uploaded.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'current_version': APP_VERSION}), 500


@main_bp.route('/api/update/install', methods=['POST'])
def api_update_install():
    try:
        manifest = fetch_update_manifest()
        latest = manifest['latest_version'].upper()
        if version_number(latest) <= version_number(APP_VERSION):
            return jsonify({'success': False, 'error': '当前已是最新版本', 'current_version': APP_VERSION, 'latest_version': latest}), 400
        zip_path, size, digest = download_update_zip(manifest['package_url'], manifest['sha256'], latest)
        extracted = safe_extract_update_zip(zip_path)
        candidates = [n for n in os.listdir(extracted) if n.lower().startswith('upgrade_to_') and n.lower().endswith('.exe')]
        if not candidates:
            raise RuntimeError('更新包缺少升级器 exe')
        updater = os.path.join(extracted, sorted(candidates)[0])
        if os.name != 'nt':
            return jsonify({
                'success': True,
                'downloaded': True,
                'installed': False,
                'message': '更新包已下载并校验；当前不是 Windows 环境，未启动 EXE 升级器。',
                'zip_path': zip_path,
                'extracted': extracted,
                'sha256': digest,
                'size': size
            })
        creationflags = getattr(subprocess, 'DETACHED_PROCESS', 0) | getattr(subprocess, 'CREATE_NEW_PROCESS_GROUP', 0)
        subprocess.Popen([updater], cwd=extracted, close_fds=True, creationflags=creationflags)
        return jsonify({
            'success': True,
            'downloaded': True,
            'installed': True,
            'latest_version': latest,
            'sha256': digest,
            'size': size,
            'message': '更新器已启动。系统会关闭当前服务并完成升级。'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'current_version': APP_VERSION}), 500

# ========== 物流看板 API ==========

@main_bp.route('/api/dashboard/summary')
def dashboard_summary():
    """总览数据，基于业务状态计算"""
    db = get_db()
    month_sql, params = dashboard_month_filter()
    rows = db.execute(f"""
        SELECT booking_id, delivery_date, pickup_date, shipment_status,
               actual_delivery_date, delivery_locked, manual_status,
               total_pieces, total_weight, total_volume, price_without_vat
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
          {month_sql}
    """, params).fetchall()
    total = len(rows)
    status_counts = {'已交付': 0, '待核实': 0, '运输中': 0, '异常单': 0}
    total_pieces = total_weight = total_volume = total_price = 0
    for r in rows:
        ds = calc_display_status(r)
        status_counts[ds] = status_counts.get(ds, 0) + 1
        total_pieces += r['total_pieces'] or 0
        total_weight += r['total_weight'] or 0
        total_volume += r['total_volume'] or 0
        total_price += r['price_without_vat'] or 0
    latest_logistics_upload = ''
    upload_dir = upload_dir_path()
    upload_rows = db.execute("""
        SELECT DISTINCT source_file
        FROM shipments
        WHERE source_file IS NOT NULL AND source_file != ''
          AND recycled_at IS NULL
    """).fetchall()
    for r in upload_rows:
        disk_name = source_file_to_disk_name(r['source_file'])
        path = os.path.join(upload_dir, disk_name)
        if os.path.exists(path):
            modified_at = datetime.fromtimestamp(os.path.getmtime(path)).strftime('%Y-%m-%d %H:%M:%S')
            if not latest_logistics_upload or modified_at > latest_logistics_upload:
                latest_logistics_upload = modified_at
    return jsonify({
        'total_shipments': total,
        'delivered': status_counts['已交付'],
        'pending_verify': status_counts['待核实'],
        'in_transit': status_counts['运输中'],
        'exception': status_counts['异常单'],
        'total_pieces': total_pieces,
        'total_weight': total_weight,
        'total_volume': total_volume,
        'total_price': total_price,
        'latest_logistics_upload': latest_logistics_upload,
    })

@main_bp.route('/api/dashboard/status_distribution')
def status_distribution():
    """运单状态分布（业务状态：已交付/运输中/异常单）"""
    db = get_db()
    month_sql, params = dashboard_month_filter()
    rows = db.execute(f"""
        SELECT booking_id, delivery_date, pickup_date, shipment_status,
               actual_delivery_date, delivery_locked, manual_status
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
          {month_sql}
    """, params).fetchall()
    counts = {}
    for r in rows:
        ds = calc_display_status(r)
        counts[ds] = counts.get(ds, 0) + 1
    result = [{'shipment_status': k, 'count': v} for k, v in sorted(counts.items(), key=lambda x: -x[1])]
    return jsonify(result)

@main_bp.route('/api/dashboard/type_distribution')
def type_distribution():
    """运单类型分布"""
    db = get_db()
    month_sql, params = dashboard_month_filter()
    type_labels = {
        'warehouse_central': '中央仓发出',
        'warehouse_furniture': '家具仓发出',
        'direct': '调拨',
    }
    rows = db.execute(f"""
        SELECT shipment_type, COUNT(*) as count
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
          {month_sql}
        GROUP BY shipment_type
        ORDER BY count DESC
    """, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['label'] = type_labels.get(d['shipment_type'], d['shipment_type'])
        result.append(d)
    return jsonify(result)

@main_bp.route('/api/dashboard/country_ranking')
def country_ranking():
    """各国发货量排名，含重量和平均交付时长（工作日），按取货/发货月份筛选。"""
    from flask import request
    month = request.args.get('month', '')
    db = get_db()

    demand_expr = "COALESCE(NULLIF(UPPER(TRIM(demand_country)), ''), UPPER(TRIM(consignee_country_code)))"
    where = f"COALESCE(shipment_status, '') != 'Canceled' AND recycled_at IS NULL AND {demand_expr} IS NOT NULL AND {demand_expr} != ''"
    params = []
    if month:
        where += " AND SUBSTR(pickup_date, 1, 7) = ?"
        params.append(month)

    rows = db.execute(f"""
        SELECT {demand_expr} as country, COUNT(*) as count,
               COALESCE(SUM(total_weight), 0) as weight,
               COALESCE(SUM(price_without_vat), 0) as price
        FROM shipments
        WHERE {where}
        GROUP BY {demand_expr}
        ORDER BY count DESC
    """, params).fetchall()

    # 精确工作日计算：在 Python 端处理
    from datetime import datetime, timedelta
    def calc_workdays(start_str, end_str):
        try:
            s = datetime.strptime(start_str[:10], '%Y-%m-%d')
            e = datetime.strptime(end_str[:10], '%Y-%m-%d')
            count = 0
            cur = s
            while cur < e:
                if cur.weekday() < 5:
                    count += 1
                cur += timedelta(days=1)
            return count
        except:
            return None

    # 按国家查询已交付运单的工作日
    country_data = {}
    for r in rows:
        d = dict(r)
        d['avg_delivery_days'] = None
        country_data[d['country']] = d

    # 查询各国已交付运单的 pickup_date 和 actual_delivery_date
    delivered_rows = db.execute(f"""
        SELECT {demand_expr} as demand_country, pickup_date, actual_delivery_date
        FROM shipments
        WHERE {where}
          AND actual_delivery_date IS NOT NULL AND actual_delivery_date != ''
          AND pickup_date IS NOT NULL AND pickup_date != ''
    """, params).fetchall()

    from collections import defaultdict
    country_workdays = defaultdict(list)
    for r in delivered_rows:
        wd = calc_workdays(r['pickup_date'], r['actual_delivery_date'])
        if wd is not None:
            country_workdays[r['demand_country']].append(wd)

    for country, wds in country_workdays.items():
        if country in country_data:
            country_data[country]['avg_delivery_days'] = round(sum(wds) / len(wds), 1) if wds else None

    result = list(country_data.values())

    # 不带 month 参数时，额外返回可用月份列表
    if not month:
        month_rows = db.execute("""
            SELECT DISTINCT SUBSTR(pickup_date, 1, 7) as month
            FROM shipments
            WHERE COALESCE(shipment_status, '') != 'Canceled'
              AND recycled_at IS NULL
              AND pickup_date IS NOT NULL AND pickup_date != ''
            ORDER BY month DESC
        """).fetchall()
        available_months = [r['month'] for r in month_rows]
        return jsonify({'data': result, 'available_months': available_months})

    return jsonify({'data': result, 'available_months': []})

@main_bp.route('/api/dashboard/monthly_trend')
def monthly_trend():
    """月度趋势（按发货日期）"""
    db = get_db()
    rows = db.execute("""
        SELECT
            SUBSTR(pickup_date, 1, 7) as month,
            COUNT(*) as count,
            COALESCE(SUM(total_weight), 0) as weight,
            COALESCE(SUM(total_volume), 0) as volume,
            COALESCE(SUM(price_without_vat), 0) as price
        FROM shipments
        WHERE pickup_date IS NOT NULL AND pickup_date != '' AND COALESCE(shipment_status, '') != 'Canceled' AND recycled_at IS NULL
        GROUP BY SUBSTR(pickup_date, 1, 7)
        ORDER BY month
    """).fetchall()
    return jsonify([dict(r) for r in rows])

@main_bp.route('/api/dashboard/warehouse_distribution')
def warehouse_distribution():
    """仓库发货分布"""
    db = get_db()
    rows = db.execute("""
        SELECT warehouse, COUNT(*) as count,
               COALESCE(SUM(total_weight), 0) as weight
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled' AND recycled_at IS NULL AND warehouse IS NOT NULL AND warehouse != ''
        GROUP BY warehouse
        ORDER BY count DESC
    """).fetchall()
    return jsonify([dict(r) for r in rows])

# ========== 运单列表 API ==========

@main_bp.route('/api/shipments')
def shipments_list():
    """运单列表，支持筛选，支持业务状态(ds)和仓库(warehouse)筛选"""
    from flask import request
    db = get_db()

    ds = request.args.get('ds', '')  # 业务状态：已交付/运输中/异常单
    verify = request.args.get('verify', '')  # 验收状态：已验收/待验收/未关联
    type_ = request.args.get('type', '')
    country = request.args.get('country', '')
    warehouse = request.args.get('warehouse', '')
    month = request.args.get('month', '')
    pickup_month = request.args.get('pickup_month', '') or month
    actual_delivery_month = request.args.get('actual_delivery_month', '')
    sort_by = request.args.get('sort_by', '')
    sort_dir = request.args.get('sort_dir', 'desc')
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    where_clauses = ["COALESCE(shipment_status, '') != 'Canceled'", "recycled_at IS NULL"]
    params = []

    if type_:
        where_clauses.append('shipment_type = ?')
        params.append(type_)
    if country:
        where_clauses.append("COALESCE(NULLIF(UPPER(TRIM(demand_country)), ''), UPPER(TRIM(consignee_country_code))) = ?")
        params.append(country.strip().upper())
    if warehouse:
        where_clauses.append('warehouse = ?')
        params.append(warehouse)
    if pickup_month == 'annual':
        where_clauses.append("SUBSTR(pickup_date, 1, 4) = ?")
        params.append(str(datetime.now().year))
    elif pickup_month.startswith('year:') and re.match(r'^year:\d{4}$', pickup_month):
        where_clauses.append("SUBSTR(pickup_date, 1, 4) = ?")
        params.append(pickup_month.split(':', 1)[1])
    elif pickup_month:
        where_clauses.append("SUBSTR(pickup_date, 1, 7) = ?")
        params.append(pickup_month)
    if actual_delivery_month:
        where_clauses.append("SUBSTR(actual_delivery_date, 1, 7) = ?")
        params.append(actual_delivery_month)
    if search:
        where_clauses.append('(booking_id LIKE ? OR stt_number LIKE ? OR consignee_name LIKE ? OR consignee_city LIKE ?)')
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])

    where_sql = ' AND '.join(where_clauses)

    # 先查出全部匹配行，再按业务状态过滤（数据量不大，195条足够快）
    all_rows = db.execute(f"SELECT * FROM shipments WHERE {where_sql} ORDER BY pickup_date DESC", params).fetchall()

    if ds:
        all_rows = [r for r in all_rows if calc_display_status(r) == ds]

    # 查询所有已验收的发票STT号，用于标注验收状态
    verified_stts = set()
    pending_stts = set()
    verified_amounts = {}
    verified_invoices = {}
    try:
        inv_rows = db.execute("""SELECT i.stt_number,
                                         COALESCE(NULLIF(i.acceptance_status, ''), h.status) AS status,
                                         h.invoice_number,
                                         COALESCE(i.accepted_amount, i.net_amount, 0) AS net_amount
                                  FROM invoice_items i
                                  JOIN invoice_headers h ON i.invoice_id = h.id""").fetchall()
        for ir in inv_rows:
            if ir['status'] == 'verified':
                verified_stts.add(ir['stt_number'])
                verified_amounts[ir['stt_number']] = verified_amounts.get(ir['stt_number'], 0) + float(ir['net_amount'] or 0)
                verified_invoices.setdefault(ir['stt_number'], set()).add(ir['invoice_number'])
            elif ir['status'] == 'pending':
                pending_stts.add(ir['stt_number'])
    except:
        pass

    # 按验收状态筛选
    if verify:
        def get_verify_status(r):
            stt = r['stt_number'] if 'stt_number' in r.keys() else ''
            if stt in verified_stts:
                return '已验收'
            elif stt in pending_stts:
                return '待验收'
            else:
                return '未关联'
        all_rows = [r for r in all_rows if get_verify_status(r) == verify]

    type_labels = {
        'warehouse_central': '中央仓发出',
        'warehouse_furniture': '家具仓发出',
        'direct': '调拨',
    }

    def shipment_verify_status(row):
        stt = row['stt_number'] if 'stt_number' in row.keys() else ''
        if stt in verified_stts:
            return '已验收'
        if stt in pending_stts:
            return '待验收'
        return '未关联'

    def shipment_sort_value(row, field):
        keys = row.keys()
        if field == 'stt_number':
            return (row['stt_number'] if 'stt_number' in keys else '') or (row['booking_id'] if 'booking_id' in keys else '')
        if field == 'display_status':
            return calc_display_status(row)
        if field == 'shipment_type':
            raw = row['shipment_type'] if 'shipment_type' in keys else ''
            return type_labels.get(raw, raw or '')
        if field == 'country':
            return effective_demand_country(row)
        if field == 'consignee_city':
            return row['consignee_city'] if 'consignee_city' in keys else ''
        if field in ('pickup_date', 'delivery_date', 'actual_delivery_date'):
            return row[field] if field in keys else ''
        if field == 'total_weight':
            try:
                return float(row['total_weight'] or 0)
            except Exception:
                return 0
        if field == 'verify_status':
            return shipment_verify_status(row)
        if field == 'accepted_amount':
            stt = row['stt_number'] if 'stt_number' in keys else ''
            return verified_amounts.get(stt, 0)
        return ''

    sortable_fields = {
        'stt_number', 'display_status', 'shipment_type', 'country',
        'consignee_city', 'pickup_date', 'delivery_date', 'actual_delivery_date',
        'total_weight', 'verify_status', 'accepted_amount'
    }
    if sort_by in sortable_fields:
        reverse = str(sort_dir).lower() != 'asc'
        if sort_by == 'accepted_amount':
            all_rows = sorted(
                all_rows,
                key=lambda r: (
                    0 if (r['stt_number'] if 'stt_number' in r.keys() else '') in verified_amounts else 1,
                    shipment_sort_value(r, sort_by)
                ),
                reverse=False
            )
            if reverse:
                verified = [r for r in all_rows if (r['stt_number'] if 'stt_number' in r.keys() else '') in verified_amounts]
                missing = [r for r in all_rows if (r['stt_number'] if 'stt_number' in r.keys() else '') not in verified_amounts]
                all_rows = sorted(verified, key=lambda r: shipment_sort_value(r, sort_by), reverse=True) + missing
        else:
            non_empty = [r for r in all_rows if shipment_sort_value(r, sort_by) not in ('', None)]
            empty = [r for r in all_rows if shipment_sort_value(r, sort_by) in ('', None)]
            all_rows = sorted(non_empty, key=lambda r: shipment_sort_value(r, sort_by), reverse=reverse) + empty

    total = len(all_rows)
    # 分页
    start = (page - 1) * per_page
    page_rows = all_rows[start:start + per_page]

    shipments = []
    for r in page_rows:
        d = dict(r)
        d['shipment_type_label'] = type_labels.get(d['shipment_type'], d['shipment_type'])
        d['display_status'] = calc_display_status(r)
        # 验收状态
        stt = d.get('stt_number', '')
        if stt in verified_stts:
            d['verify_status'] = '已验收'
            d['accepted_amount'] = round(verified_amounts.get(stt, 0), 2)
            d['accepted_invoices'] = sorted(verified_invoices.get(stt, []))
        elif stt in pending_stts:
            d['verify_status'] = '待验收'
            d['accepted_amount'] = None
            d['accepted_invoices'] = []
        else:
            d['verify_status'] = '未关联'
            d['accepted_amount'] = None
            d['accepted_invoices'] = []
        shipments.append(d)

    return jsonify({
        'shipments': shipments,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page if total else 1
    })


@main_bp.route('/api/uploads')
def list_uploaded_files():
    """List uploaded source files and their linked business records."""
    db = get_db()
    upload_dir = upload_dir_path()
    file_map = {}

    def ensure_item(disk_name):
        disk_name = os.path.basename(disk_name or '')
        if not disk_name:
            return None
        path = os.path.join(upload_dir, disk_name)
        item = file_map.setdefault(disk_name, {
            'filename': disk_name,
            'exists': os.path.exists(path),
            'size': os.path.getsize(path) if os.path.exists(path) else 0,
            'modified_at': datetime.fromtimestamp(os.path.getmtime(path)).strftime('%Y-%m-%d %H:%M:%S') if os.path.exists(path) else '',
            'shipment_records': 0,
            'invoice_records': 0,
            'invoice_number': '',
            'invoice_status': '',
            'types': set(),
        })
        return item

    for filename in os.listdir(upload_dir):
        path = os.path.join(upload_dir, filename)
        if os.path.isfile(path):
            item = ensure_item(filename)
            if item:
                item['types'].add('file')

    shipment_rows = db.execute("""
        SELECT source_file, COUNT(*) AS cnt
        FROM shipments
        WHERE source_file IS NOT NULL AND source_file != ''
        GROUP BY source_file
    """).fetchall()
    for r in shipment_rows:
        disk_name = source_file_to_disk_name(r['source_file'])
        item = ensure_item(disk_name)
        if not item:
            continue
        item['shipment_records'] += r['cnt'] or 0
        item['types'].add('logistics')

    invoice_rows = db.execute("""
        SELECT source_file, invoice_number, status
        FROM invoice_headers
        WHERE source_file IS NOT NULL AND source_file != ''
    """).fetchall()
    for r in invoice_rows:
        disk_name = source_file_to_disk_name(r['source_file'])
        item = ensure_item(disk_name)
        if not item:
            continue
        item['invoice_records'] += 1
        item['invoice_number'] = r['invoice_number'] or item['invoice_number']
        item['invoice_status'] = r['status'] or item['invoice_status']
        item['types'].add('invoice')

    records = []
    for item in file_map.values():
        item['types'] = sorted(item['types'])
        records.append(item)
    records.sort(key=lambda x: (not x['exists'], x['modified_at']), reverse=True)
    return jsonify({'files': records})


@main_bp.route('/api/uploads/<path:filename>', methods=['DELETE'])
def delete_uploaded_file(filename):
    data = request.get_json(silent=True) or {}
    return _delete_uploaded_file_by_name(filename, data)


@main_bp.route('/api/uploads/delete', methods=['POST'])
def delete_uploaded_file_json():
    data = request.get_json(silent=True) or {}
    return _delete_uploaded_file_by_name(data.get('filename', ''), data)


def _delete_uploaded_file_by_name(filename, data):
    if not maintenance_password_ok(data.get('password'), allow_admin=True):
        return jsonify({'error': '删除密码错误'}), 403
    safe_name = os.path.basename(filename or '')
    if not safe_name or safe_name in ('.', '..'):
        return jsonify({'error': '文件名无效'}), 400
    upload_dir = upload_dir_path()
    path = os.path.abspath(os.path.join(upload_dir, safe_name))
    if not path.startswith(os.path.abspath(upload_dir) + os.sep):
        return jsonify({'error': '文件路径无效'}), 400
    if not os.path.exists(path):
        return jsonify({'error': '文件不存在或已删除'}), 404
    remove_file_safely(path)
    return jsonify({'success': True, 'filename': safe_name})


@main_bp.route('/api/maintenance/slim', methods=['POST'])
def slim_local_package():
    data = request.get_json(silent=True) or {}
    if not maintenance_password_ok(data.get('password'), allow_admin=False):
        return jsonify({'error': '瘦身密码错误'}), 403

    clear_inventory = bool(data.get('clear_inventory', False))
    clear_images = bool(data.get('clear_images', False))
    clear_non_business_uploads = bool(data.get('clear_non_business_uploads', False))
    clear_logistics_uploads = bool(data.get('clear_logistics_uploads', False))
    clear_invoice_uploads = bool(data.get('clear_invoice_uploads', False))
    clear_cache = bool(data.get('clear_cache', False))
    clear_update_downloads = bool(data.get('clear_update_downloads', False))
    clear_backups = bool(data.get('clear_backups', False))
    base_dir = app_base_dir()
    package_root = software_root_dir()
    size_before = dir_size_bytes(package_root)

    result = {
        'success': True,
        'kept': ['物流数据', '发票/验收数据', '运单手动状态', '异常/回收站记录'],
        'deleted': {},
        'failed': [],
        'backups': [],
        'size_before': size_before,
    }

    if clear_inventory:
        backup = db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'before_slim_inventory')
        if backup:
            result['backups'].append(backup)
        wms = sqlite3.connect(current_app.config['WMS_DATABASE'])
        try:
            cur = wms.cursor()
            tables = [
                'delivery_order_items',
                'delivery_orders',
                'inventory_transactions',
                'inventory_movements',
                'warehouse_inventory',
                'product_master',
                'sku_master',
            ]
            table_counts = {}
            for table in tables:
                exists = cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
                if not exists:
                    continue
                count = cur.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                cur.execute(f"DELETE FROM {table}")
                table_counts[table] = count
            wms.commit()
            cur.execute("VACUUM")
            result['deleted']['inventory_records'] = table_counts
        finally:
            wms.close()

    if clear_images:
        image_folder = current_app.config['UPLOAD_FOLDER']
        deleted_images, failed_images = remove_tree_contents(image_folder)
        wms = sqlite3.connect(current_app.config['WMS_DATABASE'])
        try:
            cur = wms.cursor()
            cur.execute("UPDATE sku_master SET image_path = NULL, updated_at = CURRENT_TIMESTAMP WHERE COALESCE(image_path, '') <> ''")
            result['deleted']['image_mappings'] = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
            wms.commit()
        finally:
            wms.close()
        result['deleted']['image_files'] = deleted_images
        result['failed'].extend(failed_images)

    if clear_non_business_uploads or clear_logistics_uploads or clear_invoice_uploads:
        main_db = get_db()
        logistics_files = set()
        for row in main_db.execute("""
            SELECT DISTINCT source_file FROM shipments
            WHERE source_file IS NOT NULL AND TRIM(source_file) != ''
        """).fetchall():
            logistics_files.add(row['source_file'])
            logistics_files.add(source_file_to_disk_name(row['source_file']))
        invoice_files = set()
        for row in main_db.execute("""
            SELECT DISTINCT source_file FROM invoice_headers
            WHERE source_file IS NOT NULL AND TRIM(source_file) != ''
        """).fetchall():
            invoice_files.add(row['source_file'])
            invoice_files.add(source_file_to_disk_name(row['source_file']))

        upload_dir = upload_dir_path()
        deleted_uploads = 0
        kept_uploads = 0
        deleted_logistics = 0
        deleted_invoice = 0
        deleted_other = 0
        failed_uploads = []
        for name in os.listdir(upload_dir):
            path = os.path.abspath(os.path.join(upload_dir, name))
            if not path.startswith(os.path.abspath(upload_dir) + os.sep) or not os.path.isfile(path):
                continue
            is_logistics = name in logistics_files
            is_invoice = name in invoice_files
            should_delete = (
                (is_logistics and clear_logistics_uploads)
                or (is_invoice and clear_invoice_uploads)
                or (not is_logistics and not is_invoice and clear_non_business_uploads)
            )
            if not should_delete:
                kept_uploads += 1
                continue
            try:
                os.remove(path)
                deleted_uploads += 1
                if is_logistics:
                    deleted_logistics += 1
                elif is_invoice:
                    deleted_invoice += 1
                else:
                    deleted_other += 1
            except OSError:
                failed_uploads.append(path)
        result['deleted']['upload_files_total'] = deleted_uploads
        result['deleted']['logistics_upload_files'] = deleted_logistics
        result['deleted']['invoice_upload_files'] = deleted_invoice
        result['deleted']['other_upload_files'] = deleted_other
        result['deleted']['kept_upload_files'] = kept_uploads
        result['failed'].extend(failed_uploads)

    if clear_cache:
        base_dir = app_base_dir()
        cache_targets = [
            os.path.join(base_dir, 'app', '__pycache__'),
            os.path.join(base_dir, '__pycache__'),
            os.path.join(base_dir, 'app', 'data', 'tmp'),
        ]
        removed = 0
        failed = []
        for target in cache_targets:
            count, bad = remove_tree_contents(target)
            removed += count
            failed.extend(bad)
            try:
                if os.path.isdir(target) and not os.listdir(target):
                    os.rmdir(target)
            except OSError:
                pass
        result['deleted']['cache_files'] = removed
        result['failed'].extend(failed)

    if clear_update_downloads:
        update_targets = [
            os.path.join(package_root, '_downloaded_updates'),
            os.path.join(package_root, '_update_staging'),
        ]
        removed = 0
        failed = []
        removed_dirs = 0
        current_version_num = version_number(APP_VERSION)
        for target in update_targets:
            count, bad = remove_tree_contents(target)
            removed += count
            failed.extend(bad)
            try:
                if os.path.isdir(target) and not os.listdir(target):
                    os.rmdir(target)
                    removed_dirs += 1
            except OSError:
                pass
        for name in os.listdir(package_root):
            lower = name.lower()
            path = os.path.join(package_root, name)
            file_version = version_number(name)
            is_update_extract = (
                os.path.isdir(path)
                and lower.startswith('vn')
                and ('patch' in lower or 'update' in lower)
                and (not file_version or file_version <= current_version_num)
            )
            if not is_update_extract:
                continue
            count, bad = remove_tree_contents(path)
            removed += count
            failed.extend(bad)
            try:
                os.rmdir(path)
                removed_dirs += 1
            except OSError:
                pass
        result['deleted']['downloaded_update_files'] = removed
        result['deleted']['downloaded_update_dirs'] = removed_dirs
        result['failed'].extend(failed)

    if clear_backups:
        backup_targets = [
            os.path.join(base_dir, 'backups'),
            os.path.join(base_dir, 'app', 'data', 'backups'),
        ]
        removed = 0
        failed = []
        for target in backup_targets:
            count, bad = remove_tree_contents(target)
            removed += count
            failed.extend(bad)
            try:
                if os.path.isdir(target) and not os.listdir(target):
                    os.rmdir(target)
            except OSError:
                pass

        old_patch_files = 0
        old_patch_dirs = 0
        current_version_label = APP_VERSION.upper()
        current_version_num = version_number(current_version_label)
        current_keep = {
            'start.bat',
            'rollback_to_previous.bat',
            'stop_system.bat',
            f'EUCWH-{current_version_label}.exe',
            f'start_{current_version_label.lower()}.bat',
            f'debug_start_{current_version_label.lower()}.bat',
            f'launch_hidden_{current_version_label.lower()}.vbs',
            f'PATCH_NOTES_{current_version_label}.txt',
            f'{current_version_label}_GUIDE.txt',
        }
        current_keep_lower = {item.lower() for item in current_keep}
        scan_dirs = [base_dir]
        parent_dir = os.path.dirname(base_dir)
        if os.path.basename(base_dir).lower() == '_system' and parent_dir and parent_dir not in scan_dirs:
            scan_dirs.append(parent_dir)
        if package_root and package_root not in scan_dirs:
            scan_dirs.append(package_root)
        for scan_dir in scan_dirs:
            if not os.path.isdir(scan_dir):
                continue
            for name in os.listdir(scan_dir):
                lower = name.lower()
                path = os.path.join(scan_dir, name)
                file_version = version_number(name)
                is_current_launcher = file_version == current_version_num or lower in current_keep_lower
                is_old_version_file = file_version and file_version < current_version_num
                should_delete_file = (
                    (
                        (re.fullmatch(r'(start|stop)_vn\d+\.bat', lower) is not None)
                        or (re.fullmatch(r'launch(_vn\d+)?_hidden(_vn\d+)?\.vbs', lower) is not None)
                        or (re.fullmatch(r'patch_notes_vn\d+\.txt', lower) is not None)
                        or (re.fullmatch(r'vn\d+_applied\.txt', lower) is not None)
                        or (re.fullmatch(r'vn\d+_(guide|操作指南)\.txt', lower) is not None)
                        or (re.fullmatch(r'.*vn\d+.*\.exe', lower) is not None)
                        or (re.fullmatch(r'(debug启动|debug_start)_vn\d+\.bat', lower) is not None)
                        or (re.fullmatch(r'upgrade_to_vn\d+\.exe', lower) is not None)
                    )
                    and is_old_version_file
                    and not is_current_launcher
                )
                should_delete_dir = (
                    lower.startswith('backup_before_vn')
                    or lower in {'_downloaded_updates', '_update_staging'}
                    or (lower.startswith('vn') and ('patch' in lower or 'update' in lower))
                )
                if os.path.isfile(path) and should_delete_file:
                    try:
                        os.remove(path)
                        old_patch_files += 1
                    except OSError:
                        failed.append(path)
                elif os.path.isdir(path) and should_delete_dir:
                    count, bad = remove_tree_contents(path)
                    removed += count
                    failed.extend(bad)
                    try:
                        os.rmdir(path)
                        old_patch_dirs += 1
                    except OSError:
                        pass
        result['deleted']['backup_files'] = removed
        result['deleted']['old_patch_files'] = old_patch_files
        result['deleted']['old_patch_dirs'] = old_patch_dirs
        result['failed'].extend(failed)

    result['size_after'] = dir_size_bytes(package_root)
    result['freed_bytes'] = max(0, result['size_before'] - result['size_after'])
    return jsonify(result)


@main_bp.route('/api/export/full')
def export_full_package():
    """导出库存、物流、验收状态和库存图片到一个离线 ZIP。"""
    scope = request.args.get('scope', 'all')
    warehouse = request.args.get('warehouse', '')
    export_all = scope == 'all' or not warehouse

    wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
    wms_conn.row_factory = sqlite3.Row

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    label = 'all_warehouses' if export_all else re.sub(r'[^A-Za-z0-9_-]+', '_', warehouse).strip('_')
    zip_path = os.path.join(tempfile.gettempdir(), f'VN62_完整包_{safe_filename_part(label, "all")}_{ts}.zip')

    try:
        image_rows = wms_conn.execute(f"""
            SELECT DISTINCT COALESCE(pimg.image_path, sm.image_path) AS image_path
            FROM warehouse_inventory wi
            LEFT JOIN product_master pm ON wi.product_number = pm.product_number
            LEFT JOIN sku_master sm ON COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code) = sm.bom_code
                AND NOT (
                    REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                    OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                )
            LEFT JOIN sku_master pimg ON (
                REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
            ) AND wi.product_number = pimg.bom_code
            WHERE COALESCE(COALESCE(pimg.image_path, sm.image_path), '') <> ''
        """).fetchall()
        image_files = [r['image_path'] for r in image_rows if r['image_path']]

        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            backup_wb = build_core_backup_workbook()
            zf.writestr('core_backup_no_images.xlsx', _workbook_bytes(backup_wb))
            upload_dir = current_app.config['UPLOAD_FOLDER']
            for image_path in sorted(set(image_files)):
                source = os.path.join(upload_dir, image_path)
                if os.path.exists(source) and os.path.isfile(source):
                    zf.write(source, f'images/{os.path.basename(image_path)}')

        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=os.path.basename(zip_path)
        )
    finally:
        wms_conn.close()


def build_core_backup_workbook():
    from openpyxl import Workbook
    main_db = get_db()
    wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
    wms_conn.row_factory = sqlite3.Row
    wb = Workbook()

    def add_sheet(title, headers, rows):
        ws = wb.active if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active['A1'].value is None else wb.create_sheet()
        ws.title = title[:31]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(1, col_idx, str(header))
        for row in rows:
            ws.append([row.get(h, '') if isinstance(row, dict) else row[h] if h in row.keys() else '' for h in headers])
        _style_simple_sheet(ws)
        return ws

    try:
        shipment_headers = ['booking_id', 'stt_number', 'waybill_no', 'shipment_status', 'manual_status', 'shipment_type', 'warehouse', 'demand_country', 'demand_country_source', 'consignee_country_code', 'consignee_name', 'consignee_city', 'pickup_date', 'delivery_date', 'actual_delivery_date', 'delivery_locked', 'total_pieces', 'total_weight', 'total_volume', 'price_without_vat', 'product', 'source_file', 'recycled_at', 'recycle_reason', 'import_time']
        add_sheet('物流底表', shipment_headers, main_db.execute(f"SELECT {', '.join(shipment_headers)} FROM shipments ORDER BY pickup_date DESC, booking_id").fetchall())

        invoice_headers = ['invoice_number', 'invoice_date', 'invoice_type', 'currency', 'status', 'stt_number', 'net_amount', 'ref_date', 'matched_booking_id', 'accepted_amount', 'acceptance_status', 'exception_reason', 'acceptance_remark']
        add_sheet('发票验收底表', invoice_headers, main_db.execute("""
            SELECT h.invoice_number, h.invoice_date, h.invoice_type, h.currency, h.status,
                   i.stt_number, i.net_amount, i.ref_date, i.matched_booking_id,
                   i.accepted_amount, i.acceptance_status, i.exception_reason, i.acceptance_remark
            FROM invoice_items i JOIN invoice_headers h ON i.invoice_id = h.id
            ORDER BY h.invoice_date DESC, h.invoice_number, i.id
        """).fetchall())

        inventory_headers = ['warehouse', 'bom_code', 'product_number', 'product_description', 'category', 'category_2', 'instruction', 'inventory', 'unit_price', 'ttl_amount', 'last_inbound_date', 'last_outbound_date', 'dos_threshold', 'idle_threshold', 'comment', 'updated_at']
        add_sheet('库存底表', inventory_headers, wms_conn.execute("""
            SELECT wi.warehouse, wi.bom_code, wi.product_number, pm.product_description, pm.category, pm.category_2,
                   wi.instruction, wi.inventory, pm.unit_price, COALESCE(pm.unit_price,0) * COALESCE(wi.inventory,0) AS ttl_amount,
                   wi.last_inbound_date, wi.last_outbound_date, pm.dos_threshold, pm.idle_threshold, wi.comment, wi.updated_at
            FROM warehouse_inventory wi LEFT JOIN product_master pm ON wi.product_number = pm.product_number
            ORDER BY wi.warehouse, wi.product_number
        """).fetchall())

        sku_headers = ['product_number', 'bom_code', 'product_description', 'category', 'category_2', 'unit_price', 'last_inbound_date', 'last_outbound_date', 'remark', 'dos_threshold', 'idle_threshold', 'created_at', 'updated_at']
        add_sheet('SKU主数据', sku_headers, wms_conn.execute(f"SELECT {', '.join(sku_headers)} FROM product_master ORDER BY product_number").fetchall())

        movement_headers = ['movement_type', 'sheet_name', 'product_number', 'movement_date', 'operation_id', 'operation_prefix', 'operation_target', 'order_date', 'quantity', 'uom', 'source_file', 'created_at']
        add_sheet('出入库流水', movement_headers, wms_conn.execute(f"SELECT {', '.join(movement_headers)} FROM inventory_movements ORDER BY movement_date DESC, product_number").fetchall())

        meta = [{'key': 'version', 'value': 'VN62'}, {'key': 'export_time', 'value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}, {'key': 'images', 'value': 'not included'}]
        add_sheet('版本信息', ['key', 'value'], meta)
        return wb
    finally:
        wms_conn.close()


def build_inventory_calibration_workbook(include_images=False):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    headers = [
        'Bom Code', 'Product Number', 'Product Description', 'Category', 'Category 2', 'Instruction',
        'Inventory (Pcs)', 'Unit Price (USD)', 'TTL Amount (USD)', '最近入库时间 Last Inbound date',
        '最近出库时间 Last Outbound date', '库龄 Dos (Days of Supply)', '库存静置时长 Days without Stock',
        '备注 comment', 'SKU条目创建时间', '图片 Photo',
    ]
    sheet_map = [('Lodz', 'Lodz warehouse'), ('Bydgoszcz', 'Bydgoszcz warehouse')]
    wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
    wms_conn.row_factory = sqlite3.Row
    image_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), '..', 'static', 'images')
    wb = Workbook()
    try:
        for idx, (sheet_title, warehouse) in enumerate(sheet_map):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = sheet_title
            ws.append(headers)
            rows = wms_conn.execute("""
                SELECT wi.warehouse,
                       COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code, sm.bom_code) AS bom_code,
                       wi.product_number,
                       COALESCE(pm.product_description, sm.product_description) AS product_description,
                       COALESCE(pm.category, sm.category) AS category,
                       COALESCE(pm.category_2, sm.category_2) AS category_2,
                       wi.instruction, wi.inventory,
                       COALESCE(pm.unit_price, sm.unit_price, 0) AS unit_price,
                       COALESCE(wi.last_inbound_date, pm.last_inbound_date, DATE(pm.created_at)) AS last_inbound_date,
                       COALESCE(wi.last_outbound_date, pm.last_outbound_date) AS last_outbound_date,
                       wi.comment, DATE(pm.created_at) AS sku_created_at,
                       COALESCE(pimg.image_path, sm.image_path) AS image_path
                FROM warehouse_inventory wi
                LEFT JOIN product_master pm ON wi.product_number = pm.product_number
                LEFT JOIN sku_master sm ON COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code) = sm.bom_code
                    AND NOT (
                        REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                        OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                    )
                LEFT JOIN sku_master pimg ON (
                    REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                    OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                ) AND wi.product_number = pimg.bom_code
                WHERE wi.warehouse = ?
                ORDER BY wi.product_number
            """, (warehouse,)).fetchall()
            for r in rows:
                inventory = float(r['inventory'] or 0)
                unit_price = float(r['unit_price'] or 0)
                last_inbound = r['last_inbound_date'] or ''
                last_outbound = r['last_outbound_date'] or ''
                dos = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_inbound,)).fetchone()[0] if last_inbound else ''
                idle = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_outbound,)).fetchone()[0] if last_outbound else ''
                ws.append([
                    r['bom_code'] or '', r['product_number'] or '', r['product_description'] or '',
                    r['category'] or '', r['category_2'] or '', r['instruction'] or '', int(inventory),
                    unit_price, inventory * unit_price, last_inbound, last_outbound, dos or '', idle or '',
                    r['comment'] or '', r['sku_created_at'] or '', '' if not r['image_path'] else ('' if include_images else r['image_path'])
                ])
                row_idx = ws.max_row
                ws.row_dimensions[row_idx].height = 58 if include_images else 22
                if include_images and r['image_path']:
                    img_path = os.path.abspath(os.path.join(image_dir, r['image_path']))
                    if img_path.startswith(os.path.abspath(image_dir) + os.sep) and os.path.exists(img_path):
                        try:
                            img = ExcelImage(img_path)
                            img.width = 72
                            img.height = 52
                            ws.add_image(img, f'P{row_idx}')
                        except Exception:
                            pass
            _style_simple_sheet(ws)
        return wb
    finally:
        wms_conn.close()


@main_bp.route('/api/export/core-backup')
def export_core_backup():
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return _xlsx_response(build_core_backup_workbook(), f'VN62_core_backup_no_images_{ts}.xlsx')


@main_bp.route('/api/import/core-backup', methods=['POST'])
def import_core_backup():
    """Import the lightweight core backup workbook exported from the full package."""
    from openpyxl import load_workbook
    if request.form.get('password', '') != 'Q84405995':
        return jsonify({'error': '密码错误，请重新输入'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xls', '.zip')):
        return jsonify({'error': 'Only Excel or full export ZIP files supported'}), 400

    db_module.backup_sqlite_db(current_app.config['DATABASE'], 'before_core_backup_import_main')
    db_module.backup_sqlite_db(current_app.config['WMS_DATABASE'], 'before_core_backup_import_wms')

    def rows_from_sheet(wb, title):
        if title not in wb.sheetnames:
            return []
        ws = wb[title]
        headers = [str(c.value or '').strip() for c in ws[1]]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, '') for v in values):
                continue
            rows.append({headers[i]: values[i] if i < len(values) else '' for i in range(len(headers)) if headers[i]})
        return rows

    def table_columns(conn, table):
        return {r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}

    def backup_is_local_bom(value):
        text = str(value or '').strip().upper()
        compact = re.sub(r'[\s/_-]+', '', text)
        return text in {'/', '／'} or compact in {'本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', 'NOBOM', 'NOBOMCODE'}

    def normalize_backup_bom(value):
        text = str(value or '').strip()
        return '/' if not text or backup_is_local_bom(text) else text

    def upsert_by_key(conn, table, key_cols, row):
        cols = table_columns(conn, table)
        data = {k: v for k, v in row.items() if k in cols and k != 'id'}
        if not all(str(data.get(k) or '').strip() for k in key_cols):
            return 'skipped'
        where = ' AND '.join([f"{k}=?" for k in key_cols])
        keys = [data[k] for k in key_cols]
        exists = conn.execute(f"SELECT 1 FROM {table} WHERE {where} LIMIT 1", keys).fetchone()
        if exists:
            update_cols = [c for c in data.keys() if c not in key_cols]
            if update_cols:
                assignments = []
                values = []
                for c in update_cols:
                    if table == 'product_master' and c == 'unit_price':
                        assignments.append("unit_price = CASE WHEN COALESCE(?, 0) > 0 THEN ? ELSE unit_price END")
                        values.extend([data[c], data[c]])
                    elif table == 'shipments' and c in ('actual_delivery_date', 'manual_status'):
                        assignments.append(f"{c} = COALESCE(NULLIF(?, ''), {c})")
                        values.append(data[c])
                    elif table == 'shipments' and c == 'delivery_locked':
                        assignments.append("delivery_locked = CASE WHEN COALESCE(?, 0) IN (1, '1', 'Y', 'YES', 'TRUE', '是') THEN 1 ELSE delivery_locked END")
                        values.append(data[c])
                    else:
                        assignments.append(c + '=?')
                        values.append(data[c])
                conn.execute(f"UPDATE {table} SET {', '.join(assignments)} WHERE {where}",
                             values + keys)
            return 'updated'
        insert_cols = list(data.keys())
        conn.execute(f"INSERT INTO {table} ({', '.join(insert_cols)}) VALUES ({', '.join(['?'] * len(insert_cols))})",
                     [data[c] for c in insert_cols])
        return 'created'

    extracted_zip_images = []
    try:
        if filename.lower().endswith('.zip'):
            raw = file.read()
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                candidate = None
                for name in zf.namelist():
                    if os.path.basename(name).lower() == 'core_backup_no_images.xlsx':
                        candidate = name
                        break
                if not candidate:
                    return jsonify({'error': 'ZIP 中未找到 core_backup_no_images.xlsx'}), 400
                for name in zf.namelist():
                    base = os.path.basename(name)
                    ext = base.rsplit('.', 1)[-1].lower() if '.' in base else ''
                    if not name.lower().startswith('images/') or not base or ext not in {'png', 'jpg', 'jpeg', 'gif', 'webp'}:
                        continue
                    extracted_zip_images.append((base, zf.read(name)))
                wb = load_workbook(io.BytesIO(zf.read(candidate)), data_only=True)
        else:
            wb = load_workbook(file, data_only=True)
        main_db = get_db()
        wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
        wms_conn.row_factory = sqlite3.Row
        report = {'shipments': {}, 'invoices': {}, 'inventory': {}, 'sku': {}, 'movements': {}}

        for row in rows_from_sheet(wb, '物流底表'):
            row = dict(row)
            if not str(row.get('shipment_status') or '').strip():
                row['shipment_status'] = 'Active'
            result = upsert_by_key(main_db, 'shipments', ['booking_id'], row)
            report['shipments'][result] = report['shipments'].get(result, 0) + 1

        invoice_cache = {}
        for row in rows_from_sheet(wb, '发票验收底表'):
            invoice_number = str(row.get('invoice_number') or '').strip()
            stt = str(row.get('stt_number') or '').strip()
            if not invoice_number or not stt:
                report['invoices']['skipped'] = report['invoices'].get('skipped', 0) + 1
                continue
            if invoice_number not in invoice_cache:
                main_db.execute("""
                    INSERT INTO invoice_headers (invoice_number, invoice_date, invoice_type, currency, status, total_net, source_file)
                    VALUES (?, ?, ?, ?, ?, 0, ?)
                    ON CONFLICT(invoice_number) DO UPDATE SET
                        invoice_date=excluded.invoice_date,
                        invoice_type=excluded.invoice_type,
                        currency=excluded.currency,
                        status=excluded.status
                """, (invoice_number, row.get('invoice_date'), row.get('invoice_type') or 'logistics',
                      row.get('currency') or 'PLN', row.get('status') or 'pending', 'core_backup_import'))
                invoice_cache[invoice_number] = main_db.execute("SELECT id FROM invoice_headers WHERE invoice_number=?", (invoice_number,)).fetchone()['id']
            invoice_id = invoice_cache[invoice_number]
            exists = main_db.execute("SELECT id FROM invoice_items WHERE invoice_id=? AND stt_number=? AND COALESCE(ref_date,'')=COALESCE(?, '')",
                                     (invoice_id, stt, row.get('ref_date'))).fetchone()
            if exists:
                main_db.execute("""UPDATE invoice_items
                                   SET net_amount=?, matched_booking_id=?, accepted_amount=?,
                                       acceptance_status=?, exception_reason=?, acceptance_remark=?
                                   WHERE id=?""",
                                (
                                    row.get('net_amount') or 0, row.get('matched_booking_id'),
                                    row.get('accepted_amount'), row.get('acceptance_status'),
                                    row.get('exception_reason'), row.get('acceptance_remark'), exists['id']
                                ))
                result = 'updated'
            else:
                main_db.execute("""INSERT INTO invoice_items (
                                      invoice_id, stt_number, net_amount, ref_date, matched_booking_id,
                                      accepted_amount, acceptance_status, exception_reason, acceptance_remark
                                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    invoice_id, stt, row.get('net_amount') or 0, row.get('ref_date'), row.get('matched_booking_id'),
                                    row.get('accepted_amount'), row.get('acceptance_status'),
                                    row.get('exception_reason'), row.get('acceptance_remark')
                                ))
                result = 'created'
            report['invoices'][result] = report['invoices'].get(result, 0) + 1
        for invoice_number, invoice_id in invoice_cache.items():
            total = main_db.execute("SELECT COALESCE(SUM(net_amount), 0) AS total FROM invoice_items WHERE invoice_id=?", (invoice_id,)).fetchone()['total']
            main_db.execute("UPDATE invoice_headers SET total_net=? WHERE id=?", (total, invoice_id))

        for row in rows_from_sheet(wb, '库存底表'):
            product_number = str(row.get('product_number') or '').strip()
            warehouse = str(row.get('warehouse') or '').strip()
            if not product_number or not warehouse:
                report['inventory']['skipped'] = report['inventory'].get('skipped', 0) + 1
                continue
            bom = normalize_backup_bom(row.get('bom_code'))
            pm = {
                'product_number': product_number,
                'bom_code': bom,
                'product_description': row.get('product_description'),
                'category': row.get('category'),
                'category_2': row.get('category_2'),
                'unit_price': row.get('unit_price') or 0,
                'last_inbound_date': row.get('last_inbound_date'),
                'last_outbound_date': row.get('last_outbound_date'),
                'remark': row.get('comment'),
                'dos_threshold': row.get('dos_threshold') or 180,
                'idle_threshold': row.get('idle_threshold') or 180,
            }
            inv = {
                'warehouse': warehouse,
                'product_number': product_number,
                'bom_code': bom,
                'instruction': row.get('instruction'),
                'inventory': row.get('inventory') or 0,
                'last_inbound_date': row.get('last_inbound_date'),
                'last_outbound_date': row.get('last_outbound_date'),
                'comment': row.get('comment'),
            }
            upsert_by_key(wms_conn, 'product_master', ['product_number'], pm)
            result = upsert_by_key(wms_conn, 'warehouse_inventory', ['warehouse', 'product_number'], inv)
            report['inventory'][result] = report['inventory'].get(result, 0) + 1

        for row in rows_from_sheet(wb, 'SKU主数据'):
            row['bom_code'] = normalize_backup_bom(row.get('bom_code'))
            result = upsert_by_key(wms_conn, 'product_master', ['product_number'], row)
            report['sku'][result] = report['sku'].get(result, 0) + 1

        for row in rows_from_sheet(wb, '出入库流水'):
            product_number = str(row.get('product_number') or '').strip()
            if not product_number:
                report['movements']['skipped'] = report['movements'].get('skipped', 0) + 1
                continue
            exists = wms_conn.execute("""
                SELECT id FROM inventory_movements
                WHERE movement_type=? AND product_number=? AND COALESCE(movement_date,'')=COALESCE(?, '') AND COALESCE(operation_id,'')=COALESCE(?, '')
                LIMIT 1
            """, (row.get('movement_type'), product_number, row.get('movement_date'), row.get('operation_id'))).fetchone()
            if exists:
                report['movements']['skipped_existing'] = report['movements'].get('skipped_existing', 0) + 1
                continue
            row = {k: v for k, v in row.items() if k in table_columns(wms_conn, 'inventory_movements') and k != 'id'}
            cols = list(row.keys())
            wms_conn.execute(f"INSERT INTO inventory_movements ({', '.join(cols)}) VALUES ({', '.join(['?'] * len(cols))})", [row[c] for c in cols])
            report['movements']['created'] = report['movements'].get('created', 0) + 1

        image_report = {'extracted': 0, 'mapped': 0, 'skipped': 0}
        if extracted_zip_images:
            image_dir = current_app.config['UPLOAD_FOLDER']
            os.makedirs(image_dir, exist_ok=True)
            for base, content in extracted_zip_images:
                safe_name = secure_filename(base) or base
                ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else ''
                if ext not in {'png', 'jpg', 'jpeg', 'gif', 'webp'}:
                    image_report['skipped'] += 1
                    continue
                target = os.path.abspath(os.path.join(image_dir, safe_name))
                if not target.startswith(os.path.abspath(image_dir) + os.sep):
                    image_report['skipped'] += 1
                    continue
                with open(target, 'wb') as img_file:
                    img_file.write(content)
                image_report['extracted'] += 1
            try:
                from .wms_routes import repair_image_mappings
                image_report['mapped'] = repair_image_mappings(wms_conn)
            except Exception as image_error:
                image_report['error'] = str(image_error)
        report['images'] = image_report

        main_db.commit()
        wms_conn.commit()
        return jsonify({'success': True, 'report': report})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            wms_conn.close()
        except Exception:
            pass


@main_bp.route('/api/export/logistics')
def export_logistics_list():
    """Export active logistics shipment list as a styled XLSX workbook."""
    from openpyxl import Workbook
    db = get_db()
    invoice_rows = db.execute("""
        SELECT i.stt_number,
               COALESCE(i.accepted_amount, i.net_amount, 0) AS net_amount,
               h.invoice_number,
               COALESCE(NULLIF(i.acceptance_status, ''), h.status) AS status
        FROM invoice_items i
        JOIN invoice_headers h ON i.invoice_id = h.id
    """).fetchall()
    stt_status = {}
    verified_amounts = {}
    verified_invoices = {}
    for r in invoice_rows:
        stt = r['stt_number'] or ''
        if not stt:
            continue
        status = r['status'] or 'pending'
        if status == 'verified':
            stt_status[stt] = '已验收'
            verified_amounts[stt] = verified_amounts.get(stt, 0) + float(r['net_amount'] or 0)
            verified_invoices.setdefault(stt, set()).add(r['invoice_number'])
        elif stt not in stt_status:
            stt_status[stt] = '待验收' if status == 'pending' else '已拒绝'

    ds = request.args.get('ds', '').strip()
    type_filter = request.args.get('type', '').strip()
    country = request.args.get('country', '').strip().upper()
    pickup_month = request.args.get('pickup_month', '').strip()
    where = []
    params = []
    if type_filter:
        where.append('shipment_type = ?')
        params.append(type_filter)
    if country:
        where.append("COALESCE(NULLIF(UPPER(TRIM(demand_country)), ''), UPPER(TRIM(consignee_country_code))) = ?")
        params.append(country)
    if pickup_month.startswith('year:') and re.match(r'^year:\d{4}$', pickup_month):
        where.append("SUBSTR(pickup_date, 1, 4) = ?")
        params.append(pickup_month.split(':', 1)[1])
    elif pickup_month == 'annual':
        where.append("SUBSTR(pickup_date, 1, 4) = ?")
        params.append(str(datetime.now().year))
    elif pickup_month:
        where.append("SUBSTR(pickup_date, 1, 7) = ?")
        params.append(pickup_month)
    where_sql = 'WHERE ' + ' AND '.join(where) if where else ''
    rows = db.execute(f"""
        SELECT *
        FROM shipments
        {where_sql}
        ORDER BY pickup_date DESC, booking_id
    """, params).fetchall()
    if ds:
        rows = [r for r in rows if calc_display_status(r) == ds]
    type_labels = {
        'warehouse_central': '中央仓发出',
        'warehouse_furniture': '家具仓发出',
        'direct': '调拨',
    }
    headers = [
        'Booking ID', 'STT Number', 'Waybill No', 'Supplier Status', 'Display Status',
        'Manual Status', 'Delivery Locked', 'Shipment Type', 'Warehouse',
        'Demand Country', 'Demand Country Source', 'Consignee Country',
        'Consignee Name', 'Consignee City', 'Pickup Date', 'ETA Delivery Date',
        'Actual Delivery Date', 'Pieces', 'Weight', 'Volume', 'Price Without VAT',
        'Product', 'Service Type', 'Verify Status', 'Accepted Amount PLN',
        'Accepted Invoice', 'Recycle Status', 'Recycle Reason', 'Recycle Time',
        'Recycle Operator', 'Source File'
    ]
    data_rows = []
    status_summary = {'已交付': 0, '待核实': 0, '运输中': 0, '异常单': 0}
    total_weight = total_volume = total_price = accepted_total = 0
    for r in rows:
        stt = r['stt_number'] or ''
        display_status = calc_display_status(r)
        status_summary[display_status] = status_summary.get(display_status, 0) + 1
        total_weight += float(r['total_weight'] or 0)
        total_volume += float(r['total_volume'] or 0)
        total_price += float(r['price_without_vat'] or 0)
        accepted_amount = round(verified_amounts.get(stt, 0), 2) if stt in verified_amounts else ''
        if accepted_amount != '':
            accepted_total += accepted_amount
        data_rows.append([
            r['booking_id'], stt, r['waybill_no'], r['shipment_status'], display_status,
            r['manual_status'], 'Y' if r['delivery_locked'] else '',
            type_labels.get(r['shipment_type'], r['shipment_type']), r['warehouse'],
            effective_demand_country(r), r['demand_country_source'], r['consignee_country_code'],
            r['consignee_name'], r['consignee_city'], r['pickup_date'], r['delivery_date'],
            r['actual_delivery_date'], r['total_pieces'], r['total_weight'], r['total_volume'],
            r['price_without_vat'], r['product'], r['service_type'], stt_status.get(stt, '未关联'),
            accepted_amount,
            ', '.join(sorted(verified_invoices.get(stt, []))),
            '回收站' if r['recycled_at'] else '',
            r['recycle_reason'], r['recycled_at'], r['recycle_operator'], r['source_file']
        ])

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    month_label = 'all_months'
    if pickup_month.startswith('year:') and re.match(r'^year:\d{4}$', pickup_month):
        month_label = pickup_month.split(':', 1)[1] + '全年'
    elif pickup_month == 'annual':
        month_label = str(datetime.now().year) + '全年'
    elif pickup_month:
        month_label = pickup_month
    name_parts = ['VN62', '物流清单', safe_filename_part(month_label), ts]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Logistics'
    summary = [
        ('总运单数', len(rows)),
        ('已交付 / 待核实 / 运输中 / 异常单', f"{status_summary.get('已交付', 0)} / {status_summary.get('待核实', 0)} / {status_summary.get('运输中', 0)} / {status_summary.get('异常单', 0)}"),
        ('总重量 kg', round(total_weight, 2)),
        ('总体积', round(total_volume, 2)),
        ('物流费用', round(total_price, 2)),
        ('已验收金额 PLN', round(accepted_total, 2)),
    ]
    _style_report_sheet(ws, '零售中央仓 - 全量物流清单', f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} · 包含人工调整、回收站和验收金额', headers, data_rows, summary)
    return _xlsx_response(wb, '_'.join(name_parts) + '.xlsx')


@main_bp.route('/api/export/logistics-calibration')
def export_logistics_calibration():
    """Export a controlled logistics status calibration workbook."""
    from openpyxl import Workbook
    db = get_db()
    rows = db.execute("""
        SELECT booking_id, stt_number, shipment_status, manual_status, pickup_date, delivery_date, actual_delivery_date,
               delivery_locked, demand_country, consignee_country_code,
               recycled_at, recycle_reason, recycle_operator
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled'
        ORDER BY pickup_date DESC, booking_id
    """).fetchall()
    headers = [
        'STT Number', 'Booking ID', '当前显示状态', '手动物流状态',
        '实际交付日期', '锁定交付日期', '需求国家', '收货国家',
        '回收站状态', '回收站备注', '回收站操作人'
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Logistics Calibration'
    ws.append(headers)
    for r in rows:
        ws.append([
            r['stt_number'] or '',
            r['booking_id'] or '',
            calc_display_status(r),
            r['manual_status'] or '',
            r['actual_delivery_date'] or '',
            'Y' if r['delivery_locked'] else '',
            effective_demand_country(r),
            r['consignee_country_code'] or '',
            'Y' if r['recycled_at'] else '',
            r['recycle_reason'] or '',
            r['recycle_operator'] or '',
        ])
    _style_simple_sheet(ws)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return _xlsx_response(wb, f'VN62_logistics_status_calibration_{ts}.xlsx')


@main_bp.route('/api/shipments/import-calibration', methods=['POST'])
def import_logistics_calibration():
    """Import logistics status calibration. Only controlled/manual fields are updated."""
    import openpyxl
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传物流状态校准 Excel'}), 400
    if uploaded_file_size(file) > 20 * 1024 * 1024:
        return jsonify({'error': 'Excel 文件不能超过 20MB'}), 400
    db_module.backup_sqlite_db(current_app.config['DATABASE'], 'logistics_calibration_import')
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    required = ['STT Number', 'Booking ID']
    aliases = {
        '当前显示状态': ['Display Status', '物流状态'],
        '手动物流状态': ['Manual Status', 'Manual Logistics Status'],
        '实际交付日期': ['Actual Delivery Date'],
        '锁定交付日期': ['Delivery Locked'],
        '需求国家': ['Demand Country'],
        '回收站状态': ['Recycle Status'],
        '回收站备注': ['Recycle Reason'],
        '回收站操作人': ['Recycle Operator']
    }
    header_row, col_map = _find_header(ws, required, aliases={**aliases, 'STT Number': ['STT'], 'Booking ID': ['Booking']}, scan_rows=8)
    if not header_row:
        return jsonify({'error': '未找到物流校准表头，请确认包含 STT Number / Booking ID'}), 400
    db = get_db()
    updated = skipped = recycled = restored = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        stt = _clean_cell(_cell(ws, row_idx, col_map, 'STT Number'))
        booking = _clean_cell(_cell(ws, row_idx, col_map, 'Booking ID'))
        if not stt and not booking:
            continue
        row = db.execute("""
            SELECT booking_id FROM shipments
            WHERE (? != '' AND stt_number = ?) OR (? != '' AND booking_id = ?)
            ORDER BY CASE WHEN booking_id = ? THEN 0 ELSE 1 END
            LIMIT 1
        """, (stt, stt, booking, booking, booking)).fetchone()
        if not row:
            skipped += 1
            continue
        booking_id = row['booking_id']
        manual_status = _clean_cell(_cell(ws, row_idx, col_map, '手动物流状态'))
        actual_delivery = _date_cell(_cell(ws, row_idx, col_map, '实际交付日期'))
        delivery_locked_raw = _clean_cell(_cell(ws, row_idx, col_map, '锁定交付日期')).upper()
        demand_country = _clean_cell(_cell(ws, row_idx, col_map, '需求国家')).upper()
        recycle_status = _clean_cell(_cell(ws, row_idx, col_map, '回收站状态')).upper()
        recycle_reason = _clean_cell(_cell(ws, row_idx, col_map, '回收站备注'))
        recycle_operator = _clean_cell(_cell(ws, row_idx, col_map, '回收站操作人'))
        db.execute("""
            UPDATE shipments
            SET manual_status = COALESCE(NULLIF(?, ''), manual_status),
                actual_delivery_date = COALESCE(NULLIF(?, ''), actual_delivery_date),
                delivery_locked = CASE WHEN ? IN ('Y','YES','1','TRUE','是') THEN 1 ELSE delivery_locked END,
                demand_country = CASE WHEN ? != '' THEN ? ELSE demand_country END,
                demand_country_source = CASE WHEN ? != '' THEN 'manual' ELSE demand_country_source END
            WHERE booking_id = ?
        """, (manual_status, actual_delivery, delivery_locked_raw, demand_country, demand_country, demand_country, booking_id))
        if recycle_status in ('Y', 'YES', '1', 'TRUE', '是'):
            db.execute("""
                UPDATE shipments SET recycled_at = COALESCE(recycled_at, CURRENT_TIMESTAMP),
                    recycle_reason = COALESCE(NULLIF(?, ''), recycle_reason),
                    recycle_operator = COALESCE(NULLIF(?, ''), recycle_operator)
                WHERE booking_id = ?
            """, (recycle_reason, recycle_operator, booking_id))
            recycled += 1
        elif recycle_status in ('N', 'NO', '0', 'FALSE', '否'):
            db.execute("UPDATE shipments SET recycled_at = NULL, recycle_reason = NULL, recycle_operator = NULL WHERE booking_id = ?", (booking_id,))
            restored += 1
        updated += 1
    db.commit()
    return jsonify({'success': True, 'updated': updated, 'skipped': skipped, 'recycled': recycled, 'restored': restored})


@main_bp.route('/api/export/invoice-calibration')
def export_invoice_calibration():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    db = get_db()
    rows = db.execute("""
        SELECT h.invoice_number, h.invoice_date, h.invoice_type, h.currency, h.status AS invoice_status,
               i.stt_number, i.net_amount, i.ref_date, i.matched_booking_id,
               COALESCE(NULLIF(i.acceptance_status, ''), h.status, 'pending') AS acceptance_status,
               COALESCE(i.accepted_amount, i.net_amount, 0) AS accepted_amount,
               i.exception_reason, i.acceptance_remark,
               s.booking_id, s.shipment_status, s.manual_status, s.shipment_type,
               s.demand_country, s.consignee_country_code, s.consignee_name, s.consignee_city,
               s.pickup_date, s.delivery_date, s.actual_delivery_date, s.delivery_locked,
               s.total_weight, s.total_volume, s.price_without_vat
        FROM invoice_items i
        JOIN invoice_headers h ON i.invoice_id = h.id
        LEFT JOIN shipments s ON (
            s.booking_id = i.matched_booking_id
            OR (COALESCE(i.matched_booking_id, '') = '' AND (s.stt_number = i.stt_number OR s.booking_id = i.stt_number))
        )
        ORDER BY h.invoice_date DESC, h.invoice_number, i.id
    """).fetchall()
    headers = [
        'Invoice Number', 'Invoice Date', 'Invoice Type', 'STT Number',
        'Original Invoice Amount PLN', 'Currency', 'Ref Date',
        'Matched Booking ID', 'Match Status',
        'Logistics Status', 'Shipment Type', 'Demand Country', 'Consignee Country',
        'Consignee Name', 'Consignee City', 'Pickup Date', 'ETA Delivery Date',
        'Actual Delivery Date', 'Weight KG', 'Volume CBM', 'Logistics List Amount',
        'Amount Difference', 'Acceptance Status', 'Accepted Amount PLN',
        'Exception Reason', 'Acceptance Remark'
    ]
    type_labels = {
        'warehouse_central': '中央仓发出',
        'warehouse_furniture': '家具仓发出',
        'direct': '调拨',
    }
    status_labels = {'verified': '已验收', 'pending': '待验收', 'rejected': '已拒绝'}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice Acceptance'
    ws.append(headers)
    for r in rows:
        logistics_amount = float(r['price_without_vat'] or 0)
        invoice_amount = float(r['net_amount'] or 0)
        matched_booking = r['matched_booking_id'] or r['booking_id'] or ''
        ws.append([
            r['invoice_number'] or '',
            r['invoice_date'] or '',
            r['invoice_type'] or '',
            r['stt_number'] or '',
            invoice_amount,
            r['currency'] or 'PLN',
            r['ref_date'] or '',
            matched_booking,
            '已匹配' if matched_booking else '未匹配',
            calc_display_status(r) if r['booking_id'] else '',
            type_labels.get(r['shipment_type'], r['shipment_type'] or ''),
            r['demand_country'] or '',
            r['consignee_country_code'] or '',
            r['consignee_name'] or '',
            r['consignee_city'] or '',
            r['pickup_date'] or '',
            r['delivery_date'] or '',
            r['actual_delivery_date'] or '',
            r['total_weight'] or '',
            r['total_volume'] or '',
            logistics_amount if logistics_amount else '',
            round(invoice_amount - logistics_amount, 2) if logistics_amount else '',
            status_labels.get(r['acceptance_status'], r['acceptance_status'] or '待验收'),
            r['accepted_amount'] if r['accepted_amount'] is not None else invoice_amount,
            r['exception_reason'] or '',
            r['acceptance_remark'] or '',
        ])
    _style_simple_sheet(ws)
    manual_fill = PatternFill('solid', fgColor='FEF3C7')
    readonly_fill = PatternFill('solid', fgColor='E0F2FE')
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        if title in ('Acceptance Status', 'Accepted Amount PLN', 'Exception Reason', 'Acceptance Remark', 'Matched Booking ID'):
            cell.fill = manual_fill
        else:
            cell.fill = readonly_fill
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=False)
    note = wb.create_sheet('Read Me')
    note.append(['用途', '说明'])
    note.append(['工作方式', '本文件既是发票验收材料，也是可回传系统的发票验收校准 Excel。'])
    note.append(['黄色字段', '可人工维护：Matched Booking ID / Acceptance Status / Accepted Amount PLN / Exception Reason / Acceptance Remark。'])
    note.append(['状态值', 'Acceptance Status 支持：已验收 / 待验收 / 已拒绝，也兼容 verified / pending / rejected。'])
    note.append(['保护规则', '回传时不覆盖原始发票金额和物流原始数据，只更新人工验收字段和匹配关系。'])
    _style_simple_sheet(note)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return _xlsx_response(wb, f'VN62_invoice_acceptance_calibration_{ts}.xlsx')


@main_bp.route('/api/invoices/import-calibration', methods=['POST'])
def import_invoice_calibration():
    import openpyxl
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传发票验收校准 Excel'}), 400
    if uploaded_file_size(file) > 20 * 1024 * 1024:
        return jsonify({'error': 'Excel 文件不能超过 20MB'}), 400
    db_module.backup_sqlite_db(current_app.config['DATABASE'], 'invoice_acceptance_calibration_import')
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    required = ['Invoice Number', 'STT Number']
    aliases = {
        'Acceptance Status': ['Status', '验收状态', '验收结论', '是否确认验收'],
        'Accepted Amount PLN': ['Accepted Amount', '验收金额', '验收金额 PLN', 'Net Amount'],
        'Matched Booking ID': ['Booking ID', '匹配运单', '匹配 Booking ID'],
        'Ref Date': ['Ref日期'],
        'Exception Reason': ['异常原因', '问题类型', 'Reason'],
        'Acceptance Remark': ['人工备注', '备注', 'Comment'],
    }
    header_row, col_map = _find_header(ws, required, aliases=aliases, scan_rows=8)
    if not header_row:
        return jsonify({'error': '未找到发票校准表头，请确认包含 Invoice Number / STT Number'}), 400
    db = get_db()
    updated_headers = set()
    updated_items = skipped = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        invoice_number = _clean_cell(_cell(ws, row_idx, col_map, 'Invoice Number'))
        stt = _clean_cell(_cell(ws, row_idx, col_map, 'STT Number'))
        if not invoice_number or not stt:
            continue
        header = db.execute("SELECT id FROM invoice_headers WHERE invoice_number = ?", (invoice_number,)).fetchone()
        if not header:
            skipped += 1
            continue
        invoice_id = header['id']
        status = normalize_acceptance_status(_clean_cell(_cell(ws, row_idx, col_map, 'Acceptance Status')))
        accepted_amount = _float_cell(_cell(ws, row_idx, col_map, 'Accepted Amount PLN'), None)
        ref_date = _date_cell(_cell(ws, row_idx, col_map, 'Ref Date'))
        matched_booking = _clean_cell(_cell(ws, row_idx, col_map, 'Matched Booking ID'))
        exception_reason = _clean_cell(_cell(ws, row_idx, col_map, 'Exception Reason'))
        acceptance_remark = _clean_cell(_cell(ws, row_idx, col_map, 'Acceptance Remark'))
        fields = []
        params = []
        if status:
            fields.append('acceptance_status = ?')
            params.append(status)
        if accepted_amount is not None:
            fields.append('accepted_amount = ?')
            params.append(accepted_amount)
        if matched_booking:
            fields.append('matched_booking_id = ?')
            params.append(matched_booking)
        if exception_reason:
            fields.append('exception_reason = ?')
            params.append(exception_reason)
        if acceptance_remark:
            fields.append('acceptance_remark = ?')
            params.append(acceptance_remark)
        if fields:
            where = "invoice_id = ? AND stt_number = ?"
            params.extend([invoice_id, stt])
            if ref_date:
                where += " AND COALESCE(ref_date, '') = COALESCE(?, '')"
                params.append(ref_date)
            cur = db.execute(f"UPDATE invoice_items SET {', '.join(fields)} WHERE {where}", params)
            updated_items += cur.rowcount
            updated_headers.add(invoice_id)
    for invoice_id in updated_headers:
        recalc_invoice_header_status(db, invoice_id)
    db.commit()
    return jsonify({'success': True, 'updated_headers': len(updated_headers), 'updated_items': updated_items, 'skipped': skipped})


@main_bp.route('/api/export/inventory')
def export_inventory_list():
    """Export inventory calibration workbook: Lodz + Bydgoszcz sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    headers = [
        'Bom Code',
        'Product Number',
        'Product Description',
        'Category',
        'Category 2',
        'Instruction',
        'Inventory (Pcs)',
        'Unit Price (USD)',
        'TTL Amount (USD)',
        '最近入库时间 Last Inbound date',
        '最近出库时间 Last Outbound date',
        '库龄 Dos (Days of Supply)',
        '库存静置时长 Days without Stock',
        '备注 comment',
        'SKU条目创建时间',
        '图片 Photo',
    ]
    sheet_map = [
        ('Lodz', 'Lodz warehouse'),
        ('Bydgoszcz', 'Bydgoszcz warehouse'),
    ]

    def style_calibration_sheet(ws, row_count):
        header_fill = PatternFill('solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='D9E2F3')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        widths = [16, 22, 42, 18, 18, 16, 15, 16, 16, 24, 24, 22, 24, 30, 20, 24]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for row in ws.iter_rows(min_row=2, max_row=max(row_count, 2), max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=False)
        if row_count >= 2:
            ref = f"A1:{get_column_letter(len(headers))}{row_count}"
            table = Table(displayName=re.sub(r'[^A-Za-z0-9_]', '_', f"{ws.title}_Inventory"), ref=ref)
            style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
        ws.sheet_view.showGridLines = False

    wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
    wms_conn.row_factory = sqlite3.Row
    include_images = request.args.get('include_images') in ('1', 'true', 'yes')
    warehouse_filter = (request.args.get('warehouse') or '').strip()
    category_values = split_request_values(request.args.get('category') or '')
    category2_values = split_request_values(request.args.get('category2') or '')
    instruction_values = split_request_values(request.args.get('instruction') or '')
    try:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        wb = Workbook()
        image_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), '..', 'static', 'images')
        created_sheets = 0
        for idx, (sheet_title, warehouse) in enumerate(sheet_map):
            if warehouse_filter and warehouse_filter != warehouse:
                continue
            ws = wb.active if created_sheets == 0 else wb.create_sheet()
            created_sheets += 1
            ws.title = sheet_title
            ws.append(headers)
            where = ["wi.warehouse = ?"]
            params = [warehouse]
            if category_values:
                placeholders = ','.join(['?'] * len(category_values))
                where.append(f"UPPER(TRIM(COALESCE(pm.category, sm.category, ''))) IN ({placeholders})")
                params.extend(category_values)
            if category2_values:
                placeholders = ','.join(['?'] * len(category2_values))
                where.append(f"UPPER(TRIM(COALESCE(pm.category_2, sm.category_2, ''))) IN ({placeholders})")
                params.extend(category2_values)
            if instruction_values:
                placeholders = ','.join(['?'] * len(instruction_values))
                where.append(f"UPPER(TRIM(COALESCE(wi.instruction, ''))) IN ({placeholders})")
                params.extend(instruction_values)
            rows = wms_conn.execute(f"""
                SELECT wi.warehouse,
                       COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code, sm.bom_code) AS bom_code,
                       wi.product_number,
                       COALESCE(pm.product_description, sm.product_description) AS product_description,
                       COALESCE(pm.category, sm.category) AS category,
                       COALESCE(pm.category_2, sm.category_2) AS category_2,
                       wi.instruction,
                       wi.inventory,
                       COALESCE(pm.unit_price, sm.unit_price, 0) AS unit_price,
                       COALESCE(wi.last_inbound_date, pm.last_inbound_date, DATE(pm.created_at)) AS last_inbound_date,
                       COALESCE(wi.last_outbound_date, pm.last_outbound_date) AS last_outbound_date,
                       wi.comment,
                       DATE(pm.created_at) AS sku_created_at,
                       COALESCE(pimg.image_path, sm.image_path) AS image_path
                FROM warehouse_inventory wi
                LEFT JOIN product_master pm ON wi.product_number = pm.product_number
                LEFT JOIN sku_master sm ON COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code) = sm.bom_code
                    AND NOT (
                        REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                        OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                    )
                LEFT JOIN sku_master pimg ON (
                    REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                    OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                ) AND wi.product_number = pimg.bom_code
                WHERE {' AND '.join(where)}
                ORDER BY wi.product_number
            """, params).fetchall()
            for r in rows:
                inventory = float(r['inventory'] or 0)
                unit_price = float(r['unit_price'] or 0)
                ttl_amount = inventory * unit_price
                last_inbound = r['last_inbound_date'] or ''
                last_outbound = r['last_outbound_date'] or ''
                dos = ''
                idle = ''
                if last_inbound:
                    days = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_inbound,)).fetchone()[0]
                    dos = days if days is not None else ''
                if last_outbound:
                    days = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_outbound,)).fetchone()[0]
                    idle = days if days is not None else ''
                ws.append([
                    r['bom_code'] or '',
                    r['product_number'] or '',
                    r['product_description'] or '',
                    r['category'] or '',
                    r['category_2'] or '',
                    r['instruction'] or '',
                    int(inventory),
                    unit_price,
                    ttl_amount,
                    last_inbound,
                    last_outbound,
                    dos,
                    idle,
                    r['comment'] or '',
                    r['sku_created_at'] or '',
                    '',
                ])
                row_idx = ws.max_row
                ws.row_dimensions[row_idx].height = 58
                if include_images and r['image_path']:
                    img_path = os.path.abspath(os.path.join(image_dir, r['image_path']))
                    if img_path.startswith(os.path.abspath(image_dir) + os.sep) and os.path.exists(img_path):
                        try:
                            img = ExcelImage(img_path)
                            img.width = 72
                            img.height = 52
                            ws.add_image(img, f'P{row_idx}')
                        except Exception:
                            pass
                for money_cell in (ws.cell(row_idx, 8), ws.cell(row_idx, 9)):
                    money_cell.number_format = '$#,##0.00'
                ws.cell(row_idx, 7).number_format = '#,##0'
            style_calibration_sheet(ws, ws.max_row)
        if created_sheets == 0:
            ws = wb.active
            ws.title = 'No Data'
            ws.append(headers)
            style_calibration_sheet(ws, ws.max_row)
        suffix = 'with_images' if include_images else 'calibration'
        category_label = '-'.join(category_values[:2]) + (f'等{len(category_values)}项' if len(category_values) > 2 else '') if category_values else '全部品类'
        cat2_label = '-'.join(category2_values[:2]) + (f'等{len(category2_values)}项' if len(category2_values) > 2 else '') if category2_values else '全部产品'
        instruction_label = '-'.join(instruction_values[:2]) + (f'等{len(instruction_values)}项' if len(instruction_values) > 2 else '') if instruction_values else '全部国家'
        name_parts = ['VN62', '库存清单', safe_filename_part(category_label), safe_filename_part(cat2_label), safe_filename_part(instruction_label), suffix, ts]
        return _xlsx_response(wb, '_'.join(name_parts) + '.xlsx')
    finally:
        wms_conn.close()


@main_bp.route('/api/export/inventory/category2')
def export_inventory_category2_report():
    """Export a styled inventory report filtered by Category 2."""
    from openpyxl import Workbook
    category2 = (request.args.get('category2') or '').strip().upper()
    category2_values = split_request_values(category2)
    warehouse = (request.args.get('warehouse') or '').strip()
    if not category2_values:
        return jsonify({'error': '请选择 Category 2 / 适用产品后再导出报告'}), 400
    if warehouse and warehouse not in ('Lodz warehouse', 'Bydgoszcz warehouse'):
        return jsonify({'error': '仓库参数无效'}), 400

    placeholders = ','.join(['?'] * len(category2_values))
    where = [f"UPPER(TRIM(COALESCE(pm.category_2, sm.category_2, ''))) IN ({placeholders})"]
    params = category2_values[:]
    if warehouse:
        where.append("wi.warehouse = ?")
        params.append(warehouse)

    wms_conn = sqlite3.connect(current_app.config['WMS_DATABASE'])
    wms_conn.row_factory = sqlite3.Row
    try:
        rows = wms_conn.execute(f"""
            SELECT wi.warehouse,
                   COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code, sm.bom_code) AS bom_code,
                   wi.product_number,
                   COALESCE(pm.product_description, sm.product_description) AS product_description,
                   COALESCE(pm.category, sm.category) AS category,
                   COALESCE(pm.category_2, sm.category_2) AS category_2,
                   wi.instruction,
                   wi.inventory,
                   COALESCE(pm.unit_price, sm.unit_price, 0) AS unit_price,
                   COALESCE(wi.last_inbound_date, pm.last_inbound_date, DATE(pm.created_at)) AS last_inbound_date,
                   COALESCE(wi.last_outbound_date, pm.last_outbound_date) AS last_outbound_date,
                   wi.comment,
                   COALESCE(pimg.image_path, sm.image_path) AS image_path
            FROM warehouse_inventory wi
            LEFT JOIN product_master pm ON wi.product_number = pm.product_number
            LEFT JOIN sku_master sm ON COALESCE(NULLIF(wi.bom_code, ''), pm.bom_code) = sm.bom_code
                AND NOT (
                    REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                    OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
                )
            LEFT JOIN sku_master pimg ON (
                REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(wi.bom_code, pm.bom_code, ''))), ' ', ''), '_', ''), '-', '') IN ('本地', 'LOCAL', 'LOCALSKU', 'LOCALITEM', '/', 'NOBOM', 'NOBOMCODE')
                OR COALESCE(NULLIF(wi.bom_code, ''), NULLIF(pm.bom_code, ''), '') = ''
            ) AND wi.product_number = pimg.bom_code
            WHERE {' AND '.join(where)}
            ORDER BY wi.warehouse, wi.product_number
        """, params).fetchall()

        headers = [
            'Warehouse', 'Bom Code', 'Product Number', 'Product Description',
            'Category', 'Category 2', 'Instruction', 'Inventory (Pcs)',
            'Unit Price (USD)', 'TTL Amount (USD)', 'Last Inbound date',
            'Last Outbound date', 'Dos (Days)', 'Days without Stock',
            'Comment', 'Photo'
        ]
        data_rows = []
        total_pcs = 0
        total_value = 0
        warehouse_summary = {}
        for r in rows:
            inventory = float(r['inventory'] or 0)
            unit_price = float(r['unit_price'] or 0)
            ttl_amount = inventory * unit_price
            total_pcs += inventory
            total_value += ttl_amount
            wh = r['warehouse'] or ''
            warehouse_summary.setdefault(wh, {'pcs': 0, 'value': 0, 'sku': 0})
            warehouse_summary[wh]['pcs'] += inventory
            warehouse_summary[wh]['value'] += ttl_amount
            warehouse_summary[wh]['sku'] += 1
            last_inbound = r['last_inbound_date'] or ''
            last_outbound = r['last_outbound_date'] or ''
            dos = ''
            idle = ''
            if last_inbound:
                days = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_inbound,)).fetchone()[0]
                dos = days if days is not None else ''
            if last_outbound:
                days = wms_conn.execute("SELECT CAST(JULIANDAY('now') - JULIANDAY(?) AS INTEGER)", (last_outbound,)).fetchone()[0]
                idle = days if days is not None else ''
            data_rows.append([
                wh,
                r['bom_code'] or '',
                r['product_number'] or '',
                r['product_description'] or '',
                r['category'] or '',
                r['category_2'] or '',
                r['instruction'] or '',
                int(inventory),
                unit_price,
                ttl_amount,
                last_inbound,
                last_outbound,
                dos,
                idle,
                r['comment'] or '',
                r['image_path'] or '',
            ])

        wb = Workbook()
        ws = wb.active
        ws.title = 'Category2 Inventory'
        scope_text = warehouse or 'All Warehouses'
        summary = [
            ('适用产品 Category 2', ', '.join(category2_values)),
            ('统计范围', scope_text),
            ('SKU 条目数', len(rows)),
            ('库存总件数 PCS', int(total_pcs)),
            ('库存总货值 USD', round(total_value, 2)),
        ]
        for wh, s in sorted(warehouse_summary.items()):
            summary.append((f"{wh} SKU / PCS / USD", f"{s['sku']} / {int(s['pcs'])} / {round(s['value'], 2)}"))
        _style_report_sheet(
            ws,
            f"零售中央仓 - {', '.join(category2_values)} 库存清单报告",
            f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} · Category 2 筛选报告',
            headers,
            data_rows,
            summary
        )
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_cat2 = safe_filename_part('-'.join(category2_values), 'CATEGORY2')
        return _xlsx_response(wb, f'VN62_inventory_category2_{safe_cat2}_{ts}.xlsx')
    finally:
        wms_conn.close()


@main_bp.route('/api/shipments/mark_delivered', methods=['POST'])
def mark_delivered():
    """手动标记运单为已交付，写入 actual_delivery_date 并锁定"""
    from flask import request
    from datetime import datetime
    data = request.get_json()
    booking_id = data.get('booking_id', '').strip()
    delivery_date = data.get('delivery_date', '').strip()  # 可选指定日期
    if not booking_id:
        return jsonify({'error': 'Missing booking_id'}), 400
    db = get_db()
    # 检查运单是否存在
    row = db.execute("SELECT delivery_locked FROM shipments WHERE booking_id = ?", (booking_id,)).fetchone()
    if not row:
        return jsonify({'error': f'运单 {booking_id} 不存在'}), 404
    if row['delivery_locked']:
        return jsonify({'error': f'运单 {booking_id} 已锁定，请先撤销再重新标记'}), 409
    if not delivery_date:
        delivery_date = datetime.now().strftime('%Y-%m-%d')
    db.execute(
        "UPDATE shipments SET actual_delivery_date = ?, delivery_locked = 1 WHERE booking_id = ?",
        (delivery_date, booking_id)
    )
    db.commit()
    return jsonify({'success': True, 'booking_id': booking_id, 'actual_delivery_date': delivery_date})

@main_bp.route('/api/shipments/unmark_delivered', methods=['POST'])
def unmark_delivered():
    """撤销已交付标记并解锁"""
    from flask import request
    data = request.get_json()
    booking_id = data.get('booking_id', '').strip()
    if not booking_id:
        return jsonify({'error': 'Missing booking_id'}), 400
    db = get_db()
    # 检查运单是否存在
    row = db.execute("SELECT delivery_locked FROM shipments WHERE booking_id = ?", (booking_id,)).fetchone()
    if not row:
        return jsonify({'error': f'运单 {booking_id} 不存在'}), 404
    if not row['delivery_locked']:
        return jsonify({'error': f'运单 {booking_id} 未锁定，无需撤销'}), 409
    db.execute(
        "UPDATE shipments SET actual_delivery_date = '', delivery_locked = 0 WHERE booking_id = ?",
        (booking_id,)
    )
    db.commit()
    return jsonify({'success': True, 'booking_id': booking_id})


@main_bp.route('/api/shipments/mark_abnormal', methods=['POST'])
def mark_abnormal():
    """手动标记运单为异常单"""
    from flask import request
    data = request.get_json()
    booking_id = data.get('booking_id', '').strip()
    if not booking_id:
        return jsonify({'error': 'Missing booking_id'}), 400
    db = get_db()
    row = db.execute("SELECT delivery_locked, shipment_status FROM shipments WHERE booking_id = ?", (booking_id,)).fetchone()
    if not row:
        return jsonify({'error': f'运单 {booking_id} 不存在'}), 404
    if row['delivery_locked']:
        return jsonify({'error': f'运单 {booking_id} 已标记已交付，请先撤销'}), 409
    if row['shipment_status'] == 'Canceled':
        return jsonify({'error': f'已取消单无需标记异常'}), 409
    db.execute(
        "UPDATE shipments SET manual_status = 'abnormal' WHERE booking_id = ?",
        (booking_id,)
    )
    db.commit()
    return jsonify({'success': True, 'booking_id': booking_id})


@main_bp.route('/api/shipments/unmark_abnormal', methods=['POST'])
def unmark_abnormal():
    """取消异常标记"""
    from flask import request
    data = request.get_json()
    booking_id = data.get('booking_id', '').strip()
    if not booking_id:
        return jsonify({'error': 'Missing booking_id'}), 400
    db = get_db()
    db.execute(
        "UPDATE shipments SET manual_status = NULL WHERE booking_id = ?",
        (booking_id,)
    )
    db.commit()
    return jsonify({'success': True, 'booking_id': booking_id})


@main_bp.route('/api/shipments/demand_country', methods=['POST'])
def update_shipment_demand_country():
    """Manually override business demand country without changing consignee country."""
    data = request.get_json(silent=True) or {}
    booking_ids = data.get('booking_ids') or data.get('booking_id') or []
    if isinstance(booking_ids, str):
        booking_ids = [booking_ids]
    booking_ids = list(dict.fromkeys(str(x).strip() for x in booking_ids if str(x).strip()))
    demand_country = (data.get('demand_country') or '').strip().upper()
    password = data.get('password') or ''

    if not maintenance_password_ok(password, allow_admin=True):
        return jsonify({'error': '密码错误'}), 403
    if not booking_ids:
        return jsonify({'error': '请选择需要修改的运单'}), 400
    if not demand_country or not re.fullmatch(r'[A-Z]{2,3}', demand_country):
        return jsonify({'error': '请输入正确的需求国家代码，例如 DE / PL / FR'}), 400

    db = get_db()
    placeholders = ','.join(['?'] * len(booking_ids))
    cursor = db.execute(f"""
        UPDATE shipments
        SET demand_country = ?, demand_country_source = 'manual'
        WHERE booking_id IN ({placeholders})
          AND COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
    """, [demand_country] + booking_ids)
    db.commit()
    return jsonify({'success': True, 'updated': cursor.rowcount, 'demand_country': demand_country})


@main_bp.route('/api/shipments/recycle', methods=['POST'])
def recycle_shipments():
    """Move abnormal shipments into the recycle bin without deleting the records."""
    data = request.get_json(silent=True) or {}
    booking_ids = data.get('booking_ids') or []
    if isinstance(booking_ids, str):
        booking_ids = [booking_ids]
    booking_ids = list(dict.fromkeys(str(x).strip() for x in booking_ids if str(x).strip()))
    reason = (data.get('reason') or '').strip()
    password = data.get('password') or ''

    if not maintenance_password_ok(password, allow_admin=True):
        return jsonify({'error': '密码错误'}), 403
    if not booking_ids:
        return jsonify({'error': '请选择需要移入回收站的异常运单'}), 400
    if not reason:
        return jsonify({'error': '请填写移入回收站原因'}), 400

    db = get_db()
    placeholders = ','.join(['?'] * len(booking_ids))
    rows = db.execute(f"""
        SELECT *
        FROM shipments
        WHERE booking_id IN ({placeholders})
          AND COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
    """, booking_ids).fetchall()
    found_ids = {r['booking_id'] for r in rows}
    missing_ids = [bid for bid in booking_ids if bid not in found_ids]
    not_abnormal = [r['booking_id'] for r in rows if calc_display_status(r) != '异常单']
    if missing_ids or not_abnormal:
        return jsonify({
            'error': '只有当前显示为异常单的运单可以移入回收站',
            'missing': missing_ids,
            'not_abnormal': not_abnormal,
        }), 400

    recycled_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor = db.execute(f"""
        UPDATE shipments
        SET recycled_at = ?, recycle_reason = ?, recycle_operator = ?
        WHERE booking_id IN ({placeholders})
          AND recycled_at IS NULL
    """, [recycled_at, reason, 'Qiteng'] + booking_ids)
    db.commit()
    return jsonify({'success': True, 'recycled': cursor.rowcount, 'requested': len(booking_ids), 'recycled_at': recycled_at})


@main_bp.route('/api/shipments/recycle_bin')
def shipment_recycle_bin():
    """List recycled shipments for audit and recovery."""
    db = get_db()
    search = (request.args.get('search') or '').strip()
    params = []
    where = "recycled_at IS NOT NULL"
    if search:
        where += " AND (booking_id LIKE ? OR stt_number LIKE ? OR recycle_reason LIKE ? OR consignee_country_code LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])
    rows = db.execute(f"""
        SELECT *
        FROM shipments
        WHERE {where}
        ORDER BY recycled_at DESC, pickup_date DESC
        LIMIT 500
    """, params).fetchall()
    items = []
    for r in rows:
        d = dict(r)
        d['display_status'] = calc_display_status(r)
        items.append(d)
    return jsonify({'shipments': items, 'total': len(items)})


@main_bp.route('/api/shipments/recycle/restore', methods=['POST'])
def restore_recycled_shipments():
    """Restore shipments from the recycle bin."""
    data = request.get_json(silent=True) or {}
    booking_ids = data.get('booking_ids') or []
    if isinstance(booking_ids, str):
        booking_ids = [booking_ids]
    booking_ids = [str(x).strip() for x in booking_ids if str(x).strip()]
    if not maintenance_password_ok(data.get('password'), allow_admin=True):
        return jsonify({'error': '密码错误'}), 403
    if not booking_ids:
        return jsonify({'error': '请选择需要恢复的运单'}), 400

    db = get_db()
    placeholders = ','.join(['?'] * len(booking_ids))
    cur = db.execute(f"""
        UPDATE shipments
        SET recycled_at = NULL, recycle_reason = NULL, recycle_operator = NULL
        WHERE booking_id IN ({placeholders})
          AND recycled_at IS NOT NULL
    """, booking_ids)
    db.commit()
    return jsonify({'success': True, 'restored': cur.rowcount})


# ========== 批量自动核实 ==========

# 全局进度存储（单进程够用）
_verify_state = {
    'running': False,
    'progress': 0,
    'total': 0,
    'results': [],
    'started_at': None,
}


@main_bp.route('/api/shipments/verify', methods=['POST'])
def start_verify():
    """启动批量核实：抓取所有'待核实'且有 STT 号的运单"""
    import threading
    import json
    from datetime import datetime as dt, timedelta

    global _verify_state
    if _verify_state['running']:
        return jsonify({'error': '核实任务正在运行中', 'progress': _verify_state['progress'], 'total': _verify_state['total']}), 409

    db = get_db()
    rows = db.execute("""
        SELECT stt_number, booking_id, delivery_date, actual_delivery_date
        FROM shipments
        WHERE COALESCE(shipment_status, '') != 'Canceled'
          AND recycled_at IS NULL
          AND stt_number IS NOT NULL AND stt_number != ''
          AND delivery_locked = 0
          AND (manual_status IS NULL OR manual_status != 'abnormal')
    """).fetchall()

    # 过滤出需要核实的：无 actual_delivery_date + delivery_date 已过期或为空
    today_str = dt.now().strftime('%Y-%m-%d')
    to_verify = []
    seen_stts = set()
    for r in rows:
        actual_dd = r['actual_delivery_date'] or ''
        if actual_dd and actual_dd.strip():
            continue  # 已有日期的跳过
        dd = r['delivery_date'] or ''
        stt = (r['stt_number'] or '').strip()
        if not stt or stt in seen_stts:
            continue
        if dd and dd.strip() and dd[:10] <= today_str:
            to_verify.append(stt)
            seen_stts.add(stt)
        elif not dd or not dd.strip():
            to_verify.append(stt)
            seen_stts.add(stt)

    if not to_verify:
        return jsonify({'message': '没有需要核实的运单', 'total': 0})

    _verify_state = {
        'running': True,
        'progress': 0,
        'total': len(to_verify),
        'results': [],
        'started_at': dt.now().isoformat(),
    }

    def _run(db_path):
        global _verify_state
        from .scraper import verify_stt_list

        def progress_cb(current, total):
            _verify_state['progress'] = current

        try:
            results = verify_stt_list(to_verify, progress_callback=progress_cb)
        except Exception as e:
            _verify_state['running'] = False
            _verify_state['error'] = f'核实抓取失败: {str(e)}'
            return

        try:
            conn = sqlite3.connect(db_path, timeout=30)
            cursor = conn.cursor()
            delivered_count = 0
            in_transport_count = 0
            cancelled_count = 0
            unresolved_count = 0
            failed_count = 0
            for r in results:
                _verify_state['results'].append(r)
                stt = (r.get('stt_number') or '').strip()
                stt_key = stt.upper()
                status = r.get('status', '')
                if status == 'delivered':
                    dd = r.get('delivery_date')
                    if dd:
                        # 有日期：写入并锁定
                        cursor.execute(
                            """
                            UPDATE shipments
                            SET actual_delivery_date = ?,
                                delivery_locked = 1,
                                manual_status = NULL,
                                last_checked = CURRENT_TIMESTAMP
                            WHERE UPPER(TRIM(stt_number)) = ?
                              AND delivery_locked = 0
                              AND recycled_at IS NULL
                            """,
                            (dd, stt_key)
                        )
                        delivered_count += cursor.rowcount
                    # 无日期：不锁定，下次继续尝试抓
                elif status in ('in_transport', 'booked'):
                    cursor.execute(
                        """
                        UPDATE shipments
                        SET manual_status = ?, last_checked = CURRENT_TIMESTAMP
                        WHERE UPPER(TRIM(stt_number)) = ?
                          AND delivery_locked = 0
                          AND recycled_at IS NULL
                          AND (manual_status IS NULL OR manual_status NOT IN ('abnormal'))
                        """,
                        ('auto_booked' if status == 'booked' else 'auto_in_transport', stt_key)
                    )
                    in_transport_count += cursor.rowcount
                elif status == 'booked_cancelled':
                    # 自动标记为 Canceled
                    cursor.execute(
                        """
                        UPDATE shipments
                        SET shipment_status = 'Canceled', last_checked = CURRENT_TIMESTAMP
                        WHERE UPPER(TRIM(stt_number)) = ?
                          AND COALESCE(shipment_status, '') != 'Canceled'
                          AND recycled_at IS NULL
                        """,
                        (stt_key,)
                    )
                    cancelled_count += cursor.rowcount
                elif status in ('unknown', 'not_found', 'error'):
                    if status == 'error':
                        failed_count += 1
                    else:
                        unresolved_count += 1
            conn.commit()
            conn.close()
            _verify_state['running'] = False
            _verify_state['delivered_count'] = delivered_count
            _verify_state['in_transport_count'] = in_transport_count
            _verify_state['cancelled_count'] = cancelled_count
            _verify_state['unresolved_count'] = unresolved_count
            _verify_state['failed_count'] = failed_count
        except Exception as e:
            _verify_state['running'] = False
            _verify_state['error'] = f'数据库写入失败: {str(e)}'

    db_path = current_app.config['DATABASE']
    thread = threading.Thread(target=_run, args=(db_path,), daemon=True)
    thread.start()

    return jsonify({
        'message': f'开始核实 {len(to_verify)} 个运单',
        'total': len(to_verify),
    })


@main_bp.route('/api/shipments/verify/status')
def verify_status():
    """查询核实进度"""
    return jsonify(_verify_state)


@main_bp.route('/api/shipments/verify/reset', methods=['POST'])
def verify_reset():
    """重置核实状态（卡住时使用）"""
    global _verify_state
    _verify_state = {
        'running': False,
        'progress': 0,
        'total': 0,
        'results': [],
        'started_at': None,
    }
    return jsonify({'success': True, 'message': '核实状态已重置'})


# ========== Invoice API ==========
@main_bp.route('/api/invoices/upload', methods=['POST'])
def upload_invoice():
    """上传物流费用发票 PDF，解析并存入数据库"""
    from flask import request
    import os

    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': '仅支持 PDF 文件'}), 400
    if uploaded_file_size(file) > 20 * 1024 * 1024:
        return jsonify({'error': 'PDF 文件不能超过 20MB'}), 400

    invoice_type = request.form.get('type', 'auto')  # auto / logistics / warehouse
    if invoice_type == 'auto':
        name = (file.filename or '').lower()
        warehouse_keywords = ['warehouse', 'warehousing', 'storage', 'handling', 'operation', 'operacja', 'magazyn', '仓储', '仓库', '操作']
        invoice_type = 'warehouse' if any(k in name for k in warehouse_keywords) else 'logistics'
    if invoice_type not in ('logistics', 'warehouse'):
        invoice_type = 'logistics'

    # 保存文件
    upload_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    saved_filename = safe_upload_filename(file.filename)
    filepath = os.path.join(upload_dir, saved_filename)
    file.save(filepath)

    try:
        from .invoice_parser import parse_invoice_pdf
        result = parse_invoice_pdf(filepath)

        if not result['invoice_number']:
            remove_file_safely(filepath)
            return jsonify({'error': '无法识别发票号，请确认 PDF 格式'}), 400

        invoice_number = result['invoice_number']
        db_module.backup_sqlite_db(current_app.config['DATABASE'], f"invoice_upload_{invoice_number}")
        db = get_db()
        parsed_items = [
            item for item in result['items']
            if item['stt_number'] and re.match(r'^[A-Z]{2,5}\d{6,}$', item['stt_number'])
        ]

        # 检查是否已存在
        existing = db.execute('SELECT id, status, total_net FROM invoice_headers WHERE invoice_number = ?', (invoice_number,)).fetchone()
        preserved_acceptance = {}
        preserved_status = 'pending'
        if existing:
            # 已存在则更新
            old_items = db.execute('''SELECT stt_number, net_amount, ref_date, matched_booking_id,
                                             accepted_amount, acceptance_status, exception_reason, acceptance_remark
                                      FROM invoice_items
                                      WHERE invoice_id = ? ORDER BY stt_number, net_amount, ref_date''',
                                   (existing[0],)).fetchall()
            old_signature = [(r['stt_number'], round(r['net_amount'] or 0, 2), r['ref_date'] or '') for r in old_items]
            for r in old_items:
                preserved_acceptance[(r['stt_number'], r['ref_date'] or '')] = {
                    'matched_booking_id': r['matched_booking_id'],
                    'accepted_amount': r['accepted_amount'],
                    'acceptance_status': r['acceptance_status'],
                    'exception_reason': r['exception_reason'],
                    'acceptance_remark': r['acceptance_remark'],
                }
            new_signature = sorted((item['stt_number'], round(item['net_amount'] or 0, 2), item.get('ref_date') or '') for item in parsed_items)
            invoice_unchanged = (
                abs((existing['total_net'] or 0) - (result['total_net'] or 0)) < 0.01
                and old_signature == new_signature
            )
            db.execute('DELETE FROM invoice_items WHERE invoice_id = ?', (existing[0],))
            invoice_id = existing[0]
            preserved_status = existing['status'] if invoice_unchanged else 'pending'
            db.execute('''UPDATE invoice_headers SET invoice_date=?, invoice_type=?, currency=?, total_net=?,
                         total_vat=0, total_gross=0, status=?, source_file=? WHERE id=?''',
                       (result['invoice_date'], invoice_type, result['currency'],
                        result['total_net'], preserved_status, saved_filename, invoice_id))
        else:
            db.execute('''INSERT INTO invoice_headers (invoice_number, invoice_date, invoice_type, currency, total_net, status, source_file)
                         VALUES (?, ?, ?, ?, ?, 'pending', ?)''',
                       (invoice_number, result['invoice_date'], invoice_type, result['currency'],
                        result['total_net'], saved_filename))
            invoice_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        # 插入明细
        matched_count = 0
        for item in parsed_items:
            # 尝试匹配运单
            matched_booking = None
            preserved = preserved_acceptance.get((item['stt_number'], item.get('ref_date') or ''), {})
            row = db.execute('SELECT booking_id FROM shipments WHERE stt_number = ? OR booking_id = ?',
                           (item['stt_number'], item['stt_number'])).fetchone()
            if row:
                matched_booking = row[0]
                matched_count += 1
            if preserved.get('matched_booking_id'):
                matched_booking = preserved.get('matched_booking_id')

            db.execute('''INSERT INTO invoice_items (
                              invoice_id, stt_number, net_amount, ref_date, matched_booking_id,
                              accepted_amount, acceptance_status, exception_reason, acceptance_remark
                          )
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (
                           invoice_id, item['stt_number'], item['net_amount'], item['ref_date'], matched_booking,
                           preserved.get('accepted_amount'), preserved.get('acceptance_status') or preserved_status,
                           preserved.get('exception_reason'), preserved.get('acceptance_remark')
                       ))

        recalc_invoice_header_status(db, invoice_id)
        db.commit()
        db.close()
        remove_file_safely(filepath)

        return jsonify({
            'success': True,
            'invoice_number': invoice_number,
            'invoice_date': result['invoice_date'],
            'total_net': result['total_net'],
            'currency': result['currency'],
            'total_items': len(result['items']),
            'matched_items': matched_count,
            'type': invoice_type
        })
    except Exception as e:
        remove_file_safely(filepath)
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/invoices')
def list_invoices():
    """列出所有发票"""
    db = get_db()
    rows = db.execute('''SELECT h.id, h.invoice_number, h.invoice_date, h.invoice_type, h.currency,
                          h.total_net, h.status, h.source_file, h.created_at,
                          COUNT(i.id) as item_count,
                          SUM(CASE WHEN i.matched_booking_id IS NOT NULL THEN 1 ELSE 0 END) as matched_count
                          FROM invoice_headers h LEFT JOIN invoice_items i ON h.id = i.invoice_id
                          GROUP BY h.id ORDER BY h.created_at DESC''').fetchall()
    db.close()
    return jsonify([{
        'id': r['id'],
        'invoice_number': r['invoice_number'],
        'invoice_date': r['invoice_date'],
        'invoice_type': r['invoice_type'],
        'currency': r['currency'],
        'total_net': r['total_net'],
        'status': r['status'],
        'source_file': r['source_file'],
        'created_at': r['created_at'],
        'item_count': r['item_count'],
        'matched_count': r['matched_count']
    } for r in rows])


@main_bp.route('/api/invoices/<int:invoice_id>')
def get_invoice(invoice_id):
    """获取单个发票详情"""
    db = get_db()
    header = db.execute('SELECT * FROM invoice_headers WHERE id = ?', (invoice_id,)).fetchone()
    if not header:
        db.close()
        return jsonify({'error': '发票不存在'}), 404
    items = db.execute('''SELECT i.*, s.booking_id as shipment_booking_id, s.shipment_type
                         FROM invoice_items i LEFT JOIN shipments s ON i.matched_booking_id = s.booking_id
                         WHERE i.invoice_id = ? ORDER BY i.id''', (invoice_id,)).fetchall()
    db.close()
    return jsonify({
        'id': header['id'],
        'invoice_number': header['invoice_number'],
        'invoice_date': header['invoice_date'],
        'invoice_type': header['invoice_type'],
        'currency': header['currency'],
        'total_net': header['total_net'],
        'status': header['status'],
        'source_file': header['source_file'],
        'created_at': header['created_at'],
        'items': [{
            'id': item['id'],
            'stt_number': item['stt_number'],
            'net_amount': item['net_amount'],
            'ref_date': item['ref_date'],
            'matched_booking_id': item['matched_booking_id'],
            'shipment_type': item['shipment_type']
        } for item in items]
    })


@main_bp.route('/api/invoices/items')
def list_invoice_items():
    """费用明细聚合接口，供费用列表一次性加载，避免逐张发票请求。"""
    db = get_db()
    invoice_type = request.args.get('type', 'logistics')
    status = request.args.get('status', '')
    params = []
    where = []
    if invoice_type:
        where.append('h.invoice_type = ?')
        params.append(invoice_type)
    if status:
        where.append("COALESCE(NULLIF(i.acceptance_status, ''), h.status) = ?")
        params.append(status)
    where_sql = 'WHERE ' + ' AND '.join(where) if where else ''
    rows = db.execute(f"""
        SELECT i.id, i.stt_number, i.net_amount, i.ref_date, i.matched_booking_id,
               i.accepted_amount, i.acceptance_status, i.exception_reason, i.acceptance_remark,
               h.id AS invoice_id, h.invoice_number, h.invoice_date, h.invoice_type,
               h.currency, COALESCE(NULLIF(i.acceptance_status, ''), h.status) AS invoice_status,
               s.booking_id AS shipment_booking_id, s.shipment_type, s.consignee_country_code,
               s.pickup_date
        FROM invoice_items i
        JOIN invoice_headers h ON i.invoice_id = h.id
        LEFT JOIN shipments s ON i.matched_booking_id = s.booking_id
        {where_sql}
        ORDER BY h.invoice_date DESC, h.invoice_number, i.id
    """, params).fetchall()
    pln_to_usd = 0.25
    items = [dict(r) for r in rows]
    total_pln = round(sum(float(r.get('accepted_amount') if r.get('accepted_amount') is not None else r.get('net_amount') or 0) for r in items), 2)
    return jsonify({
        'items': items,
        'summary': {
            'item_count': len(items),
            'matched_count': sum(1 for r in items if r.get('matched_booking_id')),
            'pending_count': sum(1 for r in items if r.get('invoice_status') == 'pending'),
            'verified_count': sum(1 for r in items if r.get('invoice_status') == 'verified'),
            'rejected_count': sum(1 for r in items if r.get('invoice_status') == 'rejected'),
            'total_pln': total_pln,
            'total_usd': round(total_pln * pln_to_usd, 2),
            'pln_to_usd': pln_to_usd
        }
    })


@main_bp.route('/api/invoices/budget')
def get_budget():
    """获取费用验收预算概览"""
    db = get_db()
    # 总预算 25 万 USD
    budget_usd = 250000

    # 已验收物流费用（PLN）——只算已验收的发票
    logistics_total = db.execute('''SELECT COALESCE(SUM(COALESCE(i.accepted_amount, i.net_amount)), 0) FROM invoice_items i
                                    JOIN invoice_headers h ON i.invoice_id = h.id
                                    WHERE h.invoice_type = 'logistics' AND COALESCE(NULLIF(i.acceptance_status, ''), h.status) = 'verified' ''').fetchone()[0]
    pending_logistics_total = db.execute('''SELECT COALESCE(SUM(COALESCE(i.accepted_amount, i.net_amount)), 0) FROM invoice_items i
                                            JOIN invoice_headers h ON i.invoice_id = h.id
                                            WHERE h.invoice_type = 'logistics' AND COALESCE(NULLIF(i.acceptance_status, ''), h.status) = 'pending' ''').fetchone()[0]
    match_row = db.execute('''SELECT COUNT(i.id) AS item_count,
                                     SUM(CASE WHEN i.matched_booking_id IS NOT NULL THEN 1 ELSE 0 END) AS matched_count
                              FROM invoice_items i
                              JOIN invoice_headers h ON i.invoice_id = h.id
                              WHERE h.invoice_type = 'logistics' ''').fetchone()

    # PLN → USD 汇率（可配置）
    pln_to_usd = 0.25  # 约 4 PLN = 1 USD
    logistics_usd = round(logistics_total * pln_to_usd, 2)

    # 仓储操作费用（暂为0）
    warehouse_total = 0
    warehouse_usd = 0

    total_spent_usd = logistics_usd + warehouse_usd
    remaining_usd = budget_usd - total_spent_usd

    # 发票列表摘要（直接查询，不调用list_invoices避免db关闭问题）
    inv_rows = db.execute('''SELECT h.id, h.invoice_number, h.invoice_date, h.invoice_type, h.currency,
                          h.total_net, h.status, h.source_file, h.created_at,
                          COUNT(i.id) as item_count,
                          SUM(CASE WHEN i.matched_booking_id IS NOT NULL THEN 1 ELSE 0 END) as matched_count
                          FROM invoice_headers h LEFT JOIN invoice_items i ON h.id = i.invoice_id
                          GROUP BY h.id ORDER BY h.created_at DESC''').fetchall()
    invoices = [{
        'id': r['id'],
        'invoice_number': r['invoice_number'],
        'invoice_date': r['invoice_date'],
        'invoice_type': r['invoice_type'],
        'currency': r['currency'],
        'total_net': r['total_net'],
        'status': r['status'],
        'source_file': r['source_file'],
        'created_at': r['created_at'],
        'item_count': r['item_count'],
        'matched_count': r['matched_count']
    } for r in inv_rows]

    db.close()

    return jsonify({
        'budget_usd': budget_usd,
        'logistics_pln': logistics_total,
        'logistics_usd': logistics_usd,
        'pending_logistics_pln': pending_logistics_total,
        'pending_logistics_usd': round(pending_logistics_total * pln_to_usd, 2),
        'invoice_item_count': match_row['item_count'] or 0,
        'invoice_matched_count': match_row['matched_count'] or 0,
        'invoice_match_rate': round(((match_row['matched_count'] or 0) / (match_row['item_count'] or 1)) * 100, 1) if match_row['item_count'] else 0,
        'warehouse_pln': warehouse_total,
        'warehouse_usd': warehouse_usd,
        'total_spent_usd': total_spent_usd,
        'remaining_usd': remaining_usd,
        'pln_to_usd': pln_to_usd,
        'invoices': invoices
    })


@main_bp.route('/api/invoices/<int:invoice_id>/verify', methods=['POST'])
def verify_invoice(invoice_id):
    """验收/拒绝发票"""
    from flask import request
    data = request.get_json(silent=True) or {}
    status = data.get('status', 'verified')  # verified / rejected
    if status not in ('verified', 'rejected'):
        return jsonify({'error': '状态必须是 verified 或 rejected'}), 400
    db = get_db()
    db.execute('UPDATE invoice_headers SET status = ? WHERE id = ?', (status, invoice_id))
    db.execute("UPDATE invoice_items SET acceptance_status = ? WHERE invoice_id = ?", (status, invoice_id))
    db.commit()
    db.close()
    return jsonify({'success': True, 'status': status})


@main_bp.route('/api/invoices/<int:invoice_id>/match', methods=['POST'])
def match_invoice_items(invoice_id):
    """重新匹配发票明细与运单"""
    db = get_db()
    items = db.execute('SELECT id, stt_number FROM invoice_items WHERE invoice_id = ?', (invoice_id,)).fetchall()
    matched = 0
    for item in items:
        row = db.execute('SELECT booking_id FROM shipments WHERE stt_number = ? OR booking_id = ?',
                         (item['stt_number'], item['stt_number'])).fetchone()
        if row:
            db.execute('UPDATE invoice_items SET matched_booking_id = ? WHERE id = ?', (row['booking_id'], item['id']))
            matched += 1
        else:
            db.execute('UPDATE invoice_items SET matched_booking_id = NULL WHERE id = ?', (item['id'],))
    db.commit()
    db.close()
    return jsonify({'success': True, 'matched': matched, 'total': len(items)})

@main_bp.route('/api/shipments/import', methods=['POST'])
def import_shipments():
    """导入 Excel 文件，type 参数指定来源：warehouse / pl_domestic / intl_direct"""
    from flask import request
    import os
    import pandas as pd
    from datetime import datetime

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only Excel files supported'}), 400
    if uploaded_file_size(file) > 20 * 1024 * 1024:
        return jsonify({'error': 'Excel 文件不能超过 20MB'}), 400

    upload_type = request.form.get('type', '')  # warehouse / pl_domestic / intl_direct

    def _clean_nan(val):
        """处理 pandas 的 NaN 值，转为空字符串"""
        if val is None or (isinstance(val, float) and str(val) == 'nan'):
            return ''
        s = str(val).strip()
        return '' if s == 'nan' or s == '' else s

    def _parse_excel_date(val):
        if pd.isna(val) or str(val).strip() == '':
            return None
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d %H:%M')
        s = str(val).strip()
        for fmt in ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M')
            except ValueError:
                continue
        return s

    upload_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    saved_filename = safe_upload_filename(file.filename)
    filepath = os.path.join(upload_dir, saved_filename)
    file.save(filepath)

    try:
        db_module.backup_sqlite_db(current_app.config['DATABASE'], f"shipments_import_{upload_type or 'unknown'}")
        df = pd.read_excel(filepath, header=18)
        conn = sqlite3.connect(current_app.config['DATABASE'])
        cursor = conn.cursor()
        imported = 0
        inserted = 0
        updated = 0
        skipped_date = 0
        skipped_missing_stt = 0

        # 只保留2026年1月1日及之后的运单（根据pickup_date）
        MIN_DATE = '2026-01-01'

        for _, row in df.iterrows():
            # 优先使用 Booking ID，如果没有则用 STT Number 作为唯一标识
            booking_id = _clean_nan(row.get('Booking ID', ''))
            stt_number = _clean_nan(row.get('STT Number', ''))
            if not stt_number:
                skipped_missing_stt += 1
                continue
            if not booking_id or booking_id == 'nan':
                booking_id = stt_number
            if not booking_id or booking_id == 'nan':
                continue

            # 只保留2026年1月1日及之后的运单
            pickup_date = _parse_excel_date(row.get('Pickup Date'))
            if pickup_date:
                pd_date = pickup_date[:10] if len(pickup_date) >= 10 else pickup_date
                if pd_date < MIN_DATE:
                    skipped_date += 1
                    continue

            # 根据 type 参数和文件内容判定 shipment_type
            if upload_type in ('pl_domestic', 'intl_direct'):
                shipment_type = 'direct'
            elif upload_type.startswith('warehouse'):
                # 仓库文件根据 pickup city 区分
                pickup_city = str(row.get('Pickup City', '')).strip()
                pickup_city_lower = pickup_city.lower()
                if 'bydgoszcz' in pickup_city_lower:
                    shipment_type = 'warehouse_furniture'
                else:
                    shipment_type = 'warehouse_central'
            else:
                # 兼容旧逻辑：自动判断
                user_group = str(row.get('User group name', ''))
                pickup_city_lower = str(row.get('Pickup City', '')).strip().lower()
                if 'PLdom' in user_group:
                    shipment_type = 'direct'
                elif 'PLint' in user_group:
                    if 'bydgoszcz' in pickup_city_lower:
                        shipment_type = 'warehouse_furniture'
                    elif 'lodz' in pickup_city_lower or 'łódź' in pickup_city_lower:
                        shipment_type = 'warehouse_central'
                    else:
                        shipment_type = 'direct'
                else:
                    shipment_type = 'unknown'

            # 仓库字段
            pickup_city = str(row.get('Pickup City', '')).strip()
            pickup_city_lower = pickup_city.lower()
            warehouse = ''
            if 'bydgoszcz' in pickup_city_lower:
                warehouse = 'Bydgoszcz'
            elif 'lodz' in pickup_city_lower or 'łódź' in pickup_city_lower:
                warehouse = 'Łódź'

            def parse_date(val):
                return _parse_excel_date(val)

            price = row.get('Price without VAT', 0)
            if pd.isna(price): price = 0
            try: price = float(str(price).replace(',', '').replace(' ', ''))
            except: price = 0

            pieces = row.get('Total Pieces', 0)
            weight = row.get('Total Weight', 0)
            volume = row.get('Total Volume', 0)
            if pd.isna(pieces): pieces = 0
            if pd.isna(weight): weight = 0
            if pd.isna(volume): volume = 0

            consignee_country_code = _clean_nan(row.get('Consignee Country Code')).strip().upper()
            demand_country = consignee_country_code
            demand_country_source = 'default' if demand_country else ''

            shipment_values = (
                booking_id,
                stt_number,
                _clean_nan(row.get('Consignment/Waybill No.')),
                _clean_nan(row.get('Shipment Status')),
                _clean_nan(row.get('Transport Mode')),
                shipment_type, warehouse,
                _clean_nan(row.get('Shipper Name', '')),
                _clean_nan(row.get('Consignee Name')),
                _clean_nan(row.get('Consignee Name2', '')),
                _clean_nan(row.get('Consignee Street', '')),
                _clean_nan(row.get('Consignee Street2', '')),
                _clean_nan(row.get('Consignee Address')) or _clean_nan(row.get('Delivery Address', '')),
                _clean_nan(row.get('Consignee City')),
                consignee_country_code,
                demand_country,
                demand_country_source,
                _clean_nan(row.get('Pickup City')),
                _clean_nan(row.get('Pickup Country Code', '')),
                _clean_nan(row.get('Delivery City', '')),
                _clean_nan(row.get('Delivery Country Code', '')),
                parse_date(row.get('Creation Date', '')),
                parse_date(row.get('Pickup Date')),
                parse_date(row.get('Delivery Date')),
                float(pieces), float(weight), float(volume),
                price,
                _clean_nan(row.get('Incoterm', '')),
                _clean_nan(row.get('Service Type', '')),
                _clean_nan(row.get('Product', '')),
                _clean_nan(row.get('Cargo Description')),
                _clean_nan(row.get('References', '')),
                upload_type + '_' + saved_filename if upload_type else 'upload_' + saved_filename
            )

            # 重复导入规则：
            # - 新 Booking ID：新增。
            # - 已存在 Booking ID/STT：只刷新 Excel 基础字段。
            # - 保护人工/自动核实字段：actual_delivery_date、delivery_locked、manual_status、last_checked。
            existing = cursor.execute(
                """SELECT id, demand_country_source FROM shipments
                   WHERE booking_id = ?
                      OR (stt_number IS NOT NULL AND stt_number != '' AND stt_number = ?)
                   ORDER BY CASE WHEN booking_id = ? THEN 0 ELSE 1 END
                   LIMIT 1""",
                (booking_id, stt_number, booking_id)
            ).fetchone()

            if existing:
                cursor.execute("""
                    UPDATE shipments SET
                        booking_id=?, stt_number=?, waybill_no=?, shipment_status=?,
                        transport_mode=?, shipment_type=?, warehouse=?,
                        shipper_name=?, consignee_name=?, consignee_name2=?,
                        consignee_street=?, consignee_street2=?, consignee_address=?,
                        consignee_city=?, consignee_country_code=?,
                        demand_country=CASE
                            WHEN COALESCE(demand_country_source, '') = 'manual' THEN demand_country
                            ELSE ?
                        END,
                        demand_country_source=CASE
                            WHEN COALESCE(demand_country_source, '') = 'manual' THEN demand_country_source
                            ELSE ?
                        END,
                        pickup_city=?, pickup_country_code=?,
                        delivery_city=?, delivery_country_code=?,
                        creation_date=?, pickup_date=?, delivery_date=?,
                        total_pieces=?, total_weight=?, total_volume=?,
                        price_without_vat=?, incoterm=?, service_type=?,
                        product=?, cargo_description=?, references_text=?, source_file=?,
                        import_time=CURRENT_TIMESTAMP
                    WHERE id=?
                """, shipment_values + (existing[0],))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO shipments (
                        booking_id, stt_number, waybill_no, shipment_status,
                        transport_mode, shipment_type, warehouse,
                        shipper_name, consignee_name, consignee_name2,
                        consignee_street, consignee_street2, consignee_address,
                        consignee_city, consignee_country_code,
                        demand_country, demand_country_source,
                        pickup_city, pickup_country_code,
                        delivery_city, delivery_country_code,
                        creation_date, pickup_date, delivery_date,
                        total_pieces, total_weight, total_volume,
                        price_without_vat, incoterm, service_type,
                        product, cargo_description, references_text, source_file
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, shipment_values)
                inserted += 1
            imported += 1

        conn.commit()
        conn.close()
        remove_file_safely(filepath)
        return jsonify({'success': True, 'imported': imported, 'inserted': inserted, 'updated': updated, 'skipped_date': skipped_date, 'skipped_missing_stt': skipped_missing_stt, 'type': upload_type})
    except Exception as e:
        remove_file_safely(filepath)
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/shipments/import-auto', methods=['POST'])
def import_shipments_auto():
    """一口上传多个物流 Excel，并按 Pickup City 自动归为中央仓、家具仓、调拨。"""
    import pandas as pd
    import unicodedata
    from datetime import datetime

    files = [f for f in request.files.getlist('files') if f and f.filename]
    single = request.files.get('file')
    if single and single.filename:
        files.append(single)
    if not files:
        return jsonify({'error': '请上传物流 Excel 文件'}), 400
    bad = [f.filename for f in files if not f.filename.lower().endswith(('.xlsx', '.xls'))]
    if bad:
        return jsonify({'error': '只支持 .xlsx / .xls 文件：' + ', '.join(bad[:5])}), 400
    if any(uploaded_file_size(f) > 20 * 1024 * 1024 for f in files):
        return jsonify({'error': '单个物流 Excel 文件不能超过 20MB'}), 400

    def _clean_nan(val):
        if val is None or (isinstance(val, float) and str(val) == 'nan'):
            return ''
        s = str(val).strip()
        return '' if s == 'nan' or s == '' else s

    def _parse_excel_date(val):
        if pd.isna(val) or str(val).strip() == '':
            return None
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d %H:%M')
        s = str(val).strip()
        for fmt in ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M')
            except ValueError:
                continue
        return s

    def _city_key(value):
        raw = str(value or '').replace('Ł', 'L').replace('ł', 'l')
        text = unicodedata.normalize('NFKD', raw).encode('ascii', 'ignore').decode('ascii')
        return re.sub(r'[^a-z0-9]+', '', text.lower())

    def classify_by_pickup_city(value):
        key = _city_key(value)
        if 'bydgoszcz' in key:
            return 'warehouse_furniture', 'Bydgoszcz'
        if 'lodz' in key:
            return 'warehouse_central', 'Łódź'
        return 'direct', ''

    def add_counts(target, source):
        for key, value in source.items():
            if isinstance(value, dict):
                target.setdefault(key, {})
                add_counts(target[key], value)
            else:
                try:
                    target[key] = target.get(key, 0) + int(value or 0)
                except (TypeError, ValueError):
                    pass

    upload_dir = os.path.join(os.path.dirname(current_app.config['DATABASE']), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    db_module.backup_sqlite_db(current_app.config['DATABASE'], 'shipments_import_auto')
    conn = sqlite3.connect(current_app.config['DATABASE'])
    cursor = conn.cursor()
    min_date = '2026-01-01'
    summary = {
        'imported': 0, 'inserted': 0, 'updated': 0,
        'skipped_date': 0, 'skipped_missing_stt': 0,
        'type_counts': {'warehouse_central': 0, 'warehouse_furniture': 0, 'direct': 0},
    }
    file_reports = []
    saved_paths = []

    try:
        for file in files:
            saved_filename = safe_upload_filename(file.filename)
            filepath = os.path.join(upload_dir, saved_filename)
            file.save(filepath)
            saved_paths.append(filepath)
            report = {
                'filename': file.filename,
                'source_file': 'logistics_auto_' + saved_filename,
                'imported': 0, 'inserted': 0, 'updated': 0,
                'skipped_date': 0, 'skipped_missing_stt': 0,
                'type_counts': {'warehouse_central': 0, 'warehouse_furniture': 0, 'direct': 0},
            }
            try:
                df = pd.read_excel(filepath, header=18)
                for _, row in df.iterrows():
                    booking_id = _clean_nan(row.get('Booking ID', ''))
                    stt_number = _clean_nan(row.get('STT Number', ''))
                    if not stt_number:
                        report['skipped_missing_stt'] += 1
                        continue
                    if not booking_id or booking_id == 'nan':
                        booking_id = stt_number
                    if not booking_id or booking_id == 'nan':
                        continue

                    pickup_date = _parse_excel_date(row.get('Pickup Date'))
                    if pickup_date:
                        pd_date = pickup_date[:10] if len(pickup_date) >= 10 else pickup_date
                        if pd_date < min_date:
                            report['skipped_date'] += 1
                            continue

                    pickup_city = _clean_nan(row.get('Pickup City'))
                    shipment_type, warehouse = classify_by_pickup_city(pickup_city)
                    report['type_counts'][shipment_type] = report['type_counts'].get(shipment_type, 0) + 1

                    price = row.get('Price without VAT', 0)
                    if pd.isna(price):
                        price = 0
                    try:
                        price = float(str(price).replace(',', '').replace(' ', ''))
                    except Exception:
                        price = 0
                    pieces = row.get('Total Pieces', 0)
                    weight = row.get('Total Weight', 0)
                    volume = row.get('Total Volume', 0)
                    if pd.isna(pieces):
                        pieces = 0
                    if pd.isna(weight):
                        weight = 0
                    if pd.isna(volume):
                        volume = 0

                    consignee_country_code = _clean_nan(row.get('Consignee Country Code')).strip().upper()
                    demand_country = consignee_country_code
                    demand_country_source = 'default' if demand_country else ''
                    source_file = 'logistics_auto_' + saved_filename
                    shipment_values = (
                        booking_id,
                        stt_number,
                        _clean_nan(row.get('Consignment/Waybill No.')),
                        _clean_nan(row.get('Shipment Status')) or 'Active',
                        _clean_nan(row.get('Transport Mode')),
                        shipment_type, warehouse,
                        _clean_nan(row.get('Shipper Name', '')),
                        _clean_nan(row.get('Consignee Name')),
                        _clean_nan(row.get('Consignee Name2', '')),
                        _clean_nan(row.get('Consignee Street', '')),
                        _clean_nan(row.get('Consignee Street2', '')),
                        _clean_nan(row.get('Consignee Address')) or _clean_nan(row.get('Delivery Address', '')),
                        _clean_nan(row.get('Consignee City')),
                        consignee_country_code,
                        demand_country,
                        demand_country_source,
                        pickup_city,
                        _clean_nan(row.get('Pickup Country Code', '')),
                        _clean_nan(row.get('Delivery City', '')),
                        _clean_nan(row.get('Delivery Country Code', '')),
                        _parse_excel_date(row.get('Creation Date', '')),
                        pickup_date,
                        _parse_excel_date(row.get('Delivery Date')),
                        float(pieces), float(weight), float(volume),
                        price,
                        _clean_nan(row.get('Incoterm', '')),
                        _clean_nan(row.get('Service Type', '')),
                        _clean_nan(row.get('Product', '')),
                        _clean_nan(row.get('Cargo Description')),
                        _clean_nan(row.get('References', '')),
                        source_file
                    )
                    existing = cursor.execute(
                        """SELECT id, demand_country_source FROM shipments
                           WHERE booking_id = ?
                              OR (stt_number IS NOT NULL AND stt_number != '' AND stt_number = ?)
                           ORDER BY CASE WHEN booking_id = ? THEN 0 ELSE 1 END
                           LIMIT 1""",
                        (booking_id, stt_number, booking_id)
                    ).fetchone()
                    if existing:
                        cursor.execute("""
                            UPDATE shipments SET
                                booking_id=?, stt_number=?, waybill_no=?, shipment_status=?,
                                transport_mode=?, shipment_type=?, warehouse=?,
                                shipper_name=?, consignee_name=?, consignee_name2=?,
                                consignee_street=?, consignee_street2=?, consignee_address=?,
                                consignee_city=?, consignee_country_code=?,
                                demand_country=CASE
                                    WHEN COALESCE(demand_country_source, '') = 'manual' THEN demand_country
                                    ELSE ?
                                END,
                                demand_country_source=CASE
                                    WHEN COALESCE(demand_country_source, '') = 'manual' THEN demand_country_source
                                    ELSE ?
                                END,
                                pickup_city=?, pickup_country_code=?,
                                delivery_city=?, delivery_country_code=?,
                                creation_date=?, pickup_date=?, delivery_date=?,
                                total_pieces=?, total_weight=?, total_volume=?,
                                price_without_vat=?, incoterm=?, service_type=?,
                                product=?, cargo_description=?, references_text=?, source_file=?,
                                import_time=CURRENT_TIMESTAMP
                            WHERE id=?
                        """, shipment_values + (existing[0],))
                        report['updated'] += 1
                    else:
                        cursor.execute("""
                            INSERT INTO shipments (
                                booking_id, stt_number, waybill_no, shipment_status,
                                transport_mode, shipment_type, warehouse,
                                shipper_name, consignee_name, consignee_name2,
                                consignee_street, consignee_street2, consignee_address,
                                consignee_city, consignee_country_code,
                                demand_country, demand_country_source,
                                pickup_city, pickup_country_code,
                                delivery_city, delivery_country_code,
                                creation_date, pickup_date, delivery_date,
                                total_pieces, total_weight, total_volume,
                                price_without_vat, incoterm, service_type,
                                product, cargo_description, references_text, source_file
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, shipment_values)
                        report['inserted'] += 1
                    report['imported'] += 1
                add_counts(summary, report)
                file_reports.append(report)
            except Exception as file_error:
                report['error'] = str(file_error)
                file_reports.append(report)
        conn.commit()
        return jsonify({'success': True, 'file_count': len(files), 'summary': summary, 'files': file_reports})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()
        for path in saved_paths:
            remove_file_safely(path)
